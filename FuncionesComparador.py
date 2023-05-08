"""
En este archivo se encuentran todas las funciones que el Software del Comparador de Bloques TESA requiere para funcionar
"""

################## Importación de librerías ##################
import RPi.GPIO as GPIO                                             # Biblioteca para el control de los motores a pasos y el servomotor
from RpiMotorLib import RpiMotorLib                                 # Biblioteca para motores a pasos
from RpiMotorLib import rpiservolib                                 # Biblioteca para servomotor
from time import sleep                                              # Biblioteca para sleep
import time
import serial                                                       # Biblioteca para configuración y adquisición de datos de dispositivos seriales
import openpyxl                                                     # Biblioteca para hojas de datos
import smtplib, ssl
import pandas as pd                                                 # Biblioteca para manejo de datos
import numpy as np                                                  # Biblioteca para trabajar con arreglos, facilita operaciones matemáticas
import openpyxl                                                     # Biblioteca para el manejo de archivos de excel
from openpyxl import load_workbook                                  # Biblioteca para cargar excel ya existente
import openpyxl.utils.cell                                          # Biblioteca para insertar columnas o filas en un excel 
from openpyxl.styles import Font, Color, Alignment, Border, Side    # Biblioteca para darle formato a archivos de excel
import shutil                                                       # Biblioteca para copiar archivos
import datetime                                                     # Biblioteca para obtener información de la fecha y hora del día
from decimal import Decimal                                         # Biblioteca para trabajar correctamente operaciones aritméticas con flotantes decimales

################################################################

global motorEnabledState
global motorDisabledState 
motorEnabledState = GPIO.HIGH
motorDisabledState = GPIO.LOW

################## Configuración de entradas/salidas ##################

GPIO_pins1 = (22, 27, 17)           #pines de modo para el motor1
direction1 = 9                      #pin de dirección para el motor1
step1 = 11                          #pin de step para el motor1

GPIO_pins2 = (5, 6, 13)             #pines de modo para el motor2
direction2 = 20                     #pin de dirección para el motor2
step2 = 21                          #pin de step para el motor2

pin_enableCalibrationMotor = 24                         #pin de enable

GPIO_pins3 = (14, 15, 18)           #Pines de modo de paso
direction3 = 19                     #Pin de sentido de giro
step3 = 16                          #Pin de dar paso
pin_enablePlateMotor = 23

sleepMot3=12                        #Pin para controlar el sleep del motor de ordenamiento
                                    #Si está en 1 está activo, en 0 está en sleep
pin_startRotationLimitSensor = 4               #Pin para el sensor infrarrojo de rotacion de angulo nicial
pin_endRotationLimitSensor = 3                 #Pin para el sensor infrarrojo de rotacion de angulo final

steperMotorPlate = RpiMotorLib.A4988Nema(direction3, step3, GPIO_pins3, "A4988") #Parámetros del motor

steperMotor1 = RpiMotorLib.A4988Nema(direction1, step1, GPIO_pins1, "A4988") #Parámetros del motor1
steperMotor2 = RpiMotorLib.A4988Nema(direction2, step2, GPIO_pins2, "A4988") #Parámetros del motor2

GPIO.setup(pin_enableCalibrationMotor, GPIO.OUT)     
                                                                                                                                           
GPIO.output(pin_enableCalibrationMotor, motorDisabledState)       #Modo seguro, motores inhabilitados

GPIO.setup(pin_enablePlateMotor, GPIO.OUT)     
                                                                                                                                           
GPIO.output(pin_enablePlateMotor, motorDisabledState)       #Modo seguro, motores de plato inhabilitados

GPIO.setup(sleepMot3, GPIO.OUT)                                                                                                                                                
GPIO.output(sleepMot3, GPIO.LOW)       #Sleep debe estar en LOW para deshabilitarse


GPIO.setmode(GPIO.BCM)              #Numeración Broadcom
GPIO.setup(pin_startRotationLimitSensor, GPIO.IN)    #Se define como entrada el sensor

posicionStep=0                      #Variable de posición angular del disco
required=0                          #Variable de pasos requeridos par llegar
                                    #a la posicion deseada
listo=0                             #Variable que determina cuando terminó

#gohome()                            #gire el disco hasta home porque se inició el programa



################## Captura de datos TESA ##################

serTESA=serial.Serial("/dev/ttyUSBI", baudrate=1200, bytesize=serial.SEVENBITS, parity=serial.PARITY_EVEN,
                          stopbits=serial.STOPBITS_TWO, xonxoff=True, timeout=0.5) #Configuración de puerto

def DatosTESA():                                   
    
    detenerse=0                     #Constante para while que captura dato
    def recv(serial):               #Definición de una función para recibir datos
        while True:
            
            data=serial.read(30)    #Lectura de 30 bytes
            if data == "":
                continue
            else:
                break
            sleep(0.02)
        return data
    while detenerse == 0:
        data=recv(serTESA)          #Llamada de la función
        print(data)
        if data != b"":             #Comparación de datos recibidos, vacío hasta que se de la medición
            try:
                medicion=float(data) #Pasando de string a float
                MedicionBloque=medicion #Guardando dato en lista
            
            except:
                divisionDatos=data.split()
                print(divisionDatos)
                medicion=float(divisionDatos[1])    #Pasando de string a decimal
                MedicionBloque=medicion #Guardando dato en lista
            detenerse = 1           #Condición para salir del while
    return MedicionBloque

################## Captura de datos Fluke ##################

serFluke=serial.Serial("/dev/ttyUSBK", baudrate=9600, bytesize=serial.EIGHTBITS, parity=serial.PARITY_NONE,
                       stopbits=serial.STOPBITS_ONE, xonxoff=True, timeout=0.5) #Configuración de puerto


def DatosFluke():
    
    serFluke.write(b"READ?1\r\n") #Envío de instrucción para capturar dato de temperatura 1
    serFluke.write(b"READ?2\r\n") #Envío de instrucción para capturar dato de temperatura 2
    serFluke.write(b"READ?3\r\n") #Envío de instrucción para capturar dato de temperatura 3
    serFluke.write(b"READ?4\r\n") #Envío de instrucción para capturar dato de temperatura 4

    detenerse=0 #Constante para while que captura dato
    
    MedicionTemp1=0 #Creación de variable para almacenar mediciones de temperatura 1
    MedicionTemp2=0 #Creación de variable para almacenar mediciones de temperatura 2
    MedicionTemp3=0 #Creación de variable para almacenar mediciones de temperatura 3
    MedicionTemp4=0 #Creación de variable para almacenar mediciones de temperatura 4


    def recv(serial): #Definición de una función para recibir datos
        while True:
            data=serial.read(32) #Lectura de 32 bytes
            if data == "":
                continue
            else:
                break
            sleep(0.02)
        return data
    while detenerse == 0:
        data=recv(serFluke) #Llamada de la función
        print(data)
        if data != b"": #Comparación de datos recibidos, vacío hasta que se de la medición
            todas=data.split()#Separar los 4 datos en una lista
            MedicionTemp1=float(todas[0]) #Guardando temperatura 1 en lista
            MedicionTemp2=float(todas[1]) #Guardando temperatura 2 en lista
            MedicionTemp3=float(todas[2]) #Guardando temperatura 3 en lista
            MedicionTemp4=float(todas[3]) #Guardando temperatura 4 en lista
            detenerse = 1  #Condición para salir del while
    return MedicionTemp1, MedicionTemp2, MedicionTemp3, MedicionTemp4

################## Captura de datos Vaisala ##################

serVaisala=serial.Serial("/dev/ttyUSBD", baudrate=4800, bytesize=serial.SEVENBITS,
                             parity=serial.PARITY_EVEN, stopbits=serial.STOPBITS_ONE, timeout= 0.5) #Configuración de puerto

def DatosVaisala():
    
    #serVaisala=serial.Serial("/dev/ttyAMA0", baudrate=4800, bytesize=serial.SEVENBITS, parity=serial.PARITY_EVEN, stopbits=serial.STOPBITS_ONE, timeout= 0.5)
    #serVaisala=serial.Serial("/dev/ttyUSB0", baudrate=19200, bytesize=serial.EIGHTBITS, parity=serial.PARITY_NONE, stopbits=serial.STOPBITS_ONE, timeout= 0.5)

    #serVaisala.write(b"RUN\r\n")
    #serVaisala.write(b"R\r/\n")
    #serVaisala.write(b"form "P=" 4.2 P " " U6 \t "T=" t " " U3 \t "RH=" 4.2 rh " " U5 \r\n")
    #serVaisala.write(b"form 4.2 rh " " \r\n")
    #serVaisala.write(b"form 4.2 P " " \t 4.2 t " " \t 4.2 rh " " \r\n")

    
    serVaisala.write(b"SEND\r\n") #Envío de instrucción para capturar datos del Vaisala
    
    detenerse=0 #Constante para while que captura dato
    
    #DatoPresVaisala=0 #Creación de variable para almacenar mediciones de presión atmosférica
    #DatoTempVaisala=0 #Creación de variable para almacenar mediciones de temperatura ambiente
    DatoHumeVaisala=0 #Creación de variable para almacenar mediciones de humedad relativa
    
    def recv(serial): #Definición de una función para recibir datos
        while True:
            data=serial.read(30)
            if data == "":
                continue
            else:
                break
            sleep(0.02)
        return data
    while detenerse == 0:
        data=recv(serVaisala)                   #Llamada de la función
        print(data)
        if data != b"":                         #Comparación de datos recibidos, vacío hasta que se de la medición
            todos=data.split()                  #Separar los 4 datos en una lista
            print(todos)
            #DatoPresVaisala=float(todos[1])     #Guardando presión atmosférica en lista
            #DatoTempVaisala=float(todos[2])     #Guardando temperatura en lista
            #DatoHumeVaisala=float(todos[1])     #Guardando humedad relativa en lista
            DatoHumeVaisala=todos    #Guardando humedad relativa en lista
            detenerse = 1                       #Condición para salir del while
    return DatoHumeVaisala

################## Servo Motor ##################

servo_pin = 26 #Pin que envía la señal al servomotor

def ActivaPedal(servo_pin): 

    myservotest = rpiservolib.SG90servo("servoone", 50, 2, 12) #Parámetros del servomotor

    myservotest.servo_move(servo_pin, 2.3, .5, False, .01)     #Movimiento a posición 2.3
    myservotest.servo_move(servo_pin, 7.5, .5, False, .01)     #Movimiento a posición 7.5

################## Secuencia desviación de longitud central ##################

def Centros(tiempoinicial, tiempoestabilizacion, Repeticiones):
	# Tiempo inicial entra en minutos 
	# Tiempo de estabilización entra en segundos
	
    GPIO.output(pin_enableCalibrationMotor, motorEnabledState)       #habilita los motores
    
    global valorNominalBloque
    global dato
    
    #obtenerAnguloBloque(valorNominalBloque[dato])          #Moverse a la siguiente pareja de bloques
    global t1
    t1=time.time()                                   #finaliza el conteo de espera de bloques
    tic=time.perf_counter()                                 #Toma el tiempo inicial
    
    listaMediciones=[]
    
    #sleep(int(tiempoinicial)*60)					#Tiempo de estabilización inicial
    #Antes de empezar a medir es neceario que el palpador vuelva a subir un momento sobre el patrón
    ActivaPedal(servo_pin)								#Sube el palpador
    sleep(10)					#Se le da un tiempo al palpador arriba sobre el bloque patrón
    ActivaPedal(servo_pin)								#Baja el palpador
    
    for i in range(int(Repeticiones)):
		
		#Medición del bloque patrón (inicia con el palpador abajo)
        sleep(int(tiempoestabilizacion))						#Se le da un tiempo al palpador abajo en el bloque patrón
        ActivaPedal(servo_pin)								#Sube el palpador
        MedicionBloque=DatosTESA()                   		#Justo después de levantar el palpador TESA toma la medición
        print(MedicionBloque)
        listaMediciones.append(MedicionBloque)              	#Valor del patrón en posición 1 (centro patrón)
        
        #Movimiento de posición 1 a 2 con el palpador arriba
        steperMotor1.motor_go(True, "Half", 407, .005, False, 2)	
        
        #Medición del calibrando
        ActivaPedal(servo_pin)                              #Baja el palpador
        sleep(int(tiempoestabilizacion))                    #Se le da un tiempo al palpador arriba sobre el calibrando
        ActivaPedal(servo_pin)                              #Sube el palpador
        MedicionBloque=DatosTESA()                   		#Justo después de levantar el palpador TESA toma la medición
        listaMediciones.append(MedicionBloque)               	#Valor del calibrando en posición 2 (centro calibrando)
        print(MedicionBloque)
        
        #Movimiento de 2 a 1 con el palpador arriba
        steperMotor1.motor_go(False, "Half", 410, .005, False, 2)
        ActivaPedal(servo_pin)								#Baja el palpador (termina cada repetición con el palpador abajo)

    #Una vez finalizadas las mediciones de los bloques el palpador se mueve a HOME
    sleep(int(tiempoestabilizacion))
    ActivaPedal(servo_pin)                                  #Sube palpador
	#Movimiento de 1 a HOME con el palpador arriba
    steperMotor1.motor_go(True, "Half", 203, .005, False, 2)#Movimiento de punto1 a HOME
    ActivaPedal(servo_pin)                                  #Baja palpador
    
    #obtenerAnguloBloque(valorNominalBloques[dato])          #Moverse a la siguiente pareja de bloques
    toc=time.perf_counter()                                 #Toma el tiempo final
    global tiempoCorrida
    tiempoCorrida=toc-tic                                   #retorna el tiempo de corrida en segundos
    global t0
    t0=time.time()                                   #inicia el conteo de espera de bloques
    GPIO.output(pin_enableCalibrationMotor, motorDisabledState)      #Inhabilita los motores
    return listaMediciones, tiempoCorrida, t0

################## Secuencia desviación de longitud central + planitud (Plantilla 1) ##################

def Completa1(tiempoinicial, tiempoestabilizacion, Repeticiones):
    GPIO.output(pin_enableCalibrationMotor, motorEnabledState)       #habilita los motores
    
    global valorNominalBloque
    global dato
    
    #obtenerAnguloBloque(valorNominalBloque[dato])          #Moverse a la siguiente pareja de bloques
    global t1
    t1=time.time()                                   #finaliza el conteo de espera de bloques
    tic=time.perf_counter()                                 #Toma el tiempo inicial
    
    listaMediciones=[]
    
    sleep(int(tiempoinicial)*60)
    
    for i in range(int(Repeticiones)):

        ActivaPedal(servo_pin) #Baja palpador
        sleep(int(tiempoestabilizacion)) #Tiempo de estabilización
        ActivaPedal(servo_pin) #Sube palpador
        MedicionBloque=DatosTESA() #Llama función TESA
        listaMediciones.append(MedicionBloque) #Valor del patrón en posición 1 (centro)
        print(MedicionBloque)

        steperMotor1.motor_go(True, "Half", 417, .005, False, 2) #Mov de 1 a 2

        ActivaPedal(servo_pin) #Baja palpador
        sleep(int(tiempoestabilizacion)) #Tiempo de estabilización
        ActivaPedal(servo_pin) #Sube palpador
        MedicionBloque=DatosTESA() #Llama función TESA
        listaMediciones.append(MedicionBloque)  #Valor del calibrando en posición 2 (esquina)
        print(MedicionBloque)

        steperMotor1.motor_go(False, "Half", 96, .005, False, 2) #Mov1 de 2 a 3
        steperMotor2.motor_go(False, "Full", 398, .005, False, 1) #Mov2 de 2 a 3
        steperMotor1.motor_go(True, "Half", 178, .005, False, 1) #Mov3 de 2 a 3

        ActivaPedal(servo_pin) #Baja palpador
        sleep(int(tiempoestabilizacion)) #Tiempo de estabilización
        ActivaPedal(servo_pin) #Sube palpador
        MedicionBloque=DatosTESA() #Llama función TESA
        listaMediciones.append(MedicionBloque) #Valor del calibrando en posición 3 (esquina)
        print(MedicionBloque)
            
        steperMotor1.motor_go(False, "Half", 178, .005, False, 2) #Mov de 3 a 4

        ActivaPedal(servo_pin) #Baja palpador
        sleep(int(tiempoestabilizacion)) #Tiempo de estabilización
        ActivaPedal(servo_pin) #Sube palpador
        MedicionBloque=DatosTESA() #Llama función TESA
        listaMediciones.append(MedicionBloque)  #Valor del calibrando en posición 4 (esquina)
        print(MedicionBloque)

        steperMotor2.motor_go(True, "Full", 796, .005, False, 2) #Mov de 4 a 5

        ActivaPedal(servo_pin) #Baja palpador
        sleep(int(tiempoestabilizacion)) #Tiempo de estabilización
        ActivaPedal(servo_pin) #Sube palpador
        MedicionBloque=DatosTESA() #Llama función TESA
        listaMediciones.append(MedicionBloque)  #Valor del calibrando en posición 5 (esquina)
        print(MedicionBloque)

        steperMotor1.motor_go(True, "Half", 174, .005, False, 2) #Mov de 5 a 6

        ActivaPedal(servo_pin) #Baja palpador
        sleep(int(tiempoestabilizacion)) #Tiempo de estabilización
        ActivaPedal(servo_pin) #Sube palpador
        MedicionBloque=DatosTESA() #Llama función TESA
        listaMediciones.append(MedicionBloque)  #Valor del calibrando en posición 6 (esquina)
        print(MedicionBloque)

        steperMotor1.motor_go(False, "Half", 174, .005, False, 2) #Mov de 6 a 5
        steperMotor2.motor_go(False, "Full", 398, .005, False, 1) #Mov de 5 a Esp2
        steperMotor1.motor_go(False, "Half", 330, .005, False, 1) #Mov de Esp2 a 1

        ActivaPedal(servo_pin) #Baja palpador
        sleep(int(tiempoestabilizacion)) #Tiempo de estabilización
        ActivaPedal(servo_pin) #Sube palpador
        MedicionBloque=DatosTESA() #Llama función TESA
        listaMediciones.append(MedicionBloque)  #Valor del patrón en posición 1 (centro)
        print(MedicionBloque)

    steperMotor1.motor_go(True, "Half", 208, .005, False, 2) #Mov de 1 a HOME

    ActivaPedal(servo_pin) #Baja palpador
    
    listaMediciones.append(MedicionBloque)
    #obtenerAnguloBloque(valorNominalBloques[dato])          #Moverse a la siguiente pareja de bloques
    toc=time.perf_counter()                                 #Toma el tiempo final
    global tiempoCorrida
    tiempoCorrida=toc-tic                            #retorna el tiempo de corrida en segundos
    global t0
    t0=time.time()                                   #inicia el conteo de espera de bloques
    GPIO.output(pin_enableCalibrationMotor, motorDisabledState)       #Inhabilita los motores
    return listaMediciones, tiempoCorrida, t0

################## Secuencia desviación de longitud central + planitud (Plantilla 2) ##################

def Completa2(tiempoinicial, tiempoestabilizacion, Repeticiones):
    GPIO.output(pin_enableCalibrationMotor, motorEnabledState)       #habilita los motores
    
    global valorNominalBloque
    global dato
    
    #obtenerAnguloBloque(valorNominalBloque[dato])          #Moverse a la siguiente pareja de bloques
    
    global t1
    t1=time.time()                                   #finaliza el conteo de espera de bloques
    tic=time.perf_counter()                                 #Toma el tiempo inicial

    listaMediciones=[]
    
    sleep(int(tiempoinicial)*60)
    
    for i in range(int(Repeticiones)):

        ActivaPedal(servo_pin)                              #Baja palpador
        sleep(int(tiempoestabilizacion))                    #Tiempo de estabilización
        ActivaPedal(servo_pin)                              #Sube palpador
        MedicionBloque=DatosTESA()                   #Llama función TESA
        print(MedicionBloque)
        listaMediciones.append(MedicionBloque)              #Valor del patrón en posición 1 (centro)

        steperMotor1.motor_go(True, "Half", 416, .005, False, 2) #Mov de 1 a 2

        ActivaPedal(servo_pin)                              #Baja palpador
        sleep(int(tiempoestabilizacion))                    #Tiempo de estabilización
        ActivaPedal(servo_pin)                              #Sube palpador
        MedicionBloque=DatosTESA()                   #Llama función TESA
        print(MedicionBloque)
        listaMediciones.append(MedicionBloque)              #Valor del calibrando en posición 2 (esquina)
        
        steperMotor1.motor_go(False, "Half", 96, .005, False, 2)    #Mov1 de 2 a 3
        steperMotor2.motor_go(False, "Full", 337, .005, False, 1)   #Mov2 de 2 a 3
        steperMotor1.motor_go(True, "Half", 182, .005, False, 1)    #Mov3 de 2 a 3

        ActivaPedal(servo_pin)                              #Baja palpador
        sleep(int(tiempoestabilizacion))                    #Tiempo de estabilización
        ActivaPedal(servo_pin)                              #Sube palpador
        MedicionBloque=DatosTESA()                   #Llama función TESA
        print(MedicionBloque)
        listaMediciones.append(MedicionBloque)              #Valor del calibrando en posición 3 (esquina)
        
        steperMotor1.motor_go(False, "Half", 183, .005, False, 2)
                                                            #Mov de 3 a 4

        ActivaPedal(servo_pin)                              #Baja palpador
        sleep(int(tiempoestabilizacion))                    #Tiempo de estabilización
        ActivaPedal(servo_pin)                              #Sube palpador
        MedicionBloque=DatosTESA()                   #Llama función TESA
        print(MedicionBloque)
        listaMediciones.append(MedicionBloque)              #Valor del calibrando en posición 4 (esquina)

        steperMotor2.motor_go(True, "Full", 683, .005, False, 2) #Mov de 4 a 5

        ActivaPedal(servo_pin)                              #Baja palpador
        sleep(int(tiempoestabilizacion))                    #Tiempo de estabilización
        ActivaPedal(servo_pin)                              #Sube palpador
        MedicionBloque=DatosTESA()                   #Llama función TESA
        print(MedicionBloque)
        listaMediciones.append(MedicionBloque)              #Valor del calibrando en posición 5 (esquina)

        steperMotor1.motor_go(True, "Half", 178, .005, False, 2) #Mov de 5 a 6

        ActivaPedal(servo_pin)                              #Baja palpador
        sleep(int(tiempoestabilizacion))                    #Tiempo de estabilización
        ActivaPedal(servo_pin)                              #Sube palpador
        MedicionBloque=DatosTESA()                   #Llama función TESA
        print(MedicionBloque)
        listaMediciones.append(MedicionBloque)              #Valor del calibrando en posición 6 (esquina)

        steperMotor1.motor_go(False, "Half", 178, .005, False, 2) #Mov de 6 a 5
        steperMotor2.motor_go(False, "Full", 342, .005, False, 1) #Mov de 5 a Esp2
        steperMotor1.motor_go(False, "Half", 332, .005, False, 1) #Mov de Esp2 a 1

        ActivaPedal(servo_pin)                              #Baja palpador
        sleep(int(tiempoestabilizacion))                    #Tiempo de estabilización
        ActivaPedal(servo_pin)                              #Sube palpador
        MedicionBloque=DatosTESA()                   #Llama función TESA
        print(MedicionBloque)
        listaMediciones.append(MedicionBloque)              #Valor del patrón en posición 1 (centro)

    steperMotor1.motor_go(True, "Half", 208, .005, False, 2) #Mov de 1 a HOME
        
    ActivaPedal(servo_pin)                              #Baja palpador
    listaMediciones.append(MedicionBloque)              #Valor del patrón en posición 1 (centro)
    
    #obtenerAnguloBloque(valorNominalBloques[dato])          #Moverse a la siguiente pareja de bloques
    toc=time.perf_counter()                                 #Toma el tiempo final
    global tiempoCorrida
    tiempoCorrida=toc-tic                                   #retorna el tiempo de corrida en segundos
    global t0
    t0=time.time()                                   #inicia el conteo de espera de bloques
    GPIO.output(pin_enableCalibrationMotor, motorDisabledState)       #Inhabilita los motores
    return listaMediciones, tiempoCorrida, t0

################## Búsqueda de Clientes ##################

def BusquedaClientes(nombreClienteBuscado):
    """
    Entrada: 
        nombreCliente: nombre del cliente para el cual se va a calibrar
    Salida: una lista con el nombre del cliente, su dirección y el archivo donde está almacenada su información
    """
    woorkbookClientes = load_workbook(filename="Clientes.xlsx", keep_vba=True) # Apertura del archivo de excel de clientes 
    hojaClientes = woorkbookClientes.active # Hoja del archivo de excel donde están los clientes y su información
    
    i = 3 # Se inicializa un contador que va a recorrer los números de fila, empezando por la fila 3 porque las filas anteriores son encabezados
    while hojaClientes["A"+str(i)].value != None: # Mientras la celda de cliente no esté vacía, se van a seguir recorriendo las filas
        if hojaClientes["A"+str(i)].value == nombreClienteBuscado: # Si el valor de la celda es igual al nombre del cliente que se busca:
            numFila = i # El número de fila donde se encuentra la información del cliente corresponde al valor actual del contador 
        i += 1

    # Se almacenan el variables la información del cliente:
    # La columna A corresponde a "Nombre del Cliente", B a "Dirección del Cliente" y C a "Nombre del archivo del cliente"
    # numFila permite crear la coordenada de la celda que almacena la información deseada
    nombreCliente = hojaClientes["A"+str(numFila)].value 
    direccionCliente = hojaClientes["B"+str(numFila)].value
    archivoCliente = hojaClientes["C"+str(numFila)].value

    return nombreCliente, direccionCliente, archivoCliente
    
################## Selector de machote ##################

def selectorMachote(seleccionSecuencia):
	if seleccionSecuencia == "Desviación central" :
		machote = "./Machotes/Machote para calibración de Bloques con comparador mecánico TESA (Desviación central).xlsm"
		return machote
	elif seleccionSecuencia == "Desviación central y planitud" :
		machote = "./Machotes/Machote para calibración de Bloques con comparador mecánico TESA (Desviación central y planitud).xlsm"
		return machote
		
################## Creación de un archivo para la calibración ##################

def CrearArchivoCalibracion(seleccionSecuencia):
	# Se escoge sobre qué machote se va a trabajar a partir de la secuencia de calibración escogida por el usuario:
	machote = selectorMachote(seleccionSecuencia)

	# Se crea un duplicado del machote, nombrado con una marca temporal:
	fecha = datetime.datetime.now() # Fecha y hora del día
	archivoCalibracion = "./Calibraciones en curso/Calibración"+str(fecha.strftime("%c"))+".xlsm" # Nombre del archivo para la calibración
	shutil.copy(machote, archivoCalibracion) # Creación del duplicado del machote

	return archivoCalibracion

################## Autocompletado de la información que se tiene del cliente y la calibración ##################

def AutocompletarInformacionCliente(nombreCliente, direccionCliente, numeroCertificado, numeroSolicitud, identificacionCalibrando, 
                                    responsableCalibracion, responsableRevision, patron, materialPatron, seleccionSecuencia):
    # Lista con la información del cliente para el que se va a calibrar
    informacionCliente = BusquedaClientes(nombreCliente)
   
    # Carga del archivo de excel que contiene la información del cliente 
    workbookSolicitantes = load_workbook(filename=informacionCliente[2], keep_vba=True)
    hojaJuego = workbookSolicitantes[identificacionCalibrando] # Selección de la hoja que contiene la información del juego a calibrar

    # Carga del archivo de excel creado para la calibración:
    workbookCalibracion = load_workbook(filename=CrearArchivoCalibracion(seleccionSecuencia), keep_vba=True)
    #Definir los nombres de las hojas del libro de excel en el que se está trabajando:
    hojaConversionDatos = workbookCalibracion["conversion datos"]
    hojaIdentificacionBloques = workbookCalibracion["Ident.Bloques a calibrar"]
    hojaResultadosCalibracion = workbookCalibracion["Introduccion de datos de Calib."]

    #Almacenamiento de la información del calibrando
    listaValores = [] #Se crea una lista para guardar los valores nominales de los bloques del juego seleccionado
    for filaValores in hojaJuego.iter_rows(min_row=12,
                                           min_col=2,
                                           max_col=2):
        for celdaValores in filaValores:
            listaValores.append(float(celdaValores.value))
        
    listaSeries = [] #Se crea una lista para guardar los valores de las series de los bloques del juego seleccionado
    for filaSeries in hojaJuego.iter_rows(min_row=12,
                                          min_col=3,
                                          max_col=3):
        for celdaSeries in filaSeries:
            listaSeries.append(int(celdaSeries.value))

    objetoCalibrar = hojaJuego["C2"].value #Extrae el objeto a calibrar del archivo del cliente
    marcaInstrumento = hojaJuego["C3"].value #Extrae la marca del instrumento del archivo del cliente
    serieInstrumento = hojaJuego["C4"].value #Extrae la serie del instrumento del archivo del cliente
    material = hojaJuego["C5"].value #Extrae el material del calibrando del cliente
    modelo = hojaJuego["C6"].value #Extrae el modelo del calibrando del cliente
    gradoDeclarado = hojaJuego["C7"].value #Extra el gradoDeclarado del calibrando del cliente

    #Copiar información de los valores nominales y series de los bloques del juego a la hoja de la calibración
    i = 12
    while i <= (len(listaValores)+11):
        #Copiar información de los valores nominales:
        coordenadaCeldaValor = "B"+str(i-8) #Coordenada de la celda donde se encuentra el valor nominal de cada bloque a lo largo del while
        hojaIdentificacionBloques[coordenadaCeldaValor] = listaValores[i-12] #Escritura del valor nominal del bloque en la hoja correspondiente a la calibración
        #Copiar información de las series:
        coordenadaCeldaSerie = "C"+str(i-8) #Coordenada de la celda donde se encuentra la serie de cada bloque a lo largo del while
        hojaIdentificacionBloques[coordenadaCeldaSerie] = listaSeries[i-12] #Escritura de la serie del bloque en la hoja correspondiente a la calibración

        i+=1

    #Autocompletado de la información del calibrando en la hoja de la calibración
    hojaConversionDatos["L6"] = numeroCertificado
    hojaConversionDatos["L9"] = nombreCliente
    hojaConversionDatos["L11"] = numeroSolicitud
    hojaConversionDatos["L12"] = direccionCliente

    hojaConversionDatos["L17"] = objetoCalibrar
    hojaConversionDatos["L19"] = marcaInstrumento
    hojaConversionDatos["L20"] = serieInstrumento
    hojaConversionDatos["L21"] = material 
    hojaConversionDatos["L22"] = modelo
    hojaConversionDatos["L23"] = gradoDeclarado

    hojaConversionDatos["L27"] = responsableCalibracion
    hojaConversionDatos["L28"] = responsableRevision

    hojaConversionDatos["L30"] = patron
    hojaConversionDatos["L31"] = materialPatron

    return workbookSolicitantes, hojaResultadosCalibracion

def EncabezadosDesviacionCentral(numRepeticiones, hojaResultadosCalibracion):
    #SI NO SE HA REANUDADO LA CALIBRACIÓN
    #Modificar la hoja que va a almacenar los resultados de la calibración siguiendo la secuencia COMPLETA
    #Manejar los resultados de la calibración obtenidos con la secuencia COMPLETA
    #Completa: 1 Patrón, 1 Calibrando por repetición -> 2 columnas por repetición

    hojaResultadosCalibracion["H2"] = numRepeticiones
    numNuevasColumnas = 2*numRepeticiones #Se usan dos columnas por cada repetición: una para el patrón y otra para el calibrando
    hojaResultadosCalibracion.insert_cols(idx=19, amount=numNuevasColumnas) # Insertar el número de columnas necesarias al final de las columnas llenas en la hoja

    #Definir estilos 
    texto_negrita = Font(bold = True)
    texto_centrado = Alignment(horizontal = "center", vertical="center", wrapText=True)
    borde_sencillo = Side(border_style = "thin")
    borde_cuadrado = Border(top = borde_sencillo,
                            right = borde_sencillo,
                            bottom = borde_sencillo,
                            left = borde_sencillo)

    j = 19 #Se inicializa el contador para las columnas
    k = 1 #Se inicializa el contador para las repeticiones

    while j <= (19+numNuevasColumnas)-1 and k <= numRepeticiones:
        letraColumnaPatron = openpyxl.utils.cell.get_column_letter(j) #Obtener la letra de la columna en la que se está trabajando
        letraColumnaCalibrando = openpyxl.utils.cell.get_column_letter(j+1) #Obtener la letra de la columna de la derecha a la que se está trabajando
        #Coordenadas de las celdas del patrón y calibrando #repetición
        coordenadaEncabezadoPatron = letraColumnaPatron + "1"
        coordenadaEncabezadoCalibrando = letraColumnaCalibrando + "1"
        #Escribir los encabezados de las nuevas celdas:
        hojaResultadosCalibracion[coordenadaEncabezadoPatron] = "Patron #"+str(k)
        hojaResultadosCalibracion[coordenadaEncabezadoCalibrando] = "Calibrando #"+str(k)
        #Darle formato a las nuevas celdas:
        hojaResultadosCalibracion[coordenadaEncabezadoPatron].font = texto_negrita
        hojaResultadosCalibracion[coordenadaEncabezadoPatron].alignment = texto_centrado
        hojaResultadosCalibracion[coordenadaEncabezadoPatron].border = borde_cuadrado
        hojaResultadosCalibracion[coordenadaEncabezadoCalibrando].font = texto_negrita
        hojaResultadosCalibracion[coordenadaEncabezadoCalibrando].alignment = texto_centrado
        hojaResultadosCalibracion[coordenadaEncabezadoCalibrando].border = borde_cuadrado
    
        k += 1
        j += 2 
    return numNuevasColumnas

def EncabezadosCentroYPlanitud(numRepeticiones, hojaResultadosCalibracion):
    """
    SI NO SE HA REANUDADO LA CALIBRACIÓN
    Modificar la hoja que va a almacenar los resultados de la calibración siguiendo la secuencia CENTROS
    Manejar los resultados de la calibración obtenidos con la secuencia CENTROS 
    Centros: 1 Patrón (centro), 5 Calibrando: centro y cuatro esquinas -> 6 columnas por repetición
    La secuencia Centros hace lo mismo que la secuencia Completa para el caso de los centros de los bloques, pero además 
    complementa esta calibración con mediciones de planitud para cada uno de los bloques
    """
    numRepeticiones = int(input("Indique el número de repeticiones para la calibración: ")) 
    hojaResultadosCalibracion["N2"] = numRepeticiones
    numNuevasColumnas = 6*numRepeticiones #Se usan dos columnas por cada repetición: una para el patrón y otra para el calibrando
    hojaResultadosCalibracion.insert_cols(idx=25, amount=numNuevasColumnas) # Insertar el número de columnas necesarias al final de las columnas llenas en la hoja

    #Definir estilos 
    texto_negrita = Font(bold = True)
    texto_centrado = Alignment(horizontal = "center", vertical="center", wrapText=True)
    borde_sencillo = Side(border_style = "thin")
    borde_cuadrado = Border(top = borde_sencillo,
                            right = borde_sencillo,
                            bottom = borde_sencillo,
                            left = borde_sencillo)

    j = 25 #Se inicializa el contador para las columnas
    k = 1 #Se inicializa el contador para las repeticiones

    while j <= (25+numNuevasColumnas)-1 and k <= numRepeticiones:
        letraColumnaPatron = openpyxl.utils.cell.get_column_letter(j) #Obtener la letra de la columna que va a guardar los datos del Patrón en cada rep k 
        letraColumnaCalibrandoCentro = openpyxl.utils.cell.get_column_letter(j+1) #Obtener la letra de la columna que va a guardar los datos del Centro del Calibrando en cada rep k 
        letraColumnaCalibrandoEsquina3 = openpyxl.utils.cell.get_column_letter(j+2) #Obtener la letra de la columna que va a guardar los datos de la Esquina 3 del Calibrando en cada rep k 
        letraColumnaCalibrandoEsquina4 = openpyxl.utils.cell.get_column_letter(j+3) #Obtener la letra de la columna que va a guardar los datos de la Esquina 4 del Calibrando en cada rep k 
        letraColumnaCalibrandoEsquina5 = openpyxl.utils.cell.get_column_letter(j+4) #Obtener la letra de la columna que va a guardar los datos de la Esquina 5 del Calibrando en cada rep k 
        letraColumnaCalibrandoEsquina6 = openpyxl.utils.cell.get_column_letter(j+5) #Obtener la letra de la columna que va a guardar los datos de la Esquina 3 del Calibrando en cada rep k 

        #Escribir los encabezados de las nuevas celdas:
        hojaResultadosCalibracion[letraColumnaPatron + "1"] = "Patron (Centro) #"+str(k)
        hojaResultadosCalibracion[letraColumnaCalibrandoCentro + "1"] = "Calibrando (Centro) #"+str(k)
        hojaResultadosCalibracion[letraColumnaCalibrandoEsquina3 + "1"] = "Calibrando (Esquina 3) #"+str(k)
        hojaResultadosCalibracion[letraColumnaCalibrandoEsquina4 + "1"] = "Calibrando (Esquina 4) #"+str(k)
        hojaResultadosCalibracion[letraColumnaCalibrandoEsquina5 + "1"] = "Calibrando (Esquina 5) #"+str(k)
        hojaResultadosCalibracion[letraColumnaCalibrandoEsquina6 + "1"] = "Calibrando (Esquina 6) #"+str(k)
    
        #Darle formato a las nuevas celdas:
        for numColumna in range(j,j+6):
            letraColumna = openpyxl.utils.cell.get_column_letter(numColumna)
            hojaResultadosCalibracion[letraColumna + "1"].font = texto_negrita
            hojaResultadosCalibracion[letraColumna + "1"].alignment = texto_centrado
            hojaResultadosCalibracion[letraColumna + "1"].border = borde_cuadrado
    
        k += 1
        j += 6
        return numNuevasColumnas

################## Cálculos del promedio y la desviación estándar ###################

def CalculosDesviacionCentral(numRepeticiones, numNuevasColumnas, hojaResultadosCalibracion):

    #Calcular el promedio de la diferencia entre el patrón y el calibrando 
    #Calcular la desviación estándar del promedio de la diferencia entre el patrón y el calibrando

    l = 2 #Se inicializa el contador para filas

    while hojaResultadosCalibracion["S"+str(l)].value != None:
        listaDiferencias = [] #Inicializamos una lista para guardar el resultado de calcular las diferencias entre el patrón y el calibrando
        j = 19 #Se vuelve a inicializar el contador para las columnas
        k = 1 #Se vuelve a inicializar el contador para las repeticiones
        while j <= (19+numNuevasColumnas)-1 and k <= numRepeticiones:
            letraColumnaPatron = openpyxl.utils.cell.get_column_letter(j) #Obtener la letra de la columna en la que se está trabajando
            letraColumnaCalibrando = openpyxl.utils.cell.get_column_letter(j+1) #Obtener la letra de la columna de la derecha a la que se está trabajando
            #Coordenadas de las celdas del patrón y calibrando #repetición: tomando en cuenta el cambio de fila
            coordenadaPatron = letraColumnaPatron + str(l)
            coordenadaCalibrando = letraColumnaCalibrando + str(l)
            #Se calcula la diferencia entre el calibrando y el patrón
            diferencia = np.subtract(hojaResultadosCalibracion[coordenadaCalibrando].value, hojaResultadosCalibracion[coordenadaPatron].value)
            #Se agrega el valor de diferencia y calibrando a sus respectivos arreglos:
            listaDiferencias.append(diferencia)
            k += 1
            j += 2 
        #Se convierten las listas en arreglos de numpy:
        arregloDiferencias = np.array(listaDiferencias)
        #Se calculan el promedio y la desviación estándar
        promedioDiferencias = np.average(arregloDiferencias)
        desviacionDiferencias = np.std(arregloDiferencias)
        #Se escriben estos valores en las celdas correspondientes en excel:
        hojaResultadosCalibracion["E"+str(l)] = promedioDiferencias
        hojaResultadosCalibracion["F"+str(l)] = desviacionDiferencias
        l += 1
    return

def CalculosDesviacionYPlanitud(hojaResultadosCalibracion, numNuevasColumnas, numRepeticiones):
    l = 2 #Se inicializa el contador para filas -> Empezamos a agregar valores en la fila 2

    while hojaResultadosCalibracion["Y"+str(l)].value != None:
        listaDiferenciasCentros = [] #Inicializamos una lista para guardar el resultado de calcular la diferencia entre el centro del calibrando y el patrón en cada repetición
        listaDiferenciasEsquina3 = [] #Inicializamos una lista para guardar el resultado de calcular la diferencia entre la esquina 3 del calibrando y el centro del patrón en cada repetición
        listaDiferenciasEsquina4 = [] #Inicializamos una lista para guardar el resultado de calcular la diferencia entre la esquina 4 del calibrando y el centro del patrón en cada repetición
        listaDiferenciasEsquina5 = [] #Incializamos una lista para guardar el resultado de calcular la diferencia entre la esquina 5 del calibrando y el centro del patrón en cada repetición
        listaDiferenciasEsquina6 = [] #Incializamos una lista para guardar el resultado de calcular la diferencia entre la esquina 6 del calibrando y el centro del patrón en cada repetición

        j = 25 #Se vuelve a inicializar el contador para las columnas
        k = 1 #Se vuelve a inicializar el contador para las repeticiones
        while j <= (25+numNuevasColumnas)-1 and k <= numRepeticiones:
            letraColumnaPatron = openpyxl.utils.cell.get_column_letter(j) #Obtener la letra de la columna en la que se encuentra la medicón del centro del patrón
            letraColumnaCalibrando = openpyxl.utils.cell.get_column_letter(j+1) #Obtener la letra de la columna en la que se encuentra el centro del calibrando
            letraColumnaEsquina3 = openpyxl.utils.cell.get_column_letter(j+2) #Obtener la letra de la columna en la que se encuentra la medición de la esquina 3 del calibrando
            letraColumnaEsquina4 = openpyxl.utils.cell.get_column_letter(j+3) #Obtener la letra de la columna en la que se encuentra la medición de la esquina 4 del calibrando
            letraColumnaEsquina5 = openpyxl.utils.cell.get_column_letter(j+4) #Obtener la letra de la columna en la que se encuentra la medición de la esquina 5 del calibrando
            letraColumnaEsquina6 = openpyxl.utils.cell.get_column_letter(j+5) #Obtener la letra de la columna en la que se encuentra la medición de la esquina 6 del calibrando
        
            #Coordenadas de las celdas del patrón y calibrando #repetición: tomando en cuenta el cambio de fila
            coordenadaPatron = letraColumnaPatron + str(l)
            coordenadaCalibrando = letraColumnaCalibrando + str(l)
            coordenadaEsquina3 = letraColumnaEsquina3 + str(l)
            coordenadaEsquina4 = letraColumnaEsquina4 + str(l)
            coordenadaEsquina5 = letraColumnaEsquina5 + str(l)
            coordenadaEsquina6 = letraColumnaEsquina6 + str(l)

            #Se calcula la diferencia entre el calibrando y el centro del patrón
            diferenciaCentros = np.subtract(hojaResultadosCalibracion[coordenadaCalibrando].value, hojaResultadosCalibracion[coordenadaPatron].value)

            #Se calculan las diferencias entre las esquinas del calibrando y el centro del patrón
            diferenciaEsquina3 = np.subtract(hojaResultadosCalibracion[coordenadaEsquina3].value, hojaResultadosCalibracion[coordenadaPatron].value)
            diferenciaEsquina4 = np.subtract(hojaResultadosCalibracion[coordenadaEsquina4].value, hojaResultadosCalibracion[coordenadaPatron].value)
            diferenciaEsquina5 = np.subtract(hojaResultadosCalibracion[coordenadaEsquina5].value, hojaResultadosCalibracion[coordenadaPatron].value)
            diferenciaEsquina6 = np.subtract(hojaResultadosCalibracion[coordenadaEsquina6].value, hojaResultadosCalibracion[coordenadaPatron].value)

            #Se agrega el valor de las diferencias a sus respectivas listas:
            listaDiferenciasCentros.append(diferenciaCentros)
            listaDiferenciasEsquina3.append(diferenciaEsquina3)
            listaDiferenciasEsquina4.append(diferenciaEsquina4)
            listaDiferenciasEsquina5.append(diferenciaEsquina5)
            listaDiferenciasEsquina6.append(diferenciaEsquina6)

            k += 1
            j += 6

        #Se convierten las listas en arreglos de numpy:
        arregloDiferenciasCentros = np.array(listaDiferenciasCentros)
        arregloDiferenciasEsquina3 = np.array(listaDiferenciasEsquina3)
        arregloDiferenciasEsquina4 = np.array(listaDiferenciasEsquina4)
        arregloDiferenciasEsquina5 = np.array(listaDiferenciasEsquina5)
        arregloDiferenciasEsquina6 = np.array(listaDiferenciasEsquina6)

        #Se calculan el promedio y la desviación estándar de la diferencia entre el patrón y el calibrando 
        promedioDiferencias = np.average(arregloDiferenciasCentros)
        desviacionDiferencias = np.std(arregloDiferenciasCentros)

        #Se calcula el promedio de las diferencias entre las esquinas del claibrando y el patrón
        promedioEsquina3 = np.average(arregloDiferenciasEsquina3)
        promedioEsquina4 = np.average(arregloDiferenciasEsquina4)
        promedioEsquina5 = np.average(arregloDiferenciasEsquina5)
        promedioEsquina6 = np.average(arregloDiferenciasEsquina6)

        #Se escriben estos valores en las celdas correspondientes en excel:
        hojaResultadosCalibracion["K"+str(l)] = promedioDiferencias
        hojaResultadosCalibracion["L"+str(l)] = desviacionDiferencias

        hojaResultadosCalibracion["E"+str(l)] = promedioEsquina3
        hojaResultadosCalibracion["F"+str(l)] = promedioEsquina4
        hojaResultadosCalibracion["G"+str(l)] = promedioEsquina5
        hojaResultadosCalibracion["H"+str(l)] = promedioEsquina6

        l += 1

    return

################## Bloque a calibrar ##################

def ContinuarCalibracion():
    respuesta = str(input("¿Desea continuar con la calibración?: "))
    if respuesta == "Sí":
        return True
    else:
        return False
    
################## Selector fila para la hoja de resultados ##################

def selectorFilaResultados(hojaResultadosCalibracion):
    i = 2 # Se inicializa el contador en 2 porque la fila 1 tiene los encabezados 
    for filaValorNominal in hojaResultadosCalibracion.iter_rows(min_row=2,
                                                                min_col=1,
                                                                max_col=1):
        for celdaValorNominal in filaValorNominal:
            if celdaValorNominal.value == None:
                numFila = i
            else:
                i += 1
    return numFila

################## Nueva Calibración ##################

def NuevaCalibracion(nombreCliente, numeroCertificado, numeroSolicitud, identificacionCalibrando, 
                                    responsableCalibracion, responsableRevision, patron, materialPatron, seleccionSecuencia, tiempoinicial, tiempoestabilizacion, numRepeticiones):
    
	nombreCliente, direccionCliente, archivoCliente = BusquedaClientes(nombreCliente)		#Búsqueda de los datos del cliente
	machote = selectorMachote(seleccionSecuencia)											#Selección de la plantilla del machote que se va a utilizar 
    
	#Creación de un duplicado del machote, nombrado con una marca temporal:
	fecha = datetime.datetime.now() 														#Fecha y hora en la que se comienza la calibración
	archivoCalibracion = "./Calibraciones en curso/Calibración"+str(fecha.strftime("%c"))+".xlsm" # Nombre del archivo para la calibración
	shutil.copy(machote, archivoCalibracion)												#Creación del duplicado del machote
    
	#Ingreso de interés del cliente y de la calibración al archivo de Excel
	archivoExcel = AutocompletarInformacionCliente(nombreCliente, direccionCliente, numeroCertificado, numeroSolicitud, identificacionCalibrando, 
                                    responsableCalibracion, responsableRevision, patron, materialPatron, seleccionSecuencia)
	libroExcel = archivoExcel[0]
	hojaResultadosCalibracion = archivoExcel[1]
    
	#Si se va calibrar los bloques solo con desviación central se hace lo siguiente:
	if seleccionSecuencia == "Desviación central":
		numNuevasColumnas = EncabezadosDesviacionCentral(numRepeticiones, hojaResultadosCalibracion)
		continuarCalibracion = "si"
		if continuarCalibracion == "si":
			valorBloque = Decimal(float(input("Indique el valor del bloque a Calibrar: ")))
			numFila = selectorFilaResultados(hojaResultadosCalibracion)
			hojaResultadosCalibracion["A"+str(numFila)] = valorBloque
			
			#Medición y registro de las condiciones ambientales iniciales
			listaMedicionesTemperatura = DatosFluke()
			hojaResultadosCalibracion["I"+str(numFila)] = listaMedicionesTemperatura[0]
			hojaResultadosCalibracion["J"+str(numFila)] = listaMedicionesTemperatura[1]
			hojaResultadosCalibracion["K"+str(numFila)] = listaMedicionesTemperatura[2]
			hojaResultadosCalibracion["L"+str(numFila)] = listaMedicionesTemperatura[3]
			#hojaResultadosCalibracion["M"+str(numFila)] = DatosVaisala()	#Dato de humedad relativa inicial
			
			numColumnaMediciones = 12 #Contador inicia en 19 porque ese es el número de la columna a partir del
			#cual se empiezan a registar las mediciones de los bloques
			
			#Se realizan las mediciones de los bloques y se guardan en una lista [patrón, calibrando, patrón, calibrando,...]
			listaMedicionesBloque = Centros(tiempoinicial, tiempoestabilizacion, numRepeticiones)[0]
			for numMedicion in range(len(listaMedicionesBloque)-1):
				letraColumnaMedicion = openpyxl.utils.cell.get_column_letter(numColumnaMediciones)
				hojaResultadosCalibracion[letraColumnaMedicion+str(numFila)] = listaMedicionesBloque[numMedicion]
				numColumnaMediciones += 1
			
			#Medición y registro de las condiciones ambientales finales
			listaMedicionesTemperatura = DatosFluke()
			hojaResultadosCalibracion["N"+str(numFila)] = listaMedicionesTemperatura[0]
			hojaResultadosCalibracion["O"+str(numFila)] = listaMedicionesTemperatura[1]
			hojaResultadosCalibracion["P"+str(numFila)] = listaMedicionesTemperatura[2]
			hojaResultadosCalibracion["Q"+str(numFila)] = listaMedicionesTemperatura[3]
			#hojaResultadosCalibracion["R"+str(numFila)] = DatosVaisala()	#Dato de humedad relativa final
			
			continuarCalibracion = input("¿Desea continuar con la calibración?: ") 
		else:
			print("Calibración finalizada")
				
		CalculosDesviacionCentral(numRepeticiones, numNuevasColumnas, hojaResultadosCalibracion)
		libroExcel.save("./Calibraciones Finalizadas/PruebaTerminada.xlsm")    
	else:
		print("No se ha programado esto aún")
    
	return

NuevaCalibracion("Instituto Costarricense de Electricidad", "LCM12345", "67890", "12345","Fernanda Quesada", "Leonardo Rojas", "Bloques Patrón de Cerámica de 0,5 mm a 100 mm", "Cerámica", "Desviación central", 0.25, 10, 2)
            
