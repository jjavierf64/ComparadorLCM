"""
En este archivo se encuentran todas las funciones que el Software del Comparador de Bloques TESA (interfaz simple) requiere para funcionar
"""

################## Importación de librerías ##################
import RPi.GPIO as GPIO                                             # Biblioteca para el control de los motores a pasos y el servomotor
from RpiMotorLib import RpiMotorLib                                 # Biblioteca para motores a pasos
from RpiMotorLib import rpiservolib                                 # Biblioteca para servomotor
from time import sleep                                              # Biblioteca para sleep
import time
import serial                                                       # Biblioteca para configuración y adquisición de datos de dispositivos seriales
import openpyxl                                                     # Biblioteca para el manejo de archivos de excel
from openpyxl import load_workbook                                  # Biblioteca para cargar excel ya existente
import openpyxl.utils.cell                                          # Biblioteca para insertar columnas o filas en un excel 
from openpyxl.styles import Font, Color, Alignment, Border, Side    # Biblioteca para darle formato a archivos de excel
from openpyxl import Workbook                                       # Biblioteca para crear nuevos archivos de excel
from copy import copy                                               
import shutil                                                       # Biblioteca para copiar archivos
from decimal import Decimal                                         # Biblioteca para trabajar correctamente operaciones aritméticas con flotantes decimales
import curses														# Biblioteca para interacción con el teclado
import os                                                           # Biblioteca para interactuar con el sistema operativo
import warnings
import csv                                                          # Biblioteca para crear archivos csv e interactuar con ellos
import tkinter as tk
from tkinter import ttk

################################################################

global motorEnabledState
global motorDisabledState 
motorEnabledState = GPIO.HIGH
motorDisabledState = GPIO.LOW

################## Configuración de entradas/salidas ##################

GPIO_pins1 = (22, 27, 17)           #pines de modo para el motor1
direction1 = 9                      #pin de dirección para el motor1
step1 = 11                          #pin de step para el motor1

GPIO_pins2 = (5, 6, 13)             #pines de modo para el motor2
direction2 = 20                     #pin de dirección para el motor2
step2 = 21                          #pin de step para el motor2

pin_enableCalibrationMotor = 24                         #pin de enable

GPIO_pins3 = (14, 15, 18)           #Pines de modo de paso
direction3 = 19                     #Pin de sentido de giro
step3 = 16                          #Pin de dar paso
pin_enablePlateMotor = 23

sleepMot3=12                        #Pin para controlar el sleep del motor de ordenamiento
                                    #Si está en 1 está activo, en 0 está en sleep
pin_startRotationLimitSensor = 4               #Pin para el sensor infrarrojo de rotacion de angulo nicial
pin_endRotationLimitSensor = 3                 #Pin para el sensor infrarrojo de rotacion de angulo final

steperMotorPlate = RpiMotorLib.A4988Nema(direction3, step3, GPIO_pins3, "A4988") #Parámetros del motor

steperMotor1 = RpiMotorLib.A4988Nema(direction1, step1, GPIO_pins1, "A4988") #Parámetros del motor1
steperMotor2 = RpiMotorLib.A4988Nema(direction2, step2, GPIO_pins2, "A4988") #Parámetros del motor2

GPIO.setup(pin_enableCalibrationMotor, GPIO.OUT)     
                                                                                                                                           
GPIO.output(pin_enableCalibrationMotor, motorDisabledState)       #Modo seguro, motores inhabilitados

GPIO.setup(pin_enablePlateMotor, GPIO.OUT)     
                                                                                                                                           
GPIO.output(pin_enablePlateMotor, motorDisabledState)       #Modo seguro, motores de plato inhabilitados

GPIO.setup(sleepMot3, GPIO.OUT)                                                                                                                                                
GPIO.output(sleepMot3, GPIO.LOW)       #Sleep debe estar en LOW para deshabilitarse


GPIO.setmode(GPIO.BCM)              #Numeración Broadcom
GPIO.setup(pin_startRotationLimitSensor, GPIO.IN)    #Se define como entrada el sensor

posicionStep=0                      #Variable de posición angular del disco
required=0                          #Variable de pasos requeridos par llegar
                                    #a la posicion deseada
listo=0                             #Variable que determina cuando terminó


#gohome()                            #gire el disco hasta home porque se inició el programa

################## Movimiento Manual de los motores ##################

def moverManualInterfaz(event):
    try:
        tecla = event.keysym
        print(tecla)
        
        
        if tecla == "Up":
            steperMotor1.motor_go(False, "Full", 20, 0, False, 0)		

        elif tecla == "Down":
            steperMotor1.motor_go(True, "Full", 20, 0, False, 0)

        elif tecla == "Left":
            steperMotor2.motor_go(False, "Full", 20, 0, False, 0)		

        elif tecla == "Right":
            steperMotor2.motor_go(True, "Full", 20, 0, False, 0)
            
    except:
        GPIO.output(pin_enableCalibrationMotor, motorDisabledState)   # Apagar motores
        


def moverManual():
    ActivaPedal(servo_pin)
    sleep(2)
    
    screen = curses.initscr()
    curses.noecho()
    curses.cbreak()
    screen.keypad(True)
    try:
        GPIO.output(pin_enableCalibrationMotor, motorEnabledState)   # Encender motores

        while True:
            char = screen.getch()   # Toma las teclas presionadas

            if char == ord("q") or char == ord("Q"):     # Tecla para salir del posicionamiento
                break

            elif char == curses.KEY_UP:
                print("UP")
                steperMotor1.motor_go(False, "Half", 2, 0, False, 0)		

            elif char == curses.KEY_DOWN:
                print("DOWN")
                steperMotor1.motor_go(True, "Half", 2, 0, False, 0)

            elif char == curses.KEY_LEFT:
                print("UP")
                steperMotor2.motor_go(False, "Half", 2, 0, False, 0)		

            elif char == curses.KEY_RIGHT:
                print("DOWN")
                steperMotor2.motor_go(True, "Half", 2, 0, False, 0)

    finally:
        curses.nocbreak(); screen.keypad(0); curses.echo()
        curses.endwin()
        GPIO.output(pin_enableCalibrationMotor, motorDisabledState)   # Apagar motores
        ActivaPedal(servo_pin)

################## Ventanas Pop-up para valores de entrada y mensajes ##################

def ventanaEntrada(mensaje):
    def on_enter_press(event):
        input_value = input_entry.get()
        top.destroy()
        return_value.set(input_value)

    top = tk.Toplevel()
    top.title("Ingresar valor")
    top.configure(bg="white")

    input_label = ttk.Label(top, text=mensaje)
    input_label.pack(pady=5)

    input_entry = ttk.Entry(top, width=30)
    input_entry.pack(pady=5)

    return_value = tk.StringVar()
    input_entry.bind("<Return>", on_enter_press)

    top.wait_variable(return_value)
    return return_value.get()

def ventanaOpciones(mensaje, opciones):
    def on_enter_press(event):
        selected_value = combo_var.get()
        top.destroy()
        return_value.set(selected_value)

    top = tk.Toplevel()
    top.title("Seleccionar una opción")
    top.configure(bg="white")

    message_label = ttk.Label(top, text=mensaje)
    message_label.pack(pady=5)

    combo_var = tk.StringVar()
    combo_box = ttk.Combobox(top, textvariable=combo_var, values=opciones, width=27)
    combo_box.pack(pady=5)

    return_value = tk.StringVar()
    combo_box.bind("<Return>", on_enter_press)

    top.wait_variable(return_value)
    return return_value.get()

def mostrarMensaje(mensaje):
    def on_enter_press(event):
        top.destroy()

    top = tk.Toplevel()
    top.title("Mensaje")
    top.configure(bg="white")

    message_label = ttk.Label(top, text=mensaje)
    message_label.pack(pady=10)

    top.bind("<Return>", on_enter_press)

################## Captura de datos TESA ##################

serTESA=serial.Serial("/dev/ttyUSBI", baudrate=1200, bytesize=serial.SEVENBITS, parity=serial.PARITY_EVEN,
                          stopbits=serial.STOPBITS_TWO, xonxoff=True, timeout=0.5) #Configuración de puerto

def DatosTESA():                                   
    
    detenerse=0                     #Constante para while que captura dato
    def recv(serial):               #Definición de una función para recibir datos
        while True:
            
            data=serial.read(30)    #Lectura de 30 bytes
            if data == "":
                continue
            else:
                break
            sleep(0.02)
        return data
    while detenerse == 0:
        data=recv(serTESA)          #Llamada de la función
        if data != b"":             #Comparación de datos recibidos, vacío hasta que se de la medición
            try:
                medicion=float(data) #Pasando de string a float
                MedicionBloque=medicion #Guardando dato en lista
            
            except:
                divisionDatos=data.split()
                medicion=float(divisionDatos[1])    #Pasando de string a decimal
                MedicionBloque=medicion #Guardando dato en lista
            detenerse = 1           #Condición para salir del while
    return MedicionBloque

################## Captura de datos Fluke ##################

serFluke=serial.Serial("/dev/ttyUSBK", baudrate=9600, bytesize=serial.EIGHTBITS, parity=serial.PARITY_NONE,
                       stopbits=serial.STOPBITS_ONE, xonxoff=True, timeout=0.5) #Configuración de puerto


def DatosFluke():
    
    serFluke.write(b"READ?1\r\n") #Envío de instrucción para capturar dato de temperatura 1
    serFluke.write(b"READ?2\r\n") #Envío de instrucción para capturar dato de temperatura 2
    serFluke.write(b"READ?3\r\n") #Envío de instrucción para capturar dato de temperatura 3
    serFluke.write(b"READ?4\r\n") #Envío de instrucción para capturar dato de temperatura 4

    detenerse=0 #Constante para while que captura dato
    
    MedicionTemp1=0 #Creación de variable para almacenar mediciones de temperatura 1
    MedicionTemp2=0 #Creación de variable para almacenar mediciones de temperatura 2
    MedicionTemp3=0 #Creación de variable para almacenar mediciones de temperatura 3
    MedicionTemp4=0 #Creación de variable para almacenar mediciones de temperatura 4


    def recv(serial): #Definición de una función para recibir datos
        while True:
            data=serial.read(32) #Lectura de 32 bytes
            if data == "":
                continue
            else:
                break
            sleep(0.02)
        return data
    while detenerse == 0:
        data=recv(serFluke) #Llamada de la función

        if data != b"": #Comparación de datos recibidos, vacío hasta que se de la medición
            todas=data.split()#Separar los 4 datos en una lista
            MedicionTemp1=float(todas[0]) #Guardando temperatura 1 en lista
            MedicionTemp2=float(todas[1]) #Guardando temperatura 2 en lista
            MedicionTemp3=float(todas[2]) #Guardando temperatura 3 en lista
            MedicionTemp4=float(todas[3]) #Guardando temperatura 4 en lista
            detenerse = 1  #Condición para salir del while
    return MedicionTemp1, MedicionTemp2, MedicionTemp3, MedicionTemp4

################## Captura de datos Vaisala ##################

serVaisala=serial.Serial("/dev/ttyUSBD", baudrate=4800, bytesize=serial.SEVENBITS,
                             parity=serial.PARITY_EVEN, stopbits=serial.STOPBITS_ONE, timeout= 0.5) #Configuración de puerto

def DatosVaisala():
    
    #serVaisala=serial.Serial("/dev/ttyAMA0", baudrate=4800, bytesize=serial.SEVENBITS, parity=serial.PARITY_EVEN, stopbits=serial.STOPBITS_ONE, timeout= 0.5)
    #serVaisala=serial.Serial("/dev/ttyUSB0", baudrate=19200, bytesize=serial.EIGHTBITS, parity=serial.PARITY_NONE, stopbits=serial.STOPBITS_ONE, timeout= 0.5)

    #serVaisala.write(b"RUN\r\n")
    #serVaisala.write(b"R\r/\n")
    #serVaisala.write(b"form "P=" 4.2 P " " U6 \t "T=" t " " U3 \t "RH=" 4.2 rh " " U5 \r\n")
    #serVaisala.write(b"form 4.2 rh " " \r\n")
    #serVaisala.write(b"form 4.2 P " " \t 4.2 t " " \t 4.2 rh " " \r\n")

    
    serVaisala.write(b'FORM 4.2 " P=" P " " U6 3.2 "T=" T " " U3 3.2 "RH=" RH " " U4\r\n') # Formato para la toma de datos
    serVaisala.write(b"SEND\r\n") #Envío de instrucción para capturar datos del Vaisala

    
    detenerse=0 #Constante para while que captura dato
    
    DatoPresVaisala=0 #Creación de variable para almacenar mediciones de presión atmosférica
    DatoTempVaisala=0 #Creación de variable para almacenar mediciones de temperatura ambiente
    DatoHumeVaisala=0 #Creación de variable para almacenar mediciones de humedad relativa
    
    def recv(serial): #Definición de una función para recibir datos
        while True:
            data=serial.read(85)
            if data == "":
                continue
            else:
                break
            sleep(0.02)
        return data
    
    
    while detenerse == 0:
        data=recv(serVaisala)                   #Llamada de la función
        if data.split()[0] == b'OK': 						#Comparación de datos recibidos, vacío hasta que se de la medición
            todos=data.split()					#Separar los 4 datos en una lista
            
            DatoPresVaisala=float(todos[3]) #Guardando presión atmosférica en lista
            DatoTempVaisala=float(todos[6]) #Guardando temperatura en lista
            DatoHumeVaisala=float(todos[9]) #Guardando humedad relativa 3 en lista
            
            detenerse = 1                    #Condición para salir del while
            
            
    return DatoHumeVaisala

################## Servo Motor ##################

servo_pin = 26 #Pin que envía la señal al servomotor

def ActivaPedal(servo_pin=26): 

    myservotest = rpiservolib.SG90servo("servoone", 50, 2, 12) #Parámetros del servomotor

    myservotest.servo_move(servo_pin, 2.3, .5, False, .01)     #Movimiento a posición 2.3
    myservotest.servo_move(servo_pin, 7.5, .5, False, .01)     #Movimiento a posición 7.5

################## Secuencia desviación de longitud central ##################

def Centros(tiempoestabilizacion, Repeticiones):
	# Tiempo de estabilización entra en segundos
	
    global valorNominalBloque
    global dato
    
    #obtenerAnguloBloque(valorNominalBloque[dato])          #Moverse a la siguiente pareja de bloques
    global t1
    t1=time.time()                                   #finaliza el conteo de espera de bloques
    tic=time.perf_counter()                                 #Toma el tiempo inicial
    
    listaMediciones=[]
    
    #Antes de empezar a medir es necesario que el palpador vuelva a subir un momento sobre el patrón
    """
    ActivaPedal(servo_pin)								#Sube el palpador
    sleep(10)					#Se le da un tiempo al palpador arriba sobre el bloque patrón
    ActivaPedal(servo_pin)								#Baja el palpador  
    """
    for i in range(int(Repeticiones)):
		
		#Medición del bloque patrón (inicia con el palpador abajo)
        sleep(int(tiempoestabilizacion))						#Se le da un tiempo al palpador abajo en el bloque patrón
        ActivaPedal(servo_pin)								#Sube el palpador
        MedicionBloque=DatosTESA()                   		#Justo después de levantar el palpador TESA toma la medición
        #print(MedicionBloque)
        listaMediciones.append(MedicionBloque)              	#Valor del patrón en posición 1 (centro patrón)
        
        #Movimiento de posición 1 a 2 con el palpador arriba
        GPIO.output(pin_enableCalibrationMotor, motorEnabledState)       #habilita los motores
        steperMotor1.motor_go(True, "1/8", 835, .0025, False, 2)
        GPIO.output(pin_enableCalibrationMotor, motorDisabledState)       #Modo seguro, motores inhabilitados	
        
        #Medición del calibrando
        ActivaPedal(servo_pin)                              #Baja el palpador
        sleep(int(tiempoestabilizacion))                    #Se le da un tiempo al palpador arriba sobre el calibrando
        ActivaPedal(servo_pin)                              #Sube el palpador
        MedicionBloque=DatosTESA()                   		#Justo después de levantar el palpador TESA toma la medición
        listaMediciones.append(MedicionBloque)               	#Valor del calibrando en posición 2 (centro calibrando)
        #print(MedicionBloque)
        
        #Movimiento de 2 a 1 con el palpador arriba
        GPIO.output(pin_enableCalibrationMotor, motorEnabledState)       #habilita los motores
        steperMotor1.motor_go(False, "1/8", 841, .0025, False, 2)
        GPIO.output(pin_enableCalibrationMotor, motorDisabledState)       #Modo seguro, motores inhabilitados
        ActivaPedal(servo_pin)								#Baja el palpador (termina cada repetición con el palpador abajo)

    #Una vez finalizadas las mediciones de los bloques el palpador se mueve a HOME
    sleep(int(tiempoestabilizacion))
    ActivaPedal(servo_pin)                                  #Sube palpador
	#Movimiento de 1 a HOME con el palpador arriba
    GPIO.output(pin_enableCalibrationMotor, motorEnabledState)       #habilita los motores
    steperMotor1.motor_go(True, "Half", 213, .005, False, 2)#Movimiento de punto1 a HOME
    GPIO.output(pin_enableCalibrationMotor, motorDisabledState)       #Modo seguro, motores inhabilitados
    ActivaPedal(servo_pin)                                  #Baja palpador
    
    #obtenerAnguloBloque(valorNominalBloques[dato])          #Moverse a la siguiente pareja de bloques
    toc=time.perf_counter()                                 #Toma el tiempo final
    global tiempoCorrida
    tiempoCorrida=toc-tic                                   #retorna el tiempo de corrida en segundos
    global t0
    t0=time.time()                                   #inicia el conteo de espera de bloques
    return listaMediciones, tiempoCorrida, t0

################## Secuencia desviación de longitud central + planitud (Plantilla 1) ##################

def Completa1(tiempoinicial, tiempoestabilizacion, Repeticiones):
    
    global valorNominalBloque
    global dato
    
    #obtenerAnguloBloque(valorNominalBloque[dato])          #Moverse a la siguiente pareja de bloques
    global t1
    t1=time.time()                                   #finaliza el conteo de espera de bloques
    tic=time.perf_counter()                                 #Toma el tiempo inicial
    
    listaMediciones=[]
    
    sleep(int(tiempoinicial)*60)
    
    for i in range(int(Repeticiones)):

        ActivaPedal(servo_pin) #Baja palpador
        sleep(int(tiempoestabilizacion)) #Tiempo de estabilización
        ActivaPedal(servo_pin) #Sube palpador
        MedicionBloque=DatosTESA() #Llama función TESA
        listaMediciones.append(MedicionBloque) #Valor del patrón en posición 1 (centro)
        print(MedicionBloque)

        GPIO.output(pin_enableCalibrationMotor, motorEnabledState)       #habilita los motores
        steperMotor1.motor_go(True, "Half", 417, .005, False, 2) #Mov de 1 a 2
        GPIO.output(pin_enableCalibrationMotor, motorDisabledState)       #Modo seguro, motores inhabilitados

        ActivaPedal(servo_pin) #Baja palpador
        sleep(int(tiempoestabilizacion)) #Tiempo de estabilización
        ActivaPedal(servo_pin) #Sube palpador
        MedicionBloque=DatosTESA() #Llama función TESA
        listaMediciones.append(MedicionBloque)  #Valor del calibrando en posición 2 (esquina)
        print(MedicionBloque)
        
        GPIO.output(pin_enableCalibrationMotor, motorEnabledState)       #habilita los motores
        steperMotor1.motor_go(False, "Half", 96, .005, False, 2) #Mov1 de 2 a 3
        steperMotor2.motor_go(False, "Full", 398, .005, False, 1) #Mov2 de 2 a 3
        steperMotor1.motor_go(True, "Half", 178, .005, False, 1) #Mov3 de 2 a 3
        GPIO.output(pin_enableCalibrationMotor, motorDisabledState)       #Modo seguro, motores inhabilitados

        ActivaPedal(servo_pin) #Baja palpador
        sleep(int(tiempoestabilizacion)) #Tiempo de estabilización
        ActivaPedal(servo_pin) #Sube palpador
        MedicionBloque=DatosTESA() #Llama función TESA
        listaMediciones.append(MedicionBloque) #Valor del calibrando en posición 3 (esquina)
        print(MedicionBloque)
        
        GPIO.output(pin_enableCalibrationMotor, motorEnabledState)       #habilita los motores            
        steperMotor1.motor_go(False, "Half", 178, .005, False, 2) #Mov de 3 a 4
        GPIO.output(pin_enableCalibrationMotor, motorDisabledState)       #Modo seguro, motores inhabilitados

        ActivaPedal(servo_pin) #Baja palpador
        sleep(int(tiempoestabilizacion)) #Tiempo de estabilización
        ActivaPedal(servo_pin) #Sube palpador
        MedicionBloque=DatosTESA() #Llama función TESA
        listaMediciones.append(MedicionBloque)  #Valor del calibrando en posición 4 (esquina)
        print(MedicionBloque)
        
        GPIO.output(pin_enableCalibrationMotor, motorEnabledState)       #habilita los motores
        steperMotor2.motor_go(True, "Full", 796, .005, False, 2) #Mov de 4 a 5
        GPIO.output(pin_enableCalibrationMotor, motorDisabledState)       #Modo seguro, motores inhabilitados

        ActivaPedal(servo_pin) #Baja palpador
        sleep(int(tiempoestabilizacion)) #Tiempo de estabilización
        ActivaPedal(servo_pin) #Sube palpador
        MedicionBloque=DatosTESA() #Llama función TESA
        listaMediciones.append(MedicionBloque)  #Valor del calibrando en posición 5 (esquina)
        print(MedicionBloque)
        
        GPIO.output(pin_enableCalibrationMotor, motorEnabledState)       #habilita los motores
        steperMotor1.motor_go(True, "Half", 174, .005, False, 2) #Mov de 5 a 6
        GPIO.output(pin_enableCalibrationMotor, motorDisabledState)       #Modo seguro, motores inhabilitados

        ActivaPedal(servo_pin) #Baja palpador
        sleep(int(tiempoestabilizacion)) #Tiempo de estabilización
        ActivaPedal(servo_pin) #Sube palpador
        MedicionBloque=DatosTESA() #Llama función TESA
        listaMediciones.append(MedicionBloque)  #Valor del calibrando en posición 6 (esquina)
        print(MedicionBloque)
        
        GPIO.output(pin_enableCalibrationMotor, motorEnabledState)       #habilita los motores
        steperMotor1.motor_go(False, "Half", 174, .005, False, 2) #Mov de 6 a 5
        steperMotor2.motor_go(False, "Full", 398, .005, False, 1) #Mov de 5 a Esp2
        steperMotor1.motor_go(False, "Half", 330, .005, False, 1) #Mov de Esp2 a 1
        GPIO.output(pin_enableCalibrationMotor, motorDisabledState)       #Modo seguro, motores inhabilitados

        ActivaPedal(servo_pin) #Baja palpador
        sleep(int(tiempoestabilizacion)) #Tiempo de estabilización
        ActivaPedal(servo_pin) #Sube palpador
        MedicionBloque=DatosTESA() #Llama función TESA
        listaMediciones.append(MedicionBloque)  #Valor del patrón en posición 1 (centro)
        print(MedicionBloque)

    
    GPIO.output(pin_enableCalibrationMotor, motorEnabledState)       #habilita los motores
    steperMotor1.motor_go(True, "Half", 208, .005, False, 2) #Mov de 1 a HOME
    GPIO.output(pin_enableCalibrationMotor, motorDisabledState)       #Modo seguro, motores inhabilitados

    ActivaPedal(servo_pin) #Baja palpador
    
    listaMediciones.append(MedicionBloque)
    #obtenerAnguloBloque(valorNominalBloques[dato])          #Moverse a la siguiente pareja de bloques
    toc=time.perf_counter()                                 #Toma el tiempo final
    global tiempoCorrida
    tiempoCorrida=toc-tic                            #retorna el tiempo de corrida en segundos
    global t0
    t0=time.time()                                   #inicia el conteo de espera de bloques
    return listaMediciones, tiempoCorrida, t0

################## Secuencia desviación de longitud central + planitud (Plantilla 2) ##################

def Completa2(tiempoinicial, tiempoestabilizacion, Repeticiones):
        
    global valorNominalBloque
    global dato
    
    #obtenerAnguloBloque(valorNominalBloque[dato])          #Moverse a la siguiente pareja de bloques
    
    global t1
    t1=time.time()                                   #finaliza el conteo de espera de bloques
    tic=time.perf_counter()                                 #Toma el tiempo inicial

    listaMediciones=[]
    
    sleep(int(tiempoinicial)*60)
    
    for i in range(int(Repeticiones)):

        ActivaPedal(servo_pin)                              #Baja palpador
        sleep(int(tiempoestabilizacion))                    #Tiempo de estabilización
        ActivaPedal(servo_pin)                              #Sube palpador
        MedicionBloque=DatosTESA()                   #Llama función TESA
        print(MedicionBloque)
        listaMediciones.append(MedicionBloque)              #Valor del patrón en posición 1 (centro)
        
        GPIO.output(pin_enableCalibrationMotor, motorEnabledState)       #habilita los motores
        steperMotor1.motor_go(True, "Half", 416, .005, False, 2) #Mov de 1 a 2
        GPIO.output(pin_enableCalibrationMotor, motorDisabledState)       #Modo seguro, motores inhabilitados

        ActivaPedal(servo_pin)                              #Baja palpador
        sleep(int(tiempoestabilizacion))                    #Tiempo de estabilización
        ActivaPedal(servo_pin)                              #Sube palpador
        MedicionBloque=DatosTESA()                   #Llama función TESA
        print(MedicionBloque)
        listaMediciones.append(MedicionBloque)              #Valor del calibrando en posición 2 (esquina)
        
        GPIO.output(pin_enableCalibrationMotor, motorEnabledState)       #habilita los motores        
        steperMotor1.motor_go(False, "Half", 96, .005, False, 2)    #Mov1 de 2 a 3
        steperMotor2.motor_go(False, "Full", 337, .005, False, 1)   #Mov2 de 2 a 3
        steperMotor1.motor_go(True, "Half", 182, .005, False, 1)    #Mov3 de 2 a 3
        GPIO.output(pin_enableCalibrationMotor, motorDisabledState)       #Modo seguro, motores inhabilitados

        ActivaPedal(servo_pin)                              #Baja palpador
        sleep(int(tiempoestabilizacion))                    #Tiempo de estabilización
        ActivaPedal(servo_pin)                              #Sube palpador
        MedicionBloque=DatosTESA()                   #Llama función TESA
        print(MedicionBloque)
        listaMediciones.append(MedicionBloque)              #Valor del calibrando en posición 3 (esquina)
        
        GPIO.output(pin_enableCalibrationMotor, motorEnabledState)       #habilita los motores        
        steperMotor1.motor_go(False, "Half", 183, .005, False, 2)      #Mov de 3 a 4
        GPIO.output(pin_enableCalibrationMotor, motorDisabledState)       #Modo seguro, motores inhabilitados

        ActivaPedal(servo_pin)                              #Baja palpador
        sleep(int(tiempoestabilizacion))                    #Tiempo de estabilización
        ActivaPedal(servo_pin)                              #Sube palpador
        MedicionBloque=DatosTESA()                   #Llama función TESA
        print(MedicionBloque)
        listaMediciones.append(MedicionBloque)              #Valor del calibrando en posición 4 (esquina)
        
        GPIO.output(pin_enableCalibrationMotor, motorEnabledState)       #habilita los motores
        steperMotor2.motor_go(True, "Full", 683, .005, False, 2) #Mov de 4 a 5
        GPIO.output(pin_enableCalibrationMotor, motorDisabledState)       #Modo seguro, motores inhabilitados

        ActivaPedal(servo_pin)                              #Baja palpador
        sleep(int(tiempoestabilizacion))                    #Tiempo de estabilización
        ActivaPedal(servo_pin)                              #Sube palpador
        MedicionBloque=DatosTESA()                   #Llama función TESA
        print(MedicionBloque)
        listaMediciones.append(MedicionBloque)              #Valor del calibrando en posición 5 (esquina)
        
        GPIO.output(pin_enableCalibrationMotor, motorEnabledState)       #habilita los motores
        steperMotor1.motor_go(True, "Half", 178, .005, False, 2) #Mov de 5 a 6
        GPIO.output(pin_enableCalibrationMotor, motorDisabledState)       #Modo seguro, motores inhabilitados

        ActivaPedal(servo_pin)                              #Baja palpador
        sleep(int(tiempoestabilizacion))                    #Tiempo de estabilización
        ActivaPedal(servo_pin)                              #Sube palpador
        MedicionBloque=DatosTESA()                   #Llama función TESA
        print(MedicionBloque)
        listaMediciones.append(MedicionBloque)              #Valor del calibrando en posición 6 (esquina)
        
        GPIO.output(pin_enableCalibrationMotor, motorEnabledState)       #habilita los motores
        steperMotor1.motor_go(False, "Half", 178, .005, False, 2) #Mov de 6 a 5
        steperMotor2.motor_go(False, "Full", 342, .005, False, 1) #Mov de 5 a Esp2
        steperMotor1.motor_go(False, "Half", 332, .005, False, 1) #Mov de Esp2 a 1
        GPIO.output(pin_enableCalibrationMotor, motorDisabledState)       #Modo seguro, motores inhabilitados

        ActivaPedal(servo_pin)                              #Baja palpador
        sleep(int(tiempoestabilizacion))                    #Tiempo de estabilización
        ActivaPedal(servo_pin)                              #Sube palpador
        MedicionBloque=DatosTESA()                   #Llama función TESA
        print(MedicionBloque)
        listaMediciones.append(MedicionBloque)              #Valor del patrón en posición 1 (centro)

    GPIO.output(pin_enableCalibrationMotor, motorEnabledState)       #habilita los motores
    steperMotor1.motor_go(True, "Half", 208, .005, False, 2) #Mov de 1 a HOME
    GPIO.output(pin_enableCalibrationMotor, motorDisabledState)       #Modo seguro, motores inhabilitados
        
    ActivaPedal(servo_pin)                              #Baja palpador
    listaMediciones.append(MedicionBloque)              #Valor del patrón en posición 1 (centro)
    
    #obtenerAnguloBloque(valorNominalBloques[dato])          #Moverse a la siguiente pareja de bloques
    toc=time.perf_counter()                                 #Toma el tiempo final
    global tiempoCorrida
    tiempoCorrida=toc-tic                                   #retorna el tiempo de corrida en segundos
    global t0
    t0=time.time()                                   #inicia el conteo de espera de bloques
    return listaMediciones, tiempoCorrida, t0

################## Búsqueda de Clientes ##################

def BusquedaClientes(nombreClienteBuscado):
    """
    Entrada: 
        nombreCliente: nombre del cliente para el cual se va a calibrar
    Salida: una lista con el nombre del cliente, su dirección y el archivo donde está almacenada su información
    """
    woorkbookClientes = load_workbook(filename="Clientes.xlsx", keep_vba=True) # Apertura del archivo de excel de clientes 
    hojaClientes = woorkbookClientes.active # Hoja del archivo de excel donde están los clientes y su información
    
    i = 3 # Se inicializa un contador que va a recorrer los números de fila, empezando por la fila 3 porque las filas anteriores son encabezados
    while hojaClientes["A"+str(i)].value != None: # Mientras la celda de cliente no esté vacía, se van a seguir recorriendo las filas
        if hojaClientes["A"+str(i)].value == nombreClienteBuscado: # Si el valor de la celda es igual al nombre del cliente que se busca:
            numFila = i # El número de fila donde se encuentra la información del cliente corresponde al valor actual del contador 
        i += 1

    # Se almacenan el variables la información del cliente:
    # La columna A corresponde a "Nombre del Cliente", B a "Dirección del Cliente" y C a "Nombre del archivo del cliente"
    # numFila permite crear la coordenada de la celda que almacena la información deseada
    nombreCliente = hojaClientes["A"+str(numFila)].value 
    direccionCliente = hojaClientes["B"+str(numFila)].value
    archivoCliente = hojaClientes["C"+str(numFila)].value

    return nombreCliente, direccionCliente, archivoCliente

################## Selector de machote ##################

def selectorMachote(seleccionSecuencia):
	if seleccionSecuencia == "Desviación central" :
		machote = "./Machotes/Machote para calibración de Bloques con comparador mecánico TESA (Desviación central).xlsm"
		return machote
	elif seleccionSecuencia == "Desviación central y planitud" :
		machote = "./Machotes/Machote para calibración de Bloques con comparador mecánico TESA (Desviación central y planitud).xlsm"
		return machote
		
################## Creación de un archivo para la calibración ##################

def CrearArchivoCalibracion(seleccionSecuencia, numCertificado):
	# Se escoge sobre qué machote se va a trabajar a partir de la secuencia de calibración escogida por el usuario:
	machote = selectorMachote(seleccionSecuencia)

	# Se crea un duplicado del machote, nombrado con una marca temporal:
	archivoCalibracion = "./Calibraciones en curso/" + numCertificado + ".xlsm" # Nombre del archivo para la calibración
	shutil.copy(machote, archivoCalibracion) # Creación del duplicado del machote

	return archivoCalibracion


################## Creación de un archivo csv para Datos ##################

def CrearArchivoCSV(seleccionSecuencia, numCertificado):
	# Se crea un archivo csv, nombrado con una marca temporal:
	archivoDatos = "./Calibraciones en curso/" + numCertificado + ".csv" # Nombre del archivo para el almacenaje de datos
	open(archivoDatos, mode="w", newline="")	#Creación del Archivo

    # Se crean también para el registro de condiciones ambientales

	archivoDatosAmbientales = "./Calibraciones en curso/" + numCertificado + "-Ambientales.csv" # Nombre del archivo para el almacenaje de datos
	open(archivoDatosAmbientales, mode="w", newline="")	#Creación del Archivo


	return archivoDatos,archivoDatosAmbientales

################## Autocompletado de la información que se tiene del cliente y la calibración ##################

def AutocompletarInformacionCliente(nombreCliente, direccionCliente, numeroCertificado, numeroSolicitud, identificacionCalibrando, 
                                    responsableCalibracion, responsableRevision, patron, materialPatron, seleccionSecuencia):
    # Lista con la información del cliente para el que se va a calibrar
    informacionCliente = BusquedaClientes(nombreCliente)
   
    # Carga del archivo de excel que contiene la información del cliente 
    workbookSolicitantes = load_workbook(filename=informacionCliente[2], keep_vba = True, data_only = True)
    hojaJuego = workbookSolicitantes[identificacionCalibrando] # Selección de la hoja que contiene la información del juego a calibrar

    # Carga del archivo de excel creado para la calibración:
    workbookCalibracion = load_workbook(filename=CrearArchivoCalibracion(seleccionSecuencia, numeroCertificado), keep_vba=True)
    #Definir los nombres de las hojas del libro de excel en el que se está trabajando:
    hojaConversionDatos = workbookCalibracion["conversion datos"]
    hojaIdentificacionBloques = workbookCalibracion["Ident.Bloques a calibrar"]
    hojaResultadosCalibracion = workbookCalibracion["Introduccion de datos de Calib."]

    #Almacenamiento de la información del calibrando
    listaValores = [] #Se crea una lista para guardar los valores nominales de los bloques del juego seleccionado
    for filaValores in hojaJuego.iter_rows(min_row=12,
                                           min_col=2,
                                           max_col=2):
        for celdaValores in filaValores:
            listaValores.append(float(celdaValores.value))
        
    listaSeries = [] #Se crea una lista para guardar los valores de las series de los bloques del juego seleccionado
    for filaSeries in hojaJuego.iter_rows(min_row=12,
                                          min_col=3,
                                          max_col=3):
        for celdaSeries in filaSeries:
            listaSeries.append(int(celdaSeries.value))

    objetoCalibrar = hojaJuego["C2"].value #Extrae el objeto a calibrar del archivo del cliente
    marcaInstrumento = hojaJuego["C3"].value #Extrae la marca del instrumento del archivo del cliente
    serieInstrumento = hojaJuego["C4"].value #Extrae la serie del instrumento del archivo del cliente
    material = hojaJuego["C5"].value #Extrae el material del calibrando del cliente
    modelo = hojaJuego["C6"].value #Extrae el modelo del calibrando del cliente
    gradoDeclarado = hojaJuego["C7"].value #Extra el gradoDeclarado del calibrando del cliente

    #Copiar información de los valores nominales y series de los bloques del juego a la hoja de la calibración
    i = 12
    while i <= (len(listaValores)+11):
        #Copiar información de los valores nominales:
        coordenadaCeldaValor = "B"+str(i-8) #Coordenada de la celda donde se encuentra el valor nominal de cada bloque a lo largo del while
        hojaIdentificacionBloques[coordenadaCeldaValor] = listaValores[i-12] #Escritura del valor nominal del bloque en la hoja correspondiente a la calibración
        #Copiar información de las series:
        coordenadaCeldaSerie = "C"+str(i-8) #Coordenada de la celda donde se encuentra la serie de cada bloque a lo largo del while
        hojaIdentificacionBloques[coordenadaCeldaSerie] = listaSeries[i-12] #Escritura de la serie del bloque en la hoja correspondiente a la calibración

        i+=1

    #Autocompletado de la información del calibrando en la hoja de la calibración
    hojaConversionDatos["L6"] = numeroCertificado
    hojaConversionDatos["L9"] = nombreCliente
    hojaConversionDatos["L11"] = numeroSolicitud
    hojaConversionDatos["L12"] = direccionCliente

    hojaConversionDatos["L17"] = objetoCalibrar
    hojaConversionDatos["L19"] = marcaInstrumento
    hojaConversionDatos["L20"] = serieInstrumento
    hojaConversionDatos["L21"] = material 
    hojaConversionDatos["L22"] = modelo
    hojaConversionDatos["L23"] = gradoDeclarado

    hojaConversionDatos["L27"] = responsableCalibracion
    hojaConversionDatos["L28"] = responsableRevision

    hojaConversionDatos["L30"] = patron
    hojaConversionDatos["L31"] = materialPatron

    return workbookCalibracion, hojaResultadosCalibracion, hojaConversionDatos

def EncabezadosDesviacionCentral(numRepeticiones, hojaResultadosCalibracion):
    #SI NO SE HA REANUDADO LA CALIBRACIÓN
    #Modificar la hoja que va a almacenar los resultados de la calibración siguiendo la secuencia COMPLETA
    #Manejar los resultados de la calibración obtenidos con la secuencia COMPLETA
    #Completa: 1 Patrón, 1 Calibrando por repetición -> 2 columnas por repetición

    hojaResultadosCalibracion["H2"] = numRepeticiones
    numNuevasColumnas = 2*numRepeticiones #Se usan dos columnas por cada repetición: una para el patrón y otra para el calibrando
    hojaResultadosCalibracion.insert_cols(idx=19, amount=numNuevasColumnas) # Insertar el número de columnas necesarias al final de las columnas llenas en la hoja (Columna S)

    #Definir estilos 
    texto_negrita = Font(bold = True)
    texto_centrado = Alignment(horizontal = "center", vertical="center", wrapText=True)
    borde_sencillo = Side(border_style = "thin")
    borde_cuadrado = Border(top = borde_sencillo,
                            right = borde_sencillo,
                            bottom = borde_sencillo,
                            left = borde_sencillo)

    j = 19 #Se inicializa el contador para las columnas (Columna S)
    k = 1 #Se inicializa el contador para las repeticiones

    while j <= (19+numNuevasColumnas)-1 and k <= numRepeticiones:
        letraColumnaPatron = openpyxl.utils.cell.get_column_letter(j) #Obtener la letra de la columna en la que se está trabajando
        letraColumnaCalibrando = openpyxl.utils.cell.get_column_letter(j+1) #Obtener la letra de la columna de la derecha a la que se está trabajando
        #Coordenadas de las celdas del patrón y calibrando #repetición
        coordenadaEncabezadoPatron = letraColumnaPatron + "1"
        coordenadaEncabezadoCalibrando = letraColumnaCalibrando + "1"
        #Escribir los encabezados de las nuevas celdas:
        hojaResultadosCalibracion[coordenadaEncabezadoPatron] = "Patrón #"+str(k)
        hojaResultadosCalibracion[coordenadaEncabezadoCalibrando] = "Calibrando #"+str(k)
        #Darle formato a las nuevas celdas:
        hojaResultadosCalibracion[coordenadaEncabezadoPatron].font = texto_negrita
        hojaResultadosCalibracion[coordenadaEncabezadoPatron].alignment = texto_centrado
        hojaResultadosCalibracion[coordenadaEncabezadoPatron].border = borde_cuadrado
        hojaResultadosCalibracion[coordenadaEncabezadoCalibrando].font = texto_negrita
        hojaResultadosCalibracion[coordenadaEncabezadoCalibrando].alignment = texto_centrado
        hojaResultadosCalibracion[coordenadaEncabezadoCalibrando].border = borde_cuadrado
    
        k += 1
        j += 2 
    return numNuevasColumnas

def EncabezadosCentroYPlanitud(numRepeticiones, hojaResultadosCalibracion):

    """
    SI NO SE HA REANUDADO LA CALIBRACIÓN
    Modificar la hoja que va a almacenar los resultados de la calibración siguiendo la secuencia CENTROS
    Manejar los resultados de la calibración obtenidos con la secuencia CENTROS 
    Centros: 1 Patrón (centro), 5 Calibrando: centro y cuatro esquinas -> 6 columnas por repetición
    La secuencia Centros hace lo mismo que la secuencia Completa para el caso de los centros de los bloques, pero además 
    complementa esta calibración con mediciones de planitud para cada uno de los bloques
    """ 
    numRepeticiones = int(ventanaEntrada("Indique el número de repeticiones para la calibración: "))
    hojaResultadosCalibracion["N2"] = numRepeticiones
    numNuevasColumnas = 6*numRepeticiones #Se usan dos columnas por cada repetición: una para el patrón y otra para el calibrando
    hojaResultadosCalibracion.insert_cols(idx=25, amount=numNuevasColumnas) # Insertar el número de columnas necesarias al final de las columnas llenas en la hoja

    #Definir estilos 
    texto_negrita = Font(bold = True)
    texto_centrado = Alignment(horizontal = "center", vertical="center", wrapText=True)
    borde_sencillo = Side(border_style = "thin")
    borde_cuadrado = Border(top = borde_sencillo,
                            right = borde_sencillo,
                            bottom = borde_sencillo,
                            left = borde_sencillo)

    j = 25 #Se inicializa el contador para las columnas
    k = 1 #Se inicializa el contador para las repeticiones

    while j <= (25+numNuevasColumnas)-1 and k <= numRepeticiones:
        letraColumnaPatron = openpyxl.utils.cell.get_column_letter(j) #Obtener la letra de la columna que va a guardar los datos del Patrón en cada rep k 
        letraColumnaCalibrandoCentro = openpyxl.utils.cell.get_column_letter(j+1) #Obtener la letra de la columna que va a guardar los datos del Centro del Calibrando en cada rep k 
        letraColumnaCalibrandoEsquina3 = openpyxl.utils.cell.get_column_letter(j+2) #Obtener la letra de la columna que va a guardar los datos de la Esquina 3 del Calibrando en cada rep k 
        letraColumnaCalibrandoEsquina4 = openpyxl.utils.cell.get_column_letter(j+3) #Obtener la letra de la columna que va a guardar los datos de la Esquina 4 del Calibrando en cada rep k 
        letraColumnaCalibrandoEsquina5 = openpyxl.utils.cell.get_column_letter(j+4) #Obtener la letra de la columna que va a guardar los datos de la Esquina 5 del Calibrando en cada rep k 
        letraColumnaCalibrandoEsquina6 = openpyxl.utils.cell.get_column_letter(j+5) #Obtener la letra de la columna que va a guardar los datos de la Esquina 3 del Calibrando en cada rep k 

        #Escribir los encabezados de las nuevas celdas:
        hojaResultadosCalibracion[letraColumnaPatron + "1"] = "Patrón (Centro) #"+str(k)
        hojaResultadosCalibracion[letraColumnaCalibrandoCentro + "1"] = "Calibrando (Centro) #"+str(k)
        hojaResultadosCalibracion[letraColumnaCalibrandoEsquina3 + "1"] = "Calibrando (Esquina 3) #"+str(k)
        hojaResultadosCalibracion[letraColumnaCalibrandoEsquina4 + "1"] = "Calibrando (Esquina 4) #"+str(k)
        hojaResultadosCalibracion[letraColumnaCalibrandoEsquina5 + "1"] = "Calibrando (Esquina 5) #"+str(k)
        hojaResultadosCalibracion[letraColumnaCalibrandoEsquina6 + "1"] = "Calibrando (Esquina 6) #"+str(k)
        
        #Darle formato a las nuevas celdas:
        for numColumna in range(j,j+6):
            letraColumna = openpyxl.utils.cell.get_column_letter(numColumna)
            hojaResultadosCalibracion[letraColumna + "1"].font = texto_negrita
            hojaResultadosCalibracion[letraColumna + "1"].alignment = texto_centrado
            hojaResultadosCalibracion[letraColumna + "1"].border = borde_cuadrado
    
        k += 1
        j += 6
        return numNuevasColumnas

################## Cálculos del promedio y la desviación estándar ###################

def CalculosDesviacionCentral(hojaResultadosCalibracion, numNuevasColumnas, numRepeticiones):

    #Calcular el promedio de la diferencia entre el patrón y el calibrando con fórmulas en Excel
    #Calcular la desviación estándar del promedio de la diferencia entre el patrón y el calibrando con fórmulas en Excel
    
    l = 2 #Se inicializa el contador para filas

    while hojaResultadosCalibracion["S"+str(l)].value != None:
        listaDiferencias = [] #Inicializamos una lista para guardar los strings del cálculo de las diferencias, ej: A3-A2
        j = 19 #Se vuelve a inicializar el contador para las columnas en S
        k = 1 #Se vuelve a inicializar el contador para las repeticiones
        while j <= (19+numNuevasColumnas)-1 and k <= numRepeticiones:
            letraColumnaPatron = openpyxl.utils.cell.get_column_letter(j) #Obtener la letra de la columna en la que se está trabajando
            letraColumnaCalibrando = openpyxl.utils.cell.get_column_letter(j+1) #Obtener la letra de la columna de la derecha a la que se está trabajando
            #Coordenadas de las celdas del patrón y calibrando #repetición: tomando en cuenta el cambio de fila
            coordenadaPatron = letraColumnaPatron + str(l)
            coordenadaCalibrando = letraColumnaCalibrando + str(l)
            stringDiferencia = coordenadaCalibrando + "-" + coordenadaPatron
            #Se agrega el string para calcular la diferencia a listaDiferencias
            listaDiferencias.append(stringDiferencia)
            k += 1
            j += 2 
        #Se crea un string para que funcione como el argumento de la fórmula del promedio y desviación estándar a partir de listaDiferencias
        argumentoFormulaCentral = ""
        for stringDif in range(len(listaDiferencias)):
            if stringDif < len(listaDiferencias)-1:
                argumentoFormulaCentral = argumentoFormulaCentral + listaDiferencias[stringDif] + ";"
            elif stringDif == len(listaDiferencias)-1:
                argumentoFormulaCentral = argumentoFormulaCentral + listaDiferencias[stringDif]           
        
        #Se obtienen las fórmulas predefinidas en la hoja de excel
        formulaPromedio = hojaResultadosCalibracion["E"+str(l)].value
        formulaDesvst = hojaResultadosCalibracion["F"+str(l)].value
        
        #Se modifican las fórmulas agregando el argumento construído
        promedioModif = formulaPromedio.replace("))",f"{argumentoFormulaCentral}))")
        DesvstModif = formulaDesvst.replace("))",f"{argumentoFormulaCentral}))")
        
        #Se actualizan las celdas con las fórmulas modificadas
        hojaResultadosCalibracion["E"+str(l)] = promedioModif
        hojaResultadosCalibracion["F"+str(l)] = DesvstModif
        
        l += 1
    return

def CalculosDesviacionYPlanitud(hojaResultadosCalibracion, numNuevasColumnas, numRepeticiones):
    l = 2 #Se inicializa el contador para filas -> Empezamos a agregar valores en la fila 2

    while hojaResultadosCalibracion["Y"+str(l)].value != None:
        listaDiferenciasCentros = [] #Inicializamos una lista para guardar el string para calcular la diferencia entre el centro del calibrando y el patrón en cada repetición
        listaDiferenciasEsquina3 = [] #Inicializamos una lista para guardar el string para calcular la diferencia entre la esquina 3 del calibrando y el centro del patrón en cada repetición
        listaDiferenciasEsquina4 = [] #Inicializamos una lista para guardar el string para calcular la diferencia entre la esquina 4 del calibrando y el centro del patrón en cada repetición
        listaDiferenciasEsquina5 = [] #Incializamos una lista para guardar el string para calcular la diferencia entre la esquina 5 del calibrando y el centro del patrón en cada repetición
        listaDiferenciasEsquina6 = [] #Incializamos una lista para guardar el string para calcular la diferencia entre la esquina 6 del calibrando y el centro del patrón en cada repetición

        j = 25 #Se vuelve a inicializar el contador para las columnas
        k = 1 #Se vuelve a inicializar el contador para las repeticiones
        while j <= (25+numNuevasColumnas)-1 and k <= numRepeticiones:
            letraColumnaPatron = openpyxl.utils.cell.get_column_letter(j) #Obtener la letra de la columna en la que se encuentra la medicón del centro del patrón
            letraColumnaCalibrando = openpyxl.utils.cell.get_column_letter(j+1) #Obtener la letra de la columna en la que se encuentra el centro del calibrando
            letraColumnaEsquina3 = openpyxl.utils.cell.get_column_letter(j+2) #Obtener la letra de la columna en la que se encuentra la medición de la esquina 3 del calibrando
            letraColumnaEsquina4 = openpyxl.utils.cell.get_column_letter(j+3) #Obtener la letra de la columna en la que se encuentra la medición de la esquina 4 del calibrando
            letraColumnaEsquina5 = openpyxl.utils.cell.get_column_letter(j+4) #Obtener la letra de la columna en la que se encuentra la medición de la esquina 5 del calibrando
            letraColumnaEsquina6 = openpyxl.utils.cell.get_column_letter(j+5) #Obtener la letra de la columna en la que se encuentra la medición de la esquina 6 del calibrando
        
            #Coordenadas de las celdas del patrón y calibrando #repetición: tomando en cuenta el cambio de fila
            coordenadaPatron = letraColumnaPatron + str(l)
            coordenadaCalibrando = letraColumnaCalibrando + str(l)
            coordenadaEsquina3 = letraColumnaEsquina3 + str(l)
            coordenadaEsquina4 = letraColumnaEsquina4 + str(l)
            coordenadaEsquina5 = letraColumnaEsquina5 + str(l)
            coordenadaEsquina6 = letraColumnaEsquina6 + str(l)

            #String de la diferencia entre el calibrando y el centro del patrón
            diferenciaCentros = coordenadaCalibrando + "-" + coordenadaPatron

            #Strings de las diferencias entre las esquinas del calibrando y el centro del patrón
            diferenciaEsquina3 = coordenadaEsquina3 + "-" + coordenadaPatron
            diferenciaEsquina4 = coordenadaEsquina4 + "-" + coordenadaPatron
            diferenciaEsquina5 = coordenadaEsquina5 + "-" + coordenadaPatron
            diferenciaEsquina6 = coordenadaEsquina6 + "-" + coordenadaPatron

            #Se agrega el valor de las diferencias a sus respectivas listas:
            listaDiferenciasCentros.append(diferenciaCentros)
            listaDiferenciasEsquina3.append(diferenciaEsquina3)
            listaDiferenciasEsquina4.append(diferenciaEsquina4)
            listaDiferenciasEsquina5.append(diferenciaEsquina5)
            listaDiferenciasEsquina6.append(diferenciaEsquina6)

            k += 1
            j += 6

        #Formula para crear un string para que funcione como el argumento de la fórmulas a partir de las listas de diferencias
        def argumentoFormula(listaDiferencias):
            argumentoFormulaCentral = ""
            for stringDif in range(len(listaDiferencias)):
                if stringDif < len(listaDiferencias)-1:
                    argumentoFormulaCentral = argumentoFormulaCentral + listaDiferencias[stringDif] + ";"
                elif stringDif == len(listaDiferencias)-1:
                    argumentoFormulaCentral = argumentoFormulaCentral + listaDiferencias[stringDif]
            return argumentoFormulaCentral

        #Se crean los strings de los argumentos para las fórmulas
        argumentoPromedioCentros = argumentoFormula(listaDiferenciasCentros)
        argumentoPromedioEsquina3 = argumentoFormula(listaDiferenciasEsquina3)
        argumentoPromedioEsquina4 = argumentoFormula(listaDiferenciasEsquina4)
        argumentoPromedioEsquina5 = argumentoFormula(listaDiferenciasEsquina5)
        argumentoPromedioEsquina6 = argumentoFormula(listaDiferenciasEsquina6)

        #Se obtienen las fórmulas predefinidas en la hoja de excel (ya en el excel debe estar la fórmula con el argumento vacío)
        formulaPromedioCentros = hojaResultadosCalibracion["K"+str(l)].value
        formulaDesvstCentros = hojaResultadosCalibracion["L"+str(l)].value
        formulaPromedioEsquina3 = hojaResultadosCalibracion["E"+str(l)].value
        formulaPromedioEsquina4 = hojaResultadosCalibracion["F"+str(l)].value
        formulaPromedioEsquina5 = hojaResultadosCalibracion["G"+str(l)].value
        formulaPromedioEsquina6 = hojaResultadosCalibracion["H"+str(l)].value
        
        #Se modifican las fórmulas agregando el argumento construído
        promedioModifCentros = formulaPromedioCentros.replace("))",f"{argumentoPromedioCentros}))")
        DesvstModifCentros = formulaDesvstCentros.replace("))",f"{argumentoPromedioCentros}))")
        promedioModifEsquina3 = formulaPromedioEsquina3.replace("))",f"{argumentoPromedioEsquina3}))")
        promedioModifEsquina4 = formulaPromedioEsquina4.replace("))",f"{argumentoPromedioEsquina4}))")
        promedioModifEsquina5 = formulaPromedioEsquina5.replace("))",f"{argumentoPromedioEsquina5}))")
        promedioModifEsquina6 = formulaPromedioEsquina6.replace("))",f"{argumentoPromedioEsquina6}))")
        
        #Se actualizan las celdas con las fórmulas modificadas
        hojaResultadosCalibracion["K"+str(l)] = promedioModifCentros
        hojaResultadosCalibracion["L"+str(l)] = DesvstModifCentros
        hojaResultadosCalibracion["E"+str(l)] = promedioModifEsquina3
        hojaResultadosCalibracion["F"+str(l)] = promedioModifEsquina4
        hojaResultadosCalibracion["G"+str(l)] = promedioModifEsquina5
        hojaResultadosCalibracion["H"+str(l)] = promedioModifEsquina6

        l += 1

    return
    
################## Selector fila para la hoja de resultados ##################

def selectorFilaResultados(hojaResultadosCalibracion):
    i = 2 # Se inicializa el contador en 2 porque la fila 1 tiene los encabezados 
    for filaValorNominal in hojaResultadosCalibracion.iter_rows(min_row=2,
                                                                min_col=1,
                                                                max_col=1):
        for celdaValorNominal in filaValorNominal:
            if celdaValorNominal.value == None:
                numFila = i
            else:
                i += 1
    return numFila
    
################## Eliminar archivo ##################

def EliminarArchivo(rutaArchivoEliminar):
    #Revisar si el archivo existe
    if os.path.exists(rutaArchivoEliminar):
        #Borrar el archivo
        os.remove(rutaArchivoEliminar)
    else:
        mostrarMensaje("El archivo indicado no existe.")
    return 
    
################## Proceso de calibración de bloques ##################

def ProcesoCalibracion(seleccionSecuencia, tiempoinicial, tiempoestabilizacion, numRepeticiones, hojaResultadosCalibracion, hojaConversionDatos, nombreArchivoCalibracion, libroExcel,numCertificado,archivoDatos, archivoDatosAmbientales):
    
	#Si se va calibrar los bloques solo con desviación central se hace lo siguiente:
    if seleccionSecuencia == "Desviación central":
        continuarCalibracion = "sí"
        while continuarCalibracion == "sí":
            valorBloque = Decimal(float(ventanaEntrada("Indique el valor del bloque a Calibrar: "))) ## Debería agregarse un caso de error
            numFila = selectorFilaResultados(hojaResultadosCalibracion) # Se halla la fila a trabajar
            hojaResultadosCalibracion["A"+str(numFila)] = valorBloque # Se asigna el valor nominal del bloque ingresado por el usuario
            sleep(int(tiempoinicial)*60)					#Tiempo de estabilización inicial

			
			
			#Se realizan las mediciones de los bloques y se guardan en una lista [patrón, calibrando, patrón, calibrando,...]


            ########################
            # PARTE VIEJA
            ########################

			#Medición y registro de las condiciones ambientales iniciales
            """listaMedicionesTemperatura = DatosFluke()
            hojaResultadosCalibracion["I"+str(numFila)] = listaMedicionesTemperatura[0]
            hojaResultadosCalibracion["J"+str(numFila)] = listaMedicionesTemperatura[1]
            hojaResultadosCalibracion["K"+str(numFila)] = listaMedicionesTemperatura[2]
            hojaResultadosCalibracion["L"+str(numFila)] = listaMedicionesTemperatura[3]
            hojaResultadosCalibracion["M"+str(numFila)] = DatosVaisala()	#Dato de humedad relativa inicial"""
			
            numColumnaMediciones = 19 #Contador inicia en 19 porque ese es el número de la columna a partir del cual se empiezan a registar las mediciones de los bloques (Colummna S)
            """

            for numMedicion in range(len(listaMedicionesBloque)):
                letraColumnaMedicion = openpyxl.utils.cell.get_column_letter(numColumnaMediciones)
                hojaResultadosCalibracion[letraColumnaMedicion+str(numFila)] = listaMedicionesBloque[numMedicion]
                numColumnaMediciones += 1
            """
			
			#Medición y registro de las condiciones ambientales finales
            """listaMedicionesTemperatura = DatosFluke()
            hojaResultadosCalibracion["N"+str(numFila)] = listaMedicionesTemperatura[0]
            hojaResultadosCalibracion["O"+str(numFila)] = listaMedicionesTemperatura[1]
            hojaResultadosCalibracion["P"+str(numFila)] = listaMedicionesTemperatura[2]
            hojaResultadosCalibracion["Q"+str(numFila)] = listaMedicionesTemperatura[3]
            hojaResultadosCalibracion["R"+str(numFila)] = DatosVaisala()	#Dato de humedad relativa final"""
            


            ########################
            # PARTE NUEVA CSV
            ########################

            # Condiciones Ambientales Iniciales
            condicionesAmbientales = list(DatosFluke()) # 4 datos de temperatura
            condicionesAmbientales.append(DatosVaisala) # 1 dato de humedad relativa

            
            # Datos de Mediciones de Bloque Comparador
            listaMedicionesBloque = Centros(tiempoestabilizacion, numRepeticiones)[0]

            listaMedicionesBloque = [[str(num) for num in listaMedicionesBloque]] # Formato


            with open(archivoDatos, mode="a", newline="") as archivo:
                writer = csv.writer(archivo, delimiter=';')
                writer.writerows(listaMedicionesBloque)
            
            
            # Condiciones Ambientales Finales
            condicionesAmbientales = condicionesAmbientales + list(DatosFluke()) # 4 datos de temperatura
            condicionesAmbientales.append(DatosVaisala()) # 1 dato de humedad relativa

            condicionesAmbientales = [[str(num) for num in condicionesAmbientales]] #Formato

            with open(archivoDatosAmbientales, mode="a", newline="") as archivo:
                writer = csv.writer(archivo, delimiter=';')
                writer.writerows(condicionesAmbientales)





            continuarCalibracion = ventanaOpciones("¿Desea continuar con la calibración?:", ["sí", "no"]) 
            
        pausarCalibracion = ventanaOpciones("¿Desea Pausar la calibración o ya ha finalizado?", ["Pausar calibración", "Finalizar calibración"]) 
        if pausarCalibracion == "Pausar calibración": #Se Pausa la calibración (aún no se realizan cálculos)
            rutaGuardarPausa = "./Calibraciones en curso/" + nombreArchivoCalibracion
            libroExcel.save(rutaGuardarPausa)
            mostrarMensaje("Calibración pausada. \nPuede revisar el archivo correspondiente en la carpeta \"Calibraciones en curso\".")
                
        elif pausarCalibracion == "Finalizar calibración": #Se Finaliza la calibración
            duracionCalibracion = str(ventanaEntrada("Duración de la calibración (en días):"))
            hojaConversionDatos["L7"] = duracionCalibracion + " días" 
            CalculosDesviacionCentral(hojaResultadosCalibracion)
            rutaGuardar = "./Calibraciones Finalizadas/" + nombreArchivoCalibracion #Se guarda el archivo de la calibración a partir del número de Certificado
            libroExcel.save(rutaGuardar)  
            EliminarArchivo("./Calibraciones en curso/" + nombreArchivoCalibracion) #Se elimina el archivo de la calibración de la carpeta "Calibraciones en curso"

            # Mover archivos csv
            shutil.move(archivoDatos, "./Calibraciones Finalizadas/" + numCertificado +".csv")
            shutil.move(archivoDatosAmbientales, "./Calibraciones Finalizadas/" + numCertificado +"-Ambientales.csv")

            mostrarMensaje("Calibración finalizada. Puede revisar el archivo correspondiente en la carpeta \"Calibraciones Finalizadas\".")
				
	#Si se va a calibrar los bloques con desviación + planitud
    elif seleccionSecuencia == "Desviación central y planitud":
        continuarCalibracion = "sí"
        while continuarCalibracion == "sí":
            #Inluir aquí código para desviación central y planitud
            #
            #
            #
            #
            #
            #
            #
            #
            #
            #
            #
            continuarCalibracion = ventanaOpciones("¿Desea continuar con la calibración?: ", ["sí", "no"]) 
            
        pausarCalibracion = ventanaOpciones("¿Desea Pausar la calibración o ya ha finalizado?", ["Pausar calibración", "Finalizar calibración"]) 
        if pausarCalibracion == "Pausar calibración": #Se Pausa la calibración (aún no se realizan cálculos)
            rutaGuardarPausa = "./Calibraciones en curso/" + nombreArchivoCalibracion
            libroExcel.save(rutaGuardarPausa)
            mostrarMensaje("Calibración pausada. \nPuede revisar el archivo correspondiente en la carpeta \"Calibraciones en curso\".")
                
        elif pausarCalibracion == "Finalizar calibración": #Se Finaliza la calibración
            duracionCalibracion = str(ventanaEntrada("Duración de la calibración (en días):"))
            hojaConversionDatos["L7"] = duracionCalibracion + " días" 
            CalculosDesviacionYPlanitud(hojaResultadosCalibracion)
            rutaGuardar = "./Calibraciones Finalizadas/" + nombreArchivoCalibracion #Se guarda el archivo de la calibración a partir del número de Certificado
            libroExcel.save(rutaGuardar)  
            EliminarArchivo("./Calibraciones en curso/" + nombreArchivoCalibracion) #Se elimina el archivo de la calibración de la carpeta "Calibraciones en curso"
            mostrarMensaje("Calibración finalizada. Puede revisar el archivo correspondiente en la carpeta \"Calibraciones Finalizadas\".")   
    return
    
################## Nueva Calibración ##################

def NuevaCalibracion(nombreCliente, numeroCertificado, numeroSolicitud, identificacionCalibrando, 
                                    responsableCalibracion, responsableRevision, patron, materialPatron, seleccionSecuencia, tiempoinicial, tiempoestabilizacion, numRepeticiones):
    
    nombreCliente, direccionCliente, archivoCliente = BusquedaClientes(nombreCliente)		#Búsqueda de los datos del cliente
    
    
    archivoCalibracion = CrearArchivoCalibracion(seleccionSecuencia, numCertificado)
    archivoDatos, archivoDatosAmbientales = CrearArchivoCSV(seleccionSecuencia, numCertificado)
    

	#Ingreso de interés del cliente y de la calibración al archivo de Excel
    archivoExcel = AutocompletarInformacionCliente(nombreCliente, direccionCliente, numeroCertificado, numeroSolicitud, identificacionCalibrando, 
                                    responsableCalibracion, responsableRevision, patron, materialPatron, seleccionSecuencia)
    libroExcel = archivoExcel[0]
    hojaResultadosCalibracion = archivoExcel[1]
    hojaConversionDatos = archivoExcel[2]
    
    if seleccionSecuencia == "Desviación central":
        EncabezadosDesviacionCentral(numRepeticiones, hojaResultadosCalibracion)
    elif seleccionSecuencia == "Desviación central y planitud":
        EncabezadosCentroYPlanitud(numRepeticiones, hojaResultadosCalibracion)
    
    ProcesoCalibracion(seleccionSecuencia, tiempoinicial, tiempoestabilizacion, numRepeticiones, hojaResultadosCalibracion, hojaConversionDatos, nombreArchivoCalibracion, libroExcel, numCertificado,archivoDatos, archivoDatosAmbientales)
    
    return

################## Reanudar Calibración ##################

def ReanudarCalibracion(numCertificado, tiempoinicial, tiempoestabilizacion):
    
    nombreArchivoEnCurso = numCertificado + ".xlsm" 
    rutaEnCurso = "./Calibraciones en curso/" + nombreArchivoEnCurso
	
    archivoDatos = "./Calibraciones en curso/" + numCertificado + ".csv" # Nombre del archivo para el almacenaje de datos
    archivoDatosAmbientales = "./Calibraciones en curso/" + numCertificado + "-Ambientales.csv" 
    
    if os.path.exists(rutaEnCurso): #Si el archivo de la calibración en curso existe:
        workbookCalibracionEnCurso = load_workbook(filename = rutaEnCurso, keep_vba = True, data_only = True) #Apertura del archivo de excel de la calibración en curso
        hojaResultadosCalibracion = workbookCalibracionEnCurso["Introduccion de datos de Calib."] #Se abre la hoja de Excel donde se están registrando los datos de la calibración
    
        #Identificar con qué secuencia se está trabajando antes de continuar con la calibración
        if hojaResultadosCalibracion["S1"].value == "Patrón #1":
            seleccionSecuencia = "Desviación central"
        else: 
            seleccionSecuencia = "Desviación central y planitud"
        
        #Identificar el número de repeticiones con el que se está trabajando
        numRepeticiones = hojaResultadosCalibracion["H2"].value
    
        #Se continúa con el proceso de calibración
        ProcesoCalibracion(seleccionSecuencia, tiempoinicial, tiempoestabilizacion, numRepeticiones, hojaResultadosCalibracion, nombreArchivoEnCurso, workbookCalibracionEnCurso, numCertificado,archivoDatos, archivoDatosAmbientales)

    else: #Si el archivo indicado no existe
        mostrarMensaje("No hay una calibración en curso guardada con el número de certificado " + numCertificado +".")

    return

################## Agregar cliente ##################

def AgregarCliente(nombreCliente, direccionCliente):
    """
    Esta función permite agregar el nombre y la dirección de un nuevo cliente al archivo de Clientes.
    Además, crea el archivo del nuevo cliente donde se encuentra la información de sus juegos de bloques.
    """
    workbookClientes = load_workbook(filename="Clientes.xlsx", keep_vba=True) #Apertura del archivo de excel de clientes 
    hojaClientes = workbookClientes.active #Hoja del archivo de excel donde están los clientes y su información
    
    i = 3 #Se inicializa el contador para filas en 3 porque en la fila 1 y 2 están los encabezados
    #Ahora se deben recorrer las filas, empezando por la fila 3 para determinar el número de la fila que está libre para incluir un nuevo cliente
    for filaValorNominal in hojaClientes.iter_rows(min_row=3,
                                                    min_col=1,
                                                    max_col=1):
        for celdaValorNominal in filaValorNominal:
            if celdaValorNominal.value == None:
                numFila = i
            else:
                i += 1 

    machoteCliente = "./Machotes/Machote para nuevo cliente.xlsm"
    nombreArchivoCliente = nombreCliente + ".xlsx" #El nombre del archivo de Excel va a ser igual al nombre del Cliente
    shutil.copy(machoteCliente, "./Archivos de los clientes/" + nombreArchivoCliente)	

    #Se agrega la información del cliente al archivo de Clientes
    hojaClientes["A"+str(i)] = nombreCliente
    hojaClientes["B"+str(i)] = direccionCliente
    hojaClientes["C"+str(i)] = nombreArchivoCliente

    workbookClientes.save("./Clientes.xlsx")
    
    return

################## Ingresar juego de bloques/calibrando ##################

def IngresarCalibrando(nombreCliente, objeto, marca, numSerie, material, modelo, grado, unidad):
    archivoCliente = BusquedaClientes(nombreCliente)[2] #Busqueda del archivo del cliente
    workbookCliente = load_workbook(filename=archivoCliente, keep_vba=True)  #Apertura del archivo de excel del cliente

    #Revisar si ya existe algún calibrando registrado con el mismo número de serie
    existeCalibrando = False
    for serieCalibrandoRegistrado in workbookCliente.sheetnames:
        if serieCalibrandoRegistrado == numSerie:
            existeCalibrando = True
            break

    if existeCalibrando:
        mostrarMensaje("Ya existe un calibrando registrado con el númerio de serie " + numSerie + ".")

    #Crear una hoja para el nuevo calibrando
    if len(workbookCliente.sheetnames) > 1: #Si ya existen calibrandos registrados en el archivo
        hojaReferencia = workbookCliente.worksheets[0] #Se selecciona la hoja 1 como una referencia para crear la hoja para el nuevo juevo
        hojaNuevoCalibrando = workbookCliente.create_sheet(title = numSerie)
        
        #Se copia la hoja de referencia en la nueva hoja como plantilla
        for row in hojaReferencia.iter_rows():
            for cell in row:
                new_cell = hojaNuevoCalibrando.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy (cell.fill)
                    new_cell.number_format = cell.number_format
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)
    
    else: #Si no se han registrado calibrandos en el archivo del Cliente
        hojaNuevoCalibrando = workbookCliente.worksheets[0]
        hojaNuevoCalibrando.title = numSerie
    
    #Agregar la información del calibrando al archivo del cliente
    hojaNuevoCalibrando["A1"] = "Información del calibrando con identificación " + numSerie
    hojaNuevoCalibrando["C2"] = objeto
    hojaNuevoCalibrando["C3"] = marca
    hojaNuevoCalibrando["C4"] = numSerie
    hojaNuevoCalibrando["C5"] = material #Dropdown con opciones: Acero, cerámica, carburo de tungsteno, carburo de cromo
    hojaNuevoCalibrando["C6"] = modelo
    hojaNuevoCalibrando["C7"] = grado

    hojaNuevoCalibrando["B10"] = "Longitud nominal (" + unidad + ")"

    #Agregar valor nominal e identificación de los bloques del juego
    agregarNuevoBloque = "sí"
    numFila = 14 #Se inicia el contador para filas en 14 porque ahí empieza la lista de bloques
    while agregarNuevoBloque == "sí":
        valorBloqueIngresar = ventanaEntrada("Valor nomial del bloque a ingresar: ")
        idBloqueIngresar = ventanaEntrada("Identificación del bloque a ingresar: ")
        #Se agrega la información del bloque a la hoja
        hojaNuevoCalibrando["A"+str(numFila)] = numFila - 13
        hojaNuevoCalibrando["B"+str(numFila)] = valorBloqueIngresar
        hojaNuevoCalibrando["C"+str(numFila)] = idBloqueIngresar

        #Definir el estilo de los bordes de las celdas
        borde_sencillo = Side(border_style = "thin")
        borde_cuadrado = Border(top = borde_sencillo,
                                right = borde_sencillo,
                                bottom = borde_sencillo,
                                left = borde_sencillo)
        
        #Se le da estilo a la nuevas celdas
        hojaNuevoCalibrando["A"+str(numFila)].border = borde_cuadrado
        hojaNuevoCalibrando["B"+str(numFila)].border = borde_cuadrado
        hojaNuevoCalibrando["C"+str(numFila)].border = borde_cuadrado

        numFila += 1
        agregarNuevoBloque = ventanaOpciones("¿Desea agregar otro bloque?:", ["sí", "no"])
    
    mostrarMensaje("Se han ingresado exitosamente los datos del nuevo calibrando.")
    workbookCliente.save(archivoCliente)
    return
    
################## Ocultar advertencias en terminal ##################

def fxn():
    warnings.warn("deprecated", DeprecationWarning)

with warnings.catch_warnings():
    warnings.simplefilter("ignore")
    fxn()