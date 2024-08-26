"""
En este archivo se encuentran todas las funciones que el Software del Comparador de Bloques TESA (interfaz simple) requiere para funcionar
"""

################## Importación de librerías ##################
from time import sleep                                              # Biblioteca para sleep
import time
import serial                                                       # Biblioteca para configuración y adquisición de datos de dispositivos seriales
import openpyxl                                                     # Biblioteca para el manejo de archivos de excel
from openpyxl import load_workbook                                  # Biblioteca para cargar excel ya existente
import openpyxl.utils.cell                                          # Biblioteca para insertar columnas o filas en un excel 
from openpyxl.styles import Font, Color, Alignment, Border, Side    # Biblioteca para darle formato a archivos de excel
from openpyxl import Workbook                                       # Biblioteca para crear nuevos archivos de excel
from copy import copy                                               
import shutil                                                       # Biblioteca para copiar archivos
from decimal import Decimal                                         # Biblioteca para trabajar correctamente operaciones aritméticas con flotantes decimales
import curses														# Biblioteca para interacción con el teclado
import os                                                           # Biblioteca para interactuar con el sistema operativo
import warnings
import csv                                                          # Biblioteca para crear archivos csv e interactuar con ellos
import tkinter as tk
from tkinter import ttk, Scrollbar, Listbox
import requests
import win32com.client
import subprocess

################# Comunicación con el RPI - Servidor #######################


def ejecutarSecuencia(RPi_url, secuencia, tiempoinicial, tiempoestabilizacion, numRepeticiones, plantilla):
    data = {
        'secuencia':secuencia,
        'tiempoinicial':tiempoinicial,
        'tiempoestabilizacion':tiempoestabilizacion,
        'numRepeticiones':numRepeticiones,
        'plantilla':plantilla
    }
    url = RPi_url + "secuencias"
    response = requests.post(url, json=data)
    print(response)
    try:
        return response.json()
    except:
        mostrarMensaje("Error en la secuencia.\n\nPor favor reiniciar servicio y continuar\ncon la calibracion.")
        return ""


def condicionesAmbientales(RPi_url, instrumento):
    data = {
        'instrumento':instrumento
    }
    url = RPi_url + "condicionesAmbientales"
    exito = 0
    def peticion():
        response = 0
        exito = 0
        try:    
            response = requests.post(url, json=data)
            exito = 1
        except:
            response=0
            exito = 0
        return exito,response
    while exito==0:
        exito, response = peticion()
    print(response)
    return response.json()

def moverDe0a1(RPi_url):
    url = RPi_url + "moverDe0a1"
    response = requests.post(url)
    return response.json()

def moverPlatoRemoto(RPi_url, pos):
    url = RPi_url + "moverPlato"
    data = {
        'posición':pos
    }
    response = requests.post(url, json=data)
    return response.json()

def activarPedalRemoto(RPi_url):
    url = RPi_url + "activarPedalRemoto"
    response = requests.post(url)
    return response.json()




################## Ventanas Pop-up para valores de entrada y mensajes ##################

def ventanaEntrada(mensaje):
    def on_enter_press(event):
        input_value = input_entry.get()
        top.destroy()
        return_value.set(input_value)

    top = tk.Toplevel()
    top.title("Ingresar valor")
    top.configure(bg="white")

    input_label = ttk.Label(top, text=mensaje)
    input_label.pack(pady=5)

    input_entry = ttk.Entry(top, width=30)
    input_entry.pack(pady=5)

    return_value = tk.StringVar()
    input_entry.bind("<Return>", on_enter_press)

    top.wait_variable(return_value)
    return return_value.get()

def ventanaOpciones(mensaje, opciones):
    def on_enter_press(event):
        selected_value = combo_var.get()
        top.destroy()
        return_value.set(selected_value)

    top = tk.Toplevel()
    top.title("Seleccionar una opción")
    top.configure(bg="white")

    message_label = ttk.Label(top, text=mensaje)
    message_label.pack(pady=5)

    combo_var = tk.StringVar()
    combo_box = ttk.Combobox(top, textvariable=combo_var, values=opciones, width=27)
    combo_box.pack(pady=5)

    return_value = tk.StringVar()
    combo_box.bind("<Return>", on_enter_press)

    top.wait_variable(return_value)
    return return_value.get()

def mostrarMensaje(mensaje):
    def on_enter_press(event):
        top.destroy()

    top = tk.Toplevel()
    top.title("Mensaje")
    top.configure(bg="white")

    message_label = ttk.Label(top, text=mensaje, background="white")
    message_label.pack(pady=20,padx=20)

    top.bind("<Return>", on_enter_press)
    return top


################## Búsqueda de Clientes ##################

def BusquedaClientes(nombreClienteBuscado):
    """
    Entrada: 
        nombreCliente: nombre del cliente para el cual se va a calibrar
    Salida: una lista con el nombre del cliente, su dirección y el archivo donde está almacenada su información
    """
    woorkbookClientes = load_workbook(filename="Clientes/Clientes.xlsx") # Apertura del archivo de excel de clientes 
    hojaClientes = woorkbookClientes.active # Hoja del archivo de excel donde están los clientes y su información
    
    i = 3 # Se inicializa un contador que va a recorrer los números de fila, empezando por la fila 3 porque las filas anteriores son encabezados
    while hojaClientes["A"+str(i)].value != None: # Mientras la celda de cliente no esté vacía, se van a seguir recorriendo las filas
        if hojaClientes["A"+str(i)].value == nombreClienteBuscado: # Si el valor de la celda es igual al nombre del cliente que se busca:
            numFila = i # El número de fila donde se encuentra la información del cliente corresponde al valor actual del contador 
        i += 1

    # Se almacenan el variables la información del cliente:
    # La columna A corresponde a "Nombre del Cliente", B a "Dirección del Cliente" y C a "Nombre del archivo del cliente"
    # numFila permite crear la coordenada de la celda que almacena la información deseada
    nombreCliente = hojaClientes["A"+str(numFila)].value 
    direccionCliente = hojaClientes["B"+str(numFila)].value
    archivoCliente = "./Clientes/" + hojaClientes["C"+str(numFila)].value

    return nombreCliente, direccionCliente, archivoCliente

################## Creación de un archivo para la calibración ##################

def CrearArchivoCalibracion(numCertificado):
	# Se crea un duplicado del machote, nombrado con una marca temporal:
	machote_datos = "./Machotes/RegistroDatos.xlsx"
	machote_info = "./Machotes/CalibracionInfo.xlsx"

	archivoCalibracion_datos = "./Calibraciones en curso/" + numCertificado + "_Datos.xlsx" # Nombre del archivo para la calibración
	archivoCalibracion_info = "./Calibraciones en curso/" + numCertificado + "_Info.xlsx" # Nombre del archivo para la calibración

	shutil.copy(machote_datos, archivoCalibracion_datos) # Creación del duplicado del machote
	shutil.copy(machote_info, archivoCalibracion_info) # Creación del duplicado del machote

	return archivoCalibracion_datos, archivoCalibracion_info


################## Completa la información del archivo Info para la calibración ##################

def RellenarInfoCalibracion(nombreArchivo, lista_info):
    # Rellena la información brindada por la interfaz en un archivo de información
    workbookInfo = load_workbook(filename=nombreArchivo) #Abre el archivo
    hojaInfo = workbookInfo.active #Abre la hoja activa del archivo

    i = 2 # el registro empieza en la fila 2
    for elemento in lista_info:
        hojaInfo["B"+str(i)].value = elemento  # Se emplea la columa B para almacenar la información
        i+=1
    
    workbookInfo.save(nombreArchivo)
    workbookInfo.close()
    
    return




################## Autocompletado de la información que se tiene del cliente y la calibración ##################

def AutocompletarInformacionCliente(nombreCliente, direccionCliente, numeroCertificado, numeroSolicitud, identificacionCalibrando, 
                                    responsableCalibracion, responsableRevision, patron, materialPatron, seleccionSecuencia):
    # Lista con la información del cliente para el que se va a calibrar
    informacionCliente = BusquedaClientes(nombreCliente)
   
    # Carga del archivo de excel que contiene la información del cliente 
    workbookSolicitantes = load_workbook(filename="Clientes/" + informacionCliente[2], keep_vba = True, data_only = True)
    hojaJuego = workbookSolicitantes[identificacionCalibrando] # Selección de la hoja que contiene la información del juego a calibrar

    # Carga del archivo de excel creado para la calibración:
    workbookCalibracion = load_workbook(filename=CrearArchivoCalibracion(seleccionSecuencia, numeroCertificado))
    #Definir los nombres de las hojas del libro de excel en el que se está trabajando:
    hojaConversionDatos = workbookCalibracion["conversion datos"]
    hojaIdentificacionBloques = workbookCalibracion["Ident.Bloques a calibrar"]
    hojaResultadosCalibracion = workbookCalibracion["Introduccion de datos de Calib."]

    #Almacenamiento de la información del calibrando
    listaValores = [] #Se crea una lista para guardar los valores nominales de los bloques del juego seleccionado
    for filaValores in hojaJuego.iter_rows(min_row=12,
                                           min_col=2,
                                           max_col=2):
        for celdaValores in filaValores:
            listaValores.append(float(celdaValores.value))
        
    listaSeries = [] #Se crea una lista para guardar los valores de las series de los bloques del juego seleccionado
    for filaSeries in hojaJuego.iter_rows(min_row=12,
                                          min_col=3,
                                          max_col=3):
        for celdaSeries in filaSeries:
            listaSeries.append(int(celdaSeries.value))

    objetoCalibrar = hojaJuego["C2"].value #Extrae el objeto a calibrar del archivo del cliente
    marcaInstrumento = hojaJuego["C3"].value #Extrae la marca del instrumento del archivo del cliente
    serieInstrumento = hojaJuego["C4"].value #Extrae la serie del instrumento del archivo del cliente
    material = hojaJuego["C5"].value #Extrae el material del calibrando del cliente
    modelo = hojaJuego["C6"].value #Extrae el modelo del calibrando del cliente
    gradoDeclarado = hojaJuego["C7"].value #Extra el gradoDeclarado del calibrando del cliente

    #Copiar información de los valores nominales y series de los bloques del juego a la hoja de la calibración
    i = 12
    while i <= (len(listaValores)+11):
        #Copiar información de los valores nominales:
        coordenadaCeldaValor = "B"+str(i-8) #Coordenada de la celda donde se encuentra el valor nominal de cada bloque a lo largo del while
        hojaIdentificacionBloques[coordenadaCeldaValor] = listaValores[i-12] #Escritura del valor nominal del bloque en la hoja correspondiente a la calibración
        #Copiar información de las series:
        coordenadaCeldaSerie = "C"+str(i-8) #Coordenada de la celda donde se encuentra la serie de cada bloque a lo largo del while
        hojaIdentificacionBloques[coordenadaCeldaSerie] = listaSeries[i-12] #Escritura de la serie del bloque en la hoja correspondiente a la calibración

        i+=1

    #Autocompletado de la información del calibrando en la hoja de la calibración
    hojaConversionDatos["L6"] = numeroCertificado
    hojaConversionDatos["L9"] = nombreCliente
    hojaConversionDatos["L11"] = numeroSolicitud
    hojaConversionDatos["L12"] = direccionCliente

    hojaConversionDatos["L17"] = objetoCalibrar
    hojaConversionDatos["L19"] = marcaInstrumento
    hojaConversionDatos["L20"] = serieInstrumento
    hojaConversionDatos["L21"] = material 
    hojaConversionDatos["L22"] = modelo
    hojaConversionDatos["L23"] = gradoDeclarado

    hojaConversionDatos["L27"] = responsableCalibracion
    hojaConversionDatos["L28"] = responsableRevision

    hojaConversionDatos["L30"] = patron
    hojaConversionDatos["L31"] = materialPatron

    return workbookCalibracion, hojaResultadosCalibracion, hojaConversionDatos



def RellenarEncabezados(archivoCalibracion_datos, secuencia, numRepeticiones):
    
    workbookDatos = load_workbook(archivoCalibracion_datos) #Apertura del archivo de excel de la calibración
    hojaDatos = workbookDatos.active
   
    #Definir estilos 
    texto_negrita = Font(bold = True)
    texto_centrado = Alignment(horizontal = "center", vertical="center", wrapText=True)
    borde_sencillo = Side(border_style = "thin")
    borde_cuadrado = Border(top = borde_sencillo,
                            right = borde_sencillo,
                            bottom = borde_sencillo,
                            left = borde_sencillo)
    


    if str(secuencia).lower() == "desviación central":
        for k in range(int(numRepeticiones)):#Se usan dos columnas por cada repetición: una para el patrón y otra para el calibrando
            columnaActual = 14+2*k

            letraColumnaPatron = openpyxl.utils.cell.get_column_letter(columnaActual) #Obtener la letra de la columna en la que se está trabajando
            letraColumnaCalibrando = openpyxl.utils.cell.get_column_letter(columnaActual+1) #Obtener la letra de la columna de la derecha a la que se está trabajando
            #Coordenadas de las celdas del patrón y calibrando #repetición
            coordenadaEncabezadoPatron = letraColumnaPatron + "1"
            coordenadaEncabezadoCalibrando = letraColumnaCalibrando + "1"
            #Escribir los encabezados de las nuevas celdas:
            hojaDatos[coordenadaEncabezadoPatron] = "Patrón #"+str(k+1)
            hojaDatos[coordenadaEncabezadoCalibrando] = "Calibrando #"+str(k+1)
            #Darle formato a las nuevas celdas:
            hojaDatos[coordenadaEncabezadoPatron].font = texto_negrita
            hojaDatos[coordenadaEncabezadoPatron].alignment = texto_centrado
            hojaDatos[coordenadaEncabezadoPatron].border = borde_cuadrado
            hojaDatos[coordenadaEncabezadoCalibrando].font = texto_negrita
            hojaDatos[coordenadaEncabezadoCalibrando].alignment = texto_centrado
            hojaDatos[coordenadaEncabezadoCalibrando].border = borde_cuadrado
    
    elif str(secuencia).lower() == "desviación central y planitud":
        for k in range(int(numRepeticiones)):#Se usan siete columnas por cada repetición: una para el patrón y 6 para el calibrando
            columnaActual = 14+7*k

            letraColumnaPatron = openpyxl.utils.cell.get_column_letter(columnaActual) #Obtener la letra de la columna que va a guardar los datos del Patrón en cada rep k 
            letraColumnaCalibrandoCentro = openpyxl.utils.cell.get_column_letter(columnaActual+1) #Obtener la letra de la columna que va a guardar los datos del Centro del Calibrando en cada rep k 
            letraColumnaCalibrandoEsquina3 = openpyxl.utils.cell.get_column_letter(columnaActual+2) #Obtener la letra de la columna que va a guardar los datos de la Esquina 3 del Calibrando en cada rep k 
            letraColumnaCalibrandoEsquina4 = openpyxl.utils.cell.get_column_letter(columnaActual+3) #Obtener la letra de la columna que va a guardar los datos de la Esquina 4 del Calibrando en cada rep k 
            letraColumnaCalibrandoEsquina5 = openpyxl.utils.cell.get_column_letter(columnaActual+4) #Obtener la letra de la columna que va a guardar los datos de la Esquina 5 del Calibrando en cada rep k 
            letraColumnaCalibrandoEsquina6 = openpyxl.utils.cell.get_column_letter(columnaActual+5) #Obtener la letra de la columna que va a guardar los datos de la Esquina 3 del Calibrando en cada rep k 
            letraColumnaPatronRepetido = openpyxl.utils.cell.get_column_letter(columnaActual+6) #Obtener la letra de la columna que va a guardar los datos de la Esquina 3 del Calibrando en cada rep k 

            #Escribir los encabezados de las nuevas celdas:
            hojaDatos[letraColumnaPatron + "1"] = "Patrón (Centro) #"+str(k)
            hojaDatos[letraColumnaCalibrandoCentro + "1"] = "Calibrando (Centro) #"+str(k)
            hojaDatos[letraColumnaCalibrandoEsquina3 + "1"] = "Calibrando (Esquina 3) #"+str(k)
            hojaDatos[letraColumnaCalibrandoEsquina4 + "1"] = "Calibrando (Esquina 4) #"+str(k)
            hojaDatos[letraColumnaCalibrandoEsquina5 + "1"] = "Calibrando (Esquina 5) #"+str(k)
            hojaDatos[letraColumnaCalibrandoEsquina6 + "1"] = "Calibrando (Esquina 6) #"+str(k)
            hojaDatos[letraColumnaPatronRepetido + "1"] = "Patrón (Centro Repetido) #"+str(k)
            
            #Darle formato a las nuevas celdas:
            for numColumna in range(columnaActual,columnaActual+8):
                letraColumna = openpyxl.utils.cell.get_column_letter(numColumna)
                hojaDatos[letraColumna + "1"].font = texto_negrita
                hojaDatos[letraColumna + "1"].alignment = texto_centrado
                hojaDatos[letraColumna + "1"].border = borde_cuadrado
        
    workbookDatos.save(archivoCalibracion_datos)
    workbookDatos.close()



################## Cálculos del promedio y la desviación estándar ###################
###########
# OLD BOY #
###########
def CalculosDesviacionCentral(hojaResultadosCalibracion, numNuevasColumnas, numRepeticiones):

    #Calcular el promedio de la diferencia entre el patrón y el calibrando con fórmulas en Excel
    #Calcular la desviación estándar del promedio de la diferencia entre el patrón y el calibrando con fórmulas en Excel
    
    l = 2 #Se inicializa el contador para filas

    while hojaResultadosCalibracion["S"+str(l)].value != None:
        listaDiferencias = [] #Inicializamos una lista para guardar los strings del cálculo de las diferencias, ej: A3-A2
        j = 19 #Se vuelve a inicializar el contador para las columnas en S
        k = 1 #Se vuelve a inicializar el contador para las repeticiones
        while j <= (19+numNuevasColumnas)-1 and k <= numRepeticiones:
            letraColumnaPatron = openpyxl.utils.cell.get_column_letter(j) #Obtener la letra de la columna en la que se está trabajando
            letraColumnaCalibrando = openpyxl.utils.cell.get_column_letter(j+1) #Obtener la letra de la columna de la derecha a la que se está trabajando
            #Coordenadas de las celdas del patrón y calibrando #repetición: tomando en cuenta el cambio de fila
            coordenadaPatron = letraColumnaPatron + str(l)
            coordenadaCalibrando = letraColumnaCalibrando + str(l)
            stringDiferencia = coordenadaCalibrando + "-" + coordenadaPatron
            #Se agrega el string para calcular la diferencia a listaDiferencias
            listaDiferencias.append(stringDiferencia)
            k += 1
            j += 2 
        #Se crea un string para que funcione como el argumento de la fórmula del promedio y desviación estándar a partir de listaDiferencias
        argumentoFormulaCentral = ""
        for stringDif in range(len(listaDiferencias)):
            if stringDif < len(listaDiferencias)-1:
                argumentoFormulaCentral = argumentoFormulaCentral + listaDiferencias[stringDif] + ";"
            elif stringDif == len(listaDiferencias)-1:
                argumentoFormulaCentral = argumentoFormulaCentral + listaDiferencias[stringDif]           
        
        #Se obtienen las fórmulas predefinidas en la hoja de excel
        formulaPromedio = hojaResultadosCalibracion["E"+str(l)].value
        formulaDesvst = hojaResultadosCalibracion["F"+str(l)].value
        
        #Se modifican las fórmulas agregando el argumento construído
        promedioModif = formulaPromedio.replace("))",f"{argumentoFormulaCentral}))")
        DesvstModif = formulaDesvst.replace("))",f"{argumentoFormulaCentral}))")
        
        #Se actualizan las celdas con las fórmulas modificadas
        hojaResultadosCalibracion["E"+str(l)] = promedioModif
        hojaResultadosCalibracion["F"+str(l)] = DesvstModif
        
        l += 1
    return




def CalculosDesviacionYPlanitud(hojaResultadosCalibracion, numNuevasColumnas, numRepeticiones):
    l = 2 #Se inicializa el contador para filas -> Empezamos a agregar valores en la fila 2

    while hojaResultadosCalibracion["Y"+str(l)].value != None:
        listaDiferenciasCentros = [] #Inicializamos una lista para guardar el string para calcular la diferencia entre el centro del calibrando y el patrón en cada repetición
        listaDiferenciasEsquina3 = [] #Inicializamos una lista para guardar el string para calcular la diferencia entre la esquina 3 del calibrando y el centro del patrón en cada repetición
        listaDiferenciasEsquina4 = [] #Inicializamos una lista para guardar el string para calcular la diferencia entre la esquina 4 del calibrando y el centro del patrón en cada repetición
        listaDiferenciasEsquina5 = [] #Incializamos una lista para guardar el string para calcular la diferencia entre la esquina 5 del calibrando y el centro del patrón en cada repetición
        listaDiferenciasEsquina6 = [] #Incializamos una lista para guardar el string para calcular la diferencia entre la esquina 6 del calibrando y el centro del patrón en cada repetición

        j = 25 #Se vuelve a inicializar el contador para las columnas
        k = 1 #Se vuelve a inicializar el contador para las repeticiones
        while j <= (25+numNuevasColumnas)-1 and k <= numRepeticiones:
            letraColumnaPatron = openpyxl.utils.cell.get_column_letter(j) #Obtener la letra de la columna en la que se encuentra la medicón del centro del patrón
            letraColumnaCalibrando = openpyxl.utils.cell.get_column_letter(j+1) #Obtener la letra de la columna en la que se encuentra el centro del calibrando
            letraColumnaEsquina3 = openpyxl.utils.cell.get_column_letter(j+2) #Obtener la letra de la columna en la que se encuentra la medición de la esquina 3 del calibrando
            letraColumnaEsquina4 = openpyxl.utils.cell.get_column_letter(j+3) #Obtener la letra de la columna en la que se encuentra la medición de la esquina 4 del calibrando
            letraColumnaEsquina5 = openpyxl.utils.cell.get_column_letter(j+4) #Obtener la letra de la columna en la que se encuentra la medición de la esquina 5 del calibrando
            letraColumnaEsquina6 = openpyxl.utils.cell.get_column_letter(j+5) #Obtener la letra de la columna en la que se encuentra la medición de la esquina 6 del calibrando
        
            #Coordenadas de las celdas del patrón y calibrando #repetición: tomando en cuenta el cambio de fila
            coordenadaPatron = letraColumnaPatron + str(l)
            coordenadaCalibrando = letraColumnaCalibrando + str(l)
            coordenadaEsquina3 = letraColumnaEsquina3 + str(l)
            coordenadaEsquina4 = letraColumnaEsquina4 + str(l)
            coordenadaEsquina5 = letraColumnaEsquina5 + str(l)
            coordenadaEsquina6 = letraColumnaEsquina6 + str(l)

            #String de la diferencia entre el calibrando y el centro del patrón
            diferenciaCentros = coordenadaCalibrando + "-" + coordenadaPatron

            #Strings de las diferencias entre las esquinas del calibrando y el centro del patrón
            diferenciaEsquina3 = coordenadaEsquina3 + "-" + coordenadaPatron
            diferenciaEsquina4 = coordenadaEsquina4 + "-" + coordenadaPatron
            diferenciaEsquina5 = coordenadaEsquina5 + "-" + coordenadaPatron
            diferenciaEsquina6 = coordenadaEsquina6 + "-" + coordenadaPatron

            #Se agrega el valor de las diferencias a sus respectivas listas:
            listaDiferenciasCentros.append(diferenciaCentros)
            listaDiferenciasEsquina3.append(diferenciaEsquina3)
            listaDiferenciasEsquina4.append(diferenciaEsquina4)
            listaDiferenciasEsquina5.append(diferenciaEsquina5)
            listaDiferenciasEsquina6.append(diferenciaEsquina6)

            k += 1
            j += 6

        #Formula para crear un string para que funcione como el argumento de la fórmulas a partir de las listas de diferencias
        def argumentoFormula(listaDiferencias):
            argumentoFormulaCentral = ""
            for stringDif in range(len(listaDiferencias)):
                if stringDif < len(listaDiferencias)-1:
                    argumentoFormulaCentral = argumentoFormulaCentral + listaDiferencias[stringDif] + ";"
                elif stringDif == len(listaDiferencias)-1:
                    argumentoFormulaCentral = argumentoFormulaCentral + listaDiferencias[stringDif]
            return argumentoFormulaCentral

        #Se crean los strings de los argumentos para las fórmulas
        argumentoPromedioCentros = argumentoFormula(listaDiferenciasCentros)
        argumentoPromedioEsquina3 = argumentoFormula(listaDiferenciasEsquina3)
        argumentoPromedioEsquina4 = argumentoFormula(listaDiferenciasEsquina4)
        argumentoPromedioEsquina5 = argumentoFormula(listaDiferenciasEsquina5)
        argumentoPromedioEsquina6 = argumentoFormula(listaDiferenciasEsquina6)

        #Se obtienen las fórmulas predefinidas en la hoja de excel (ya en el excel debe estar la fórmula con el argumento vacío)
        formulaPromedioCentros = hojaResultadosCalibracion["K"+str(l)].value
        formulaDesvstCentros = hojaResultadosCalibracion["L"+str(l)].value
        formulaPromedioEsquina3 = hojaResultadosCalibracion["E"+str(l)].value
        formulaPromedioEsquina4 = hojaResultadosCalibracion["F"+str(l)].value
        formulaPromedioEsquina5 = hojaResultadosCalibracion["G"+str(l)].value
        formulaPromedioEsquina6 = hojaResultadosCalibracion["H"+str(l)].value
        
        #Se modifican las fórmulas agregando el argumento construído
        promedioModifCentros = formulaPromedioCentros.replace("))",f"{argumentoPromedioCentros}))")
        DesvstModifCentros = formulaDesvstCentros.replace("))",f"{argumentoPromedioCentros}))")
        promedioModifEsquina3 = formulaPromedioEsquina3.replace("))",f"{argumentoPromedioEsquina3}))")
        promedioModifEsquina4 = formulaPromedioEsquina4.replace("))",f"{argumentoPromedioEsquina4}))")
        promedioModifEsquina5 = formulaPromedioEsquina5.replace("))",f"{argumentoPromedioEsquina5}))")
        promedioModifEsquina6 = formulaPromedioEsquina6.replace("))",f"{argumentoPromedioEsquina6}))")
        
        #Se actualizan las celdas con las fórmulas modificadas
        hojaResultadosCalibracion["K"+str(l)] = promedioModifCentros
        hojaResultadosCalibracion["L"+str(l)] = DesvstModifCentros
        hojaResultadosCalibracion["E"+str(l)] = promedioModifEsquina3
        hojaResultadosCalibracion["F"+str(l)] = promedioModifEsquina4
        hojaResultadosCalibracion["G"+str(l)] = promedioModifEsquina5
        hojaResultadosCalibracion["H"+str(l)] = promedioModifEsquina6

        l += 1

    return
    
################## Calibración de Bloque ##################

def procesoCalibracion(RPi_url, archivoCalibracion_datos, secuencia, bloqueID, valorNominal, tInicial, tEstabilizacion, numReps, plantilla):

    workbookDatos = load_workbook(archivoCalibracion_datos) #Apertura del archivo de excel de la calibración
    hojaDatos = workbookDatos.active

    #Eliminar Fila si existe el registro y devuelve la fila
    #numFila = selectorFila(hojaDatos) # Elige la fila correspondiente
    numFila = seleccionarFilaSegunID(hojaDatos,bloqueID) # Elige la fila correspondiente

    hojaDatos[f"A{numFila}"] = valorNominal
    hojaDatos[f"B{numFila}"] = bloqueID
    hojaDatos[f"C{numFila}"] = numReps

    sleep(float(tInicial)*60)

    condAmb=condicionesAmbientales(RPi_url, instrumento="fluke")
    condAmb.append(condicionesAmbientales(RPi_url, instrumento="vaisala"))

    listaMedicionesBloque = ejecutarSecuencia(RPi_url, secuencia, tInicial, tEstabilizacion, numReps, plantilla)
    
    if listaMedicionesBloque:
        listaMedicionesBloque = listaMedicionesBloque[0]
    else:
        return

    condAmb += condicionesAmbientales(RPi_url, instrumento="fluke")
    condAmb.append(condicionesAmbientales(RPi_url, instrumento="vaisala"))

# Guardado de datos ambientales
    for i,columna in enumerate(hojaDatos.iter_cols(
                                                    min_row=numFila,
                                                    max_row=numFila,
                                                    min_col=4,
                                                    max_col=13)):
        for cell in columna:
            cell.value = condAmb[i]
        
# Guardado de datos de medición de bloques
    for i,columna in enumerate(hojaDatos.iter_cols(
                                                    min_row=numFila,
                                                    max_row=numFila,
                                                    min_col=14,
                                                    max_col=13+len(listaMedicionesBloque))):
        for cell in columna:
            cell.value = listaMedicionesBloque[i]
    
    workbookDatos.save(archivoCalibracion_datos)
    workbookDatos.close()

    outputString = f"Calibración del Bloque {bloqueID} Finalizada.\n\nResultados de la medición:\n"
    if secuencia.lower() == "desviación central":
        outputString += "No.  Patrón   Calibrando\n"
        for i in range(int(numReps)):
            outputString += str(i+1)+f"   {listaMedicionesBloque[i*2]}   {listaMedicionesBloque[i*2+1]}\n"
    
    elif secuencia.lower() == "desviación central y planitud":
        outputString += "No.  Patrón  Cal-C  Cal-E1  Cal-E2  Cal-E3  Cal-E4  Patrón\n"
        for i in range(int(numReps)):
            outputString += str(i+1)+f"  -  {listaMedicionesBloque[i*7]}  -   {listaMedicionesBloque[i*7+1]}  -  {listaMedicionesBloque[i*7+2]}  -  {listaMedicionesBloque[i*7+3]}  -  {listaMedicionesBloque[i*7+4]}  -  {listaMedicionesBloque[i*7+5]}  -  {listaMedicionesBloque[i*7+6]}\n"

    mostrarMensaje(outputString)
    return 






################## Selector fila para la hoja de resultados ##################

def selectorFila(hojaResultadosCalibracion):
    i = 2 # Se inicializa el contador en 2 porque la fila 1 tiene los encabezados 
    for filaValorNominal in hojaResultadosCalibracion.iter_rows(min_row=2,
                                                                max_row=500,
                                                                min_col=1,
                                                                max_col=1):
        for celdaValorNominal in filaValorNominal:
            if celdaValorNominal.value == None:
                numFila = i
            else:
                i += 1
    return numFila

def seleccionarFilaSegunID(hojaResultadosCalibracion, bloqueID):
    salir = 0
    i = 2 # Se inicializa el contador en 2 porque la fila 1 tiene los encabezados 
    for filaValorNominal in hojaResultadosCalibracion.iter_rows(min_row=2,
                                                                max_row=500,
                                                                min_col=2,
                                                                max_col=2):
        for celdaValorNominal in filaValorNominal:
            if celdaValorNominal.value == bloqueID:
                numFila = filaValorNominal[0].row
                salir = 1
                hojaResultadosCalibracion.delete_rows(filaValorNominal[0].row, 1)
            elif celdaValorNominal.value == None:
                numFila = filaValorNominal[0].row
                salir = 1
            else:
                i += 1
        
        if salir:
            break
    return numFila
################## Eliminar archivo ##################

def EliminarArchivo(rutaArchivoEliminar):
    #Revisar si el archivo existe
    if os.path.exists(rutaArchivoEliminar):
        #Borrar el archivo
        os.remove(rutaArchivoEliminar)
    else:
        mostrarMensaje("El archivo indicado no existe.")
    return 


################## Eliminar archivo ##################
def unificarArchivos(rutaDatos, rutaMacro):
    
    machoteMacro = "./Machotes/MacroFinal.xlsm"
    shutil.copy(machoteMacro, rutaMacro)	

    workbookDatos = load_workbook(filename=rutaDatos)
    workbookMacro = load_workbook(filename=rutaMacro,keep_vba=True)
    
    # Traspaso de Info

    for ws_name in ["Datos", "Información"]:
        if ws_name in workbookDatos.sheetnames:
            
            sheetDatos = workbookDatos[ws_name]
            
            sheetMacro = workbookMacro.create_sheet(title=f"Registro{ws_name}")
            
            # Copy contenido
            for row in sheetDatos.iter_rows():
                for cell in row:

                    new_cell = sheetMacro.cell(row=cell.row, column=cell.column, value=cell.value)

                    if cell.has_style:
                        new_cell.font = copy(cell.font)
                        new_cell.border = copy(cell.border)
                        new_cell.fill = copy(cell.fill)
                        new_cell.number_format = copy(cell.number_format)
                        new_cell.alignment = copy(cell.alignment)
    
    
    
    workbookDatos.close()
    
    workbookMacro.save(rutaMacro)
    workbookMacro.close()
    return

    
################## Reanudar Calibración ##################

def obtenerInfoCalibracion(numCertificado):
     
    info = [f"./Calibraciones en curso/{numCertificado}_Datos.xlsx"] #Resultados de la información, siguen la forma (archivoCalibracion_datos, cliente, certificado, solicitud, idCalibrando, responsable, revision, patron, material, secuencia, tInicial, tEstabilizacion, numReps)

    workbookInfo = load_workbook(filename=f"./Calibraciones en curso/{numCertificado}_Info.xlsx")  #Apertura del archivo de excel de la calibración
    hojaInfo = workbookInfo.active

    for fila in hojaInfo.iter_rows(min_row=2, max_row=14, min_col=2, max_col=2):
        for celda in fila:
                info.append(celda.value)
    workbookInfo.close()

    return info

def obtenerUnidadesCalibrando(cliente, idCalibrando):
    archivoCliente = BusquedaClientes(cliente)[2]
    workbookCliente = load_workbook(filename=archivoCliente)  #Apertura del archivo de excel del cliente

    #Revisar si ya existe algún calibrando registrado con el mismo número de serie
    existeCalibrando = False
    if idCalibrando in workbookCliente.sheetnames:
        existeCalibrando = True

    if not existeCalibrando:
        mostrarMensaje("El ID de calibrando no existe. \nPor favor registrar calibrando.")
        return
    
    hojaCalibrando = workbookCliente[idCalibrando]

    unidad = hojaCalibrando["Z1"]

    return unidad

################## Agregar cliente ##################

def AgregarCliente(nombreCliente, direccionCliente):
    """
    Esta función permite agregar el nombre y la dirección de un nuevo cliente al archivo de Clientes.
    Además, crea el archivo del nuevo cliente donde se encuentra la información de sus juegos de bloques.
    """
    workbookClientes = load_workbook(filename="Clientes/Clientes.xlsx") #Apertura del archivo de excel de clientes 
    hojaClientes = workbookClientes.active #Hoja del archivo de excel donde están los clientes y su información
    
    i = 3 #Se inicializa el contador para filas en 3 porque en la fila 1 y 2 están los encabezados
    #Ahora se deben recorrer las filas, empezando por la fila 3 para determinar el número de la fila que está libre para incluir un nuevo cliente
    for filaValorNominal in hojaClientes.iter_rows(min_row=3,
                                                    min_col=1,
                                                    max_col=1):
        for celdaValorNominal in filaValorNominal:
            if celdaValorNominal.value == None: #Ve si existe algún dato
                numFila = i
            elif celdaValorNominal.value == nombreCliente:
                mostrarMensaje("Existe un Cliente con el mismo nombre.\nPor favor, ingresar otro nombre o corroborar \nque es el mismo cliente.")
                return 0
            else:
                i += 1 

    machoteCliente = "./Machotes/Machote para nuevo cliente.xlsx"
    nombreArchivoCliente = nombreCliente + ".xlsx" #El nombre del archivo de Excel va a ser igual al nombre del Cliente
    shutil.copy(machoteCliente, "./Clientes/" + nombreArchivoCliente)	

    #Se agrega la información del cliente al archivo de Clientes
    hojaClientes["A"+str(i)] = nombreCliente
    hojaClientes["B"+str(i)] = direccionCliente
    hojaClientes["C"+str(i)] = nombreArchivoCliente

    workbookClientes.save("./Clientes/Clientes.xlsx")
    workbookClientes.close()
    mostrarMensaje("El cliente se ha registrado con éxito.")
    return 1





################## Ingresar juego de bloques/calibrando ##################

def IngresarCalibrando(nombreCliente, objeto, cantidad, marca, numSerie, material, modelo, grado, identificacionInterna, unidad):
    archivoCliente = BusquedaClientes(nombreCliente)[2] #Busqueda del archivo del cliente
    workbookCliente = load_workbook(filename=archivoCliente)  #Apertura del archivo de excel del cliente

    #Revisar si ya existe algún calibrando registrado con el mismo número de serie
    existeCalibrando = False
    for serieCalibrandoRegistrado in workbookCliente.sheetnames:
        if serieCalibrandoRegistrado == numSerie:
            existeCalibrando = True
            break

    if existeCalibrando:
        mostrarMensaje("Ya existe un calibrando registrado con el númerio de serie " + numSerie + ".")
        isOK = 0
        return

    #Crear una hoja para el nuevo calibrando
    hojaReferencia = workbookCliente.worksheets[0] #Se selecciona la hoja 1 como una referencia para crear la hoja para el nuevo juevo
    hojaNuevoCalibrando = workbookCliente.copy_worksheet(hojaReferencia)
    hojaNuevoCalibrando.title = numSerie

    
    #Agregar la información del calibrando al archivo del cliente
    hojaNuevoCalibrando["A1"] = "Información del calibrando con identificación " + numSerie
    hojaNuevoCalibrando["C2"] = objeto
    hojaNuevoCalibrando["C3"] = int(cantidad)
    hojaNuevoCalibrando["C4"] = marca
    hojaNuevoCalibrando["C5"] = numSerie
    hojaNuevoCalibrando["C6"] = material
    hojaNuevoCalibrando["C7"] = modelo
    try:
        hojaNuevoCalibrando["C8"] = int(grado)
    except:
        hojaNuevoCalibrando["C8"] = grado
    hojaNuevoCalibrando["C9"] = identificacionInterna
    hojaNuevoCalibrando["Z1"] = unidad

    hojaNuevoCalibrando["B12"] = "Longitud nominal (" + unidad + ")" #Agregar valor nominal e identificación de los bloques del juego
    workbookCliente.save(archivoCliente)
    workbookCliente.close()
    mostrarMensaje("Se han ingresado exitosamente los datos del nuevo calibrando.\nPor favor ingresar los datos de los bloques correspondientes.")
    return




def ingresarBloque(top, cliente, numSerie, unidad, longitudNominal_entry, idBloque_entry, ultimoBloqueDatos_label):
    valorBloqueIngresar = longitudNominal_entry.get()
    idBloqueIngresar = idBloque_entry.get()

    #Se agrega la información del bloque a la hoja
    archivoCliente = BusquedaClientes(cliente)[2] #Busqueda del archivo del cliente
    workbookCliente = load_workbook(filename=archivoCliente)  #Apertura del archivo de excel del cliente
    hojaCalibrando = workbookCliente[numSerie]

    numFila, i = 14,14 #Se inicializa el contador para filas en 14
    for fila in hojaCalibrando.iter_rows(min_row=14, max_row=500, min_col=1, max_col=1):
        for celda in fila:
            if celda.value == None: #Ve si existe algún dato
                numFila = i
            else:
                i += 1    

    hojaCalibrando["A"+str(numFila)] = numFila - 13
    hojaCalibrando["B"+str(numFila)] = valorBloqueIngresar
    hojaCalibrando["C"+str(numFila)] = idBloqueIngresar

    workbookCliente.save(archivoCliente)
    workbookCliente.close()

    ultimoBloqueDatos_label.config(text=f"{numFila - 13}) {valorBloqueIngresar} {unidad} - ID:{idBloqueIngresar}", background="white")
    # ultimoBloqueDatos_label = ttk.Label(top, text=f"{numFila - 13}) {valorBloqueIngresar} {unidad} - ID:{idBloqueIngresar}", background="white")
    # ultimoBloqueDatos_label.grid(row=40, column=10, columnspan=1, pady=10)

    return

def finalizar_agregar(workbookCliente,archivoCliente):
    workbookCliente.save(archivoCliente)
    workbookCliente.close()
    mostrarMensaje("Se han ingresado exitosamente los datos del nuevo calibrando.")
    top.destroy()
    return

################## Abrir Windows Explorer ##################

def abrir_explorador(ruta):
    try:
        subprocess.Popen(['explorer', ruta])
    except FileNotFoundError:
        print("No se encontró el comando 'explorer'.")




################## Ocultar advertencias en terminal ##################

def fxn():
    warnings.warn("deprecated", DeprecationWarning)

with warnings.catch_warnings():
    warnings.simplefilter("ignore")
    fxn()