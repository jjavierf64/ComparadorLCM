"""
En este archivo se presenta el código de una interfaz simple para el comparador de bloques TESA
"""

################## Importación de librerías ##################
import tkinter as tk
from tkinter import ttk
from ttkthemes import ThemedStyle
from PIL import Image, ImageTk
import os
import openpyxl
import requests

from FuncionesWindows import *

################## Definición variables globales ##################

cliente_combobox = None
certificado_entry = None
solicitud_entry = None
idCalibrando_entry = None
responsable_entry = None
revision_entry = None
patron_combobox = None
material_combobox = None
secuencia_combobox = None
tInicial_entry = None
tEstabilizacion_entry = None
numReps_entry = None
certificado_combobox = None
nuevoCliente_entry = None
contactoCliente_entry = None
objeto_entry = None
marca_entry = None
numSerie_entry = None
materialCalibrando_combobox = None
modelo_entry = None
grado_entry = None
unidad_combobox = None

# Dirección del servidor por defecto
global RPi_url
RPi_url = "http://192.168.196.100:5000/" # Zerotier




################## Definición de funciones de la interfaz ##################

def nueva_calibracion():
    # Ocultar la ventana del menú de opciones una vez que se selecciona una opción
    root.withdraw()

    # Crear una nueva ventana
    ventana_nuevaCalibracion = tk.Toplevel(root)
    ventana_nuevaCalibracion.title("Nueva Calibración")
    ventana_nuevaCalibracion.configure(bg="white")
    ventana_nuevaCalibracion.protocol("WM_DELETE_WINDOW", lambda: regresarVentanaPrincipal(root, ventana_nuevaCalibracion)) #Cuando se cierre la ventana secundaria, vuelva al menú de opciones
    ventana_nuevaCalibracion.iconphoto(False, winIcono)

    # Crear un nuevo layout para la ventana de Nueva Calibración
    title_label = ttk.Label(ventana_nuevaCalibracion, text="Comparador de bloques TESA", font=("Helvetica", 16, "bold"),
                            background="white")
    title_label.grid(row=0, column=0, columnspan=2, pady=20)

    subtitle_label = ttk.Label(ventana_nuevaCalibracion, text="Nueva calibración", font=("Helvetica", 14),
                               background="white")
    subtitle_label.grid(row=1, column=0, columnspan=2, pady=10)

    image = Image.open("./assets/logoLCM.png")
    image = image.resize((int(image.width * 0.25), int(image.height * 0.25)))  # Ajustar el tamaño del logo
    image = ImageTk.PhotoImage(image)

    image_label = ttk.Label(ventana_nuevaCalibracion, image=image, background="white")
    image_label.image = image
    image_label.grid(row=0, column=2, rowspan=1, padx=10, pady=10)

    global cliente_combobox, certificado_entry, solicitud_entry, idCalibrando_entry, responsable_entry, revision_entry
    global revision_entry, patron_combobox, material_combobox, secuencia_combobox, tInicial_entry, tEstabilizacion_entry, numReps_entry

    # Crear una lista con los nombres de los clientes ya registrados
    clientesRegistrados = []
    archivoClientes = openpyxl.load_workbook("Clientes/Clientes.xlsx")
    hojaClientes = archivoClientes.active

    numFila = 3  # Se empieza en la fila 3 porque antes están los encabezados
    for fila in hojaClientes.iter_rows(min_row=3,
                                       min_col=1,
                                       max_col=1):
        for celda in fila:
            if celda.value != None:
                clientesRegistrados.append(celda.value)
    archivoClientes.close()

    # Espacios para ingresar las variables requeridas para una nueva calibración
    cliente_label = ttk.Label(ventana_nuevaCalibracion, text="Nombre del cliente:", anchor=tk.CENTER,
                              background="white")
    cliente_label.grid(row=2, column=0, pady=5, sticky=tk.EW)
    cliente_combobox = ttk.Combobox(ventana_nuevaCalibracion, values=clientesRegistrados, width=40 ,state= "readonly")
    cliente_combobox.grid(row=2, column=1, columnspan=2, pady=5, padx=(20, 5), sticky="ew")

    certificado_label = ttk.Label(ventana_nuevaCalibracion, text="Número de certificado:", background="white")
    certificado_label.grid(row=3, column=0, pady=5)
    certificado_entry = ttk.Entry(ventana_nuevaCalibracion, width=42)
    certificado_entry.grid(row=3, column=1, columnspan=2, pady=5, padx=(20, 5))

    solicitud_label = ttk.Label(ventana_nuevaCalibracion, text="Número de solicitud:", background="white")
    solicitud_label.grid(row=4, column=0, pady=5)
    solicitud_entry = ttk.Entry(ventana_nuevaCalibracion, width=42)
    solicitud_entry.grid(row=4, column=1, columnspan=2, pady=5, padx=(20, 5))

    idCalibrando_label = ttk.Label(ventana_nuevaCalibracion, text="Identificación del calibrando (No. Serie):", background="white")
    idCalibrando_label.grid(row=5, column=0, pady=5)
    idCalibrando_entry = ttk.Entry(ventana_nuevaCalibracion, width=42)
    idCalibrando_entry.grid(row=5, column=1, columnspan=2, pady=5, padx=(20, 5))

    responsable_label = ttk.Label(ventana_nuevaCalibracion, text="Responsable de la calibración:", background="white")
    responsable_label.grid(row=6, column=0, pady=5)
    responsable_entry = ttk.Entry(ventana_nuevaCalibracion, width=42)
    responsable_entry.grid(row=6, column=1, columnspan=2, pady=5, padx=(20, 5))

    revision_label = ttk.Label(ventana_nuevaCalibracion, text="Responsable de la revisión:", background="white")
    revision_label.grid(row=7, column=0, pady=5)
    revision_entry = ttk.Entry(ventana_nuevaCalibracion, width=42)
    revision_entry.grid(row=7, column=1, columnspan=2, pady=5, padx=(20, 5))

    patron_label = ttk.Label(ventana_nuevaCalibracion, text="Patrón a utilizar:", anchor=tk.CENTER, background="white")
    patron_label.grid(row=8, column=0, pady=5, sticky=tk.EW)
    patron_combobox = ttk.Combobox(ventana_nuevaCalibracion, values=["Bloques Patrón de Cerámica de 0,05\" a 4\"",
                                                                     "Bloques Patrón de Cerámica de 0,5 mm a 100 mm"],
                                   width=40 ,state= "readonly")
    patron_combobox.grid(row=8, column=1, columnspan=2, pady=5, padx=(20, 5), sticky="ew")

    material_label = ttk.Label(ventana_nuevaCalibracion, text="Material de los bloques patrón: ", anchor=tk.CENTER,
                               background="white")
    material_label.grid(row=9, column=0, pady=5, sticky=tk.EW)
    material_combobox = ttk.Combobox(ventana_nuevaCalibracion,
                                     values=["Patrón en acero", "Patrón en Tungsteno", "Patrón en cerámica",
                                             "Patrón en cromo"], width=40 ,state= "readonly")
    material_combobox.grid(row=9, column=1, columnspan=2, pady=5, padx=(20, 5), sticky="ew")

    secuencia_label = ttk.Label(ventana_nuevaCalibracion, text="Secuencia de calibración:", anchor=tk.CENTER,
                                background="white")
    secuencia_label.grid(row=10, column=0, pady=5, sticky=tk.EW)
    secuencia_combobox = ttk.Combobox(ventana_nuevaCalibracion,
                                      values=["Desviación central", "Desviación central y planitud"], width=40 ,state= "readonly")
    secuencia_combobox.grid(row=10, column=1, columnspan=2, pady=5, padx=(20, 5), sticky="ew")

    tInicial_label = ttk.Label(ventana_nuevaCalibracion, text="Tiempo inicial (en minutos):", background="white")
    tInicial_label.grid(row=11, column=0, pady=5)
    tInicial_entry = ttk.Entry(ventana_nuevaCalibracion, width=42)
    tInicial_entry.grid(row=11, column=1, columnspan=2, pady=5, padx=(20, 5))

    tEstabilizacion_label = ttk.Label(ventana_nuevaCalibracion, text="Tiempo de estabilización (en segundos):",
                                      background="white")
    tEstabilizacion_label.grid(row=12, column=0, pady=5)
    tEstabilizacion_entry = ttk.Entry(ventana_nuevaCalibracion, width=42)
    tEstabilizacion_entry.grid(row=12, column=1, columnspan=2, pady=5, padx=(20, 5))

    numReps_label = ttk.Label(ventana_nuevaCalibracion, text="Número de repeticiones:", background="white")
    numReps_label.grid(row=13, column=0, pady=5)
    numReps_entry = ttk.Entry(ventana_nuevaCalibracion, width=42)
    numReps_entry.grid(row=13, column=1, columnspan=2, pady=5, padx=(20, 5))

    #motores_button = ttk.Button(ventana_nuevaCalibracion, text="Posicionar Motores", command=mover_motores)
    #motores_button.grid(row=14, column=0, columnspan=1, pady=10)

    continuar_button = ttk.Button(ventana_nuevaCalibracion, text="Continuar", command=lambda: continuarNuevaCalibracion(ventana_nuevaCalibracion))
    continuar_button.grid(row=14, column=2, columnspan=1, pady=10)

    regresar_button = ttk.Button(ventana_nuevaCalibracion, text="Regresar al menú de opciones",
                                 command=lambda: regresarVentanaPrincipal(root, ventana_nuevaCalibracion))
    regresar_button.grid(row=15, column=0, columnspan=1, pady=10)
    return


def reanudar_calibracion():
    # Ocultar la ventana del menú de opciones una vez que se selecciona una opción
    root.withdraw()

    # Crear una nueva ventana
    ventana_reanudar = tk.Toplevel(root)
    ventana_reanudar.title("Reanudar Calibración")
    ventana_reanudar.configure(bg="white")
    ventana_reanudar.protocol("WM_DELETE_WINDOW", lambda: regresarVentanaPrincipal(root, ventana_reanudar)) #Cuando se cierre la ventana secundaria, vuelva al menú de opciones
    ventana_reanudar.iconphoto(False, winIcono)

    # Crear un nuevo layout para la ventana de reanudar calibración
    title_label = ttk.Label(ventana_reanudar, text="Comparador de bloques TESA", font=("Helvetica", 16, "bold"),
                            background="white")
    title_label.grid(row=0, column=0, columnspan=2, pady=20)

    subtitle_label = ttk.Label(ventana_reanudar, text="Reanudar calibración", font=("Helvetica", 14),
                               background="white")
    subtitle_label.grid(row=1, column=0, columnspan=2, pady=10)

    image = Image.open("./assets/logoLCM.png")
    image = image.resize((int(image.width * 0.25), int(image.height * 0.25)))  # Ajustar el tamaño del logo
    image = ImageTk.PhotoImage(image)

    image_label = ttk.Label(ventana_reanudar, image=image, background="white")
    image_label.image = image
    image_label.grid(row=0, column=2, rowspan=1, padx=10, pady=10)

    global certificado_combobox, tInicial_entry, tEstabilizacion_entry

    # Se crea una lista con el nombre de los documentos que se encuentran en la carpeta de "Calibraciones en curso"
    calibracionesEnCurso = []
    for archivo in os.listdir("./Calibraciones en curso/"):
        if os.path.isfile(os.path.join("./Calibraciones en curso/", archivo)):
            nombreArchivo, extension = os.path.splitext(archivo)
            if nombreArchivo[-5:]=="_Info":
                calibracionesEnCurso.append(nombreArchivo[:-5])

    # Espacios para ingresar las variables requeridas para reanudar calibración
    certificado_label = ttk.Label(ventana_reanudar, text="Seleccione la calibración a reanudar:", anchor=tk.CENTER,
                                  background="white")
    certificado_label.grid(row=2, column=0, pady=5, sticky=tk.EW)
    certificado_combobox = ttk.Combobox(ventana_reanudar, values=calibracionesEnCurso, width=40 ,state= "readonly")
    certificado_combobox.grid(row=2, column=1, columnspan=2, pady=5, padx=(20, 5), sticky="ew")

    continuar_button = ttk.Button(ventana_reanudar, text="Reanudar calibración", command=lambda: reanudarCalibracion(ventana_reanudar))
    continuar_button.grid(row=5, column=0, columnspan=1, pady=10)

    regresar_button = ttk.Button(ventana_reanudar, text="Regresar al menú de opciones",
                                 command=lambda: regresarVentanaPrincipal(root, ventana_reanudar))
    regresar_button.grid(row=6, column=0, columnspan=1, pady=10)
    return



def calibracion_abierta(ventanaPrevia, archivoCalibracion_datos, cliente, certificado, solicitud, idCalibrando, responsable, revision, patron, material, secuencia, tInicial, tEstabilizacion, numReps, unidad):
    
    # definir función para detección automática de plantilla
    def detect_plantilla(event):
        bloqueID_oculto, valorNominal_oculto, unidad_oculto = event.widget.get().split()
        seleccionarPlantilla_combobox.current(0)

        if unidad_oculto == "mm" and float(valorNominal_oculto) > 10:
            seleccionarPlantilla_combobox.current(1)
        elif unidad_oculto == "pulg" and float(valorNominal_oculto) > 0.15:
            seleccionarPlantilla_combobox.current(1)

        
            # Movimiento de disco de bloques
        if unidad_oculto == "mm":
            if float(valorNominal_oculto) > 10:
                moverPlatoRemoto(RPi_url, 1)
            elif float(valorNominal_oculto) < 10:
                moverPlatoRemoto(RPi_url, 1)
        elif unidad_oculto == "pulg": 
            if float(valorNominal_oculto) > 0.15:
                moverPlatoRemoto(RPi_url, 1)
                


        return



    # Ocultar la ventana del menú de opciones una vez que se selecciona una opción
    root.withdraw()
    ventanaPrevia.destroy()

    # Crear una nueva ventana
    ventana_CalibracionAbierta = tk.Toplevel(root)
    ventana_CalibracionAbierta.title("Proceso de Calibración")
    ventana_CalibracionAbierta.configure(bg="white")
    ventana_CalibracionAbierta.protocol("WM_DELETE_WINDOW", lambda: regresarVentanaPrincipal(root, ventana_CalibracionAbierta)) #Cuando se cierre la ventana secundaria, vuelva al menú de opciones
    ventana_CalibracionAbierta.iconphoto(False, winIcono)

    # Crear un nuevo layout para la ventana de Nueva Calibración
    title_label = ttk.Label(ventana_CalibracionAbierta, text="Comparador de bloques TESA", font=("Helvetica", 16, "bold"),
                            background="white")
    title_label.grid(row=0, column=0, columnspan=2, pady=20)

    subtitle_label = ttk.Label(ventana_CalibracionAbierta, text="Proceso de Calibración", font=("Helvetica", 14),
                               background="white")
    subtitle_label.grid(row=9, column=0, columnspan=2, pady=10)

    image = Image.open("./assets/logoLCM.png")
    image = image.resize((int(image.width * 0.25), int(image.height * 0.25)))  # Ajustar el tamaño del logo
    image = ImageTk.PhotoImage(image)

    image_label = ttk.Label(ventana_CalibracionAbierta, image=image, background="white")
    image_label.image = image
    image_label.grid(row=0, column=10, rowspan=1, padx=10, pady=10)

    info_label = ttk.Label(ventana_CalibracionAbierta, 
        text= f"Información de la Calibración Actual:\n  · Nombre del Cliente: {cliente}\n  · Número de Certificado: {certificado}\n  · Identificación del Calibrando: {idCalibrando}\n  · Secuencia de Calibración: {secuencia}\n  · Patrón: {patron}\n", background="white"
    )
    info_label.grid(row=10, column=0, pady=20, padx=20)


# Valores por defecto
    tInicial_label = ttk.Label(ventana_CalibracionAbierta, text="Tiempo inicial (en minutos):", background="white")
    tInicial_label.grid(row=11, column=0, pady=10)
    tInicial_entry = ttk.Entry(ventana_CalibracionAbierta, width=42)
    tInicial_entry.insert(0, tInicial)
    tInicial_entry.grid(row=11, column=10, columnspan=1, pady=10, padx=10)

    tEstabilizacion_label = ttk.Label(ventana_CalibracionAbierta, text="Tiempo de estabilización (en segundos):",
                                      background="white")
    tEstabilizacion_label.grid(row=12, column=0, pady=10)
    tEstabilizacion_entry = ttk.Entry(ventana_CalibracionAbierta, width=42)
    tEstabilizacion_entry.insert(0, tEstabilizacion)
    tEstabilizacion_entry.grid(row=12, column=10, columnspan=1, pady=10, padx=10)

    numReps_label = ttk.Label(ventana_CalibracionAbierta, text="Número de repeticiones:", background="white")
    numReps_label.grid(row=13, column=0, pady=10)
    numReps_entry = ttk.Entry(ventana_CalibracionAbierta, width=42)
    numReps_entry.insert(0, numReps)
    numReps_entry.grid(row=13, column=10, columnspan=1, pady=10, padx=10)

# Parte para seleccionar el bloque y su plantilla

    seleccionarBloqueLabel = ttk.Label(ventana_CalibracionAbierta, text="Seleccione el bloque a calibrar (ID, Valor Nominal):", background="white")
    seleccionarBloqueLabel.grid(row=30, column=0, pady=10)

    bloquesCalibrando = [] #Lista para el registro de IDs y tamaños
    archivoCliente = BusquedaClientes(cliente)[2] #Busqueda del archivo del cliente
    workbookCliente = load_workbook(filename=archivoCliente)  #Apertura del archivo de excel del cliente
    hojaCalibrando = workbookCliente[idCalibrando]

    for i,fila in enumerate(hojaCalibrando.iter_rows(min_row=14, max_row=500, min_col=3, max_col=3), start=14):
        for celda in fila:
            if celda.value != None: #Ve si existe algún dato y adjunta
                bloquesCalibrando.append((celda.value, hojaCalibrando["B"+str(i)].value, unidad))
    workbookCliente.close()

    bloqueIdValor_combobox = ttk.Combobox(ventana_CalibracionAbierta, values=bloquesCalibrando, width=40,state= "readonly") # Se debe hacer split a la variable
    bloqueIdValor_combobox.grid(row=30, column=10, pady=10)
    bloqueIdValor_combobox.bind('<<ComboboxSelected>>', detect_plantilla)
    

    seleccionarPlantillaLabel =ttk.Label(ventana_CalibracionAbierta, text="Seleccione el tamaño de plantilla:", background="white")
    seleccionarPlantillaLabel.grid(row=35, column=0, pady=10, padx=10)

    seleccionarPlantilla_combobox = ttk.Combobox(ventana_CalibracionAbierta, values=["Pequeña", "Grande"], width=40 ,state= "readonly")
    seleccionarPlantilla_combobox.grid(row=35, column=10, pady=10, padx=10)
    #--

    #Mover Plantilla
    recordatorio_label = ttk.Label(ventana_CalibracionAbierta, text= "Recuerde colocar la Plantilla en Posición 1 ", background="white" )
    recordatorio_label.grid(row=40, column=0, pady=20, padx=10)

    moverDe0a1_button = ttk.Button(ventana_CalibracionAbierta, text="Mover del Centro a Posición 1", 
                                 command=lambda: moverDe0a1(RPi_url))
    moverDe0a1_button.grid(row=40, column=10, columnspan=1, pady=10, padx=10)

    #Mover Plato
    moverPlato_label = ttk.Label(ventana_CalibracionAbierta, text= "Presione el botón para mover el plato giratorio ", background="white" )
    moverPlato_label.grid(row=50, column=0, pady=20, padx=10)

    moverPlato_button = ttk.Button(ventana_CalibracionAbierta, text="Mover Plato", 
                                 command=lambda: moverPlato())
    moverPlato_button.grid(row=50, column=10, columnspan=1, pady=10, padx=10)
    
    # Continuar
    continuar_button = ttk.Button(ventana_CalibracionAbierta, text="⮩ Realizar Calibración",  command=lambda: calibrarBloque(archivoCalibracion_datos, secuencia, bloqueIdValor_combobox,seleccionarPlantilla_combobox, tInicial_entry, tEstabilizacion_entry, numReps_entry))
    continuar_button.grid(row=80, column=0, columnspan=1, pady=10, padx=10)

    regresar_button = ttk.Button(ventana_CalibracionAbierta, text="⮪ Pausar Calibración y Regresar al Menú", 
                                 command=lambda: regresarVentanaPrincipal(root, ventana_CalibracionAbierta))
    regresar_button.grid(row=100, column=0, columnspan=1, pady=10)
    
    finalizar_button = ttk.Button(ventana_CalibracionAbierta, text="🗸 Finalizar y Guardar Calibración",  command=lambda: finalizarCalibracion(root, ventana_CalibracionAbierta, archivoCalibracion_datos, cliente, certificado, secuencia, numReps) )
    finalizar_button.grid(row=100, column=10, columnspan=1, pady=10)
    return









def ingresar_cliente():
    # Ocultar la ventana del menú de opciones una vez que se selecciona una opción
    root.withdraw()

    # Crear una nueva ventana
    ventana_cliente = tk.Toplevel(root)
    ventana_cliente.title("Ingresar Cliente")
    ventana_cliente.configure(bg="white")
    ventana_cliente.protocol("WM_DELETE_WINDOW", lambda: regresarVentanaPrincipal(root, ventana_cliente)) #Cuando se cierre la ventana secundaria, vuelva al menú de opciones
    ventana_cliente.iconphoto(False, winIcono)

    # Crear un nuevo layout para la ventana de ingresar cliente
    title_label = ttk.Label(ventana_cliente, text="Comparador de bloques TESA", font=("Helvetica", 16, "bold"),
                            background="white")
    title_label.grid(row=0, column=0, columnspan=2, pady=20)

    subtitle_label = ttk.Label(ventana_cliente, text="Ingresar nuevo cliente", font=("Helvetica", 14),
                               background="white")
    subtitle_label.grid(row=1, column=0, columnspan=2, pady=10)

    image = Image.open("./assets/logoLCM.png")
    image = image.resize((int(image.width * 0.25), int(image.height * 0.25)))  # Ajustar el tamaño del logo
    image = ImageTk.PhotoImage(image)

    image_label = ttk.Label(ventana_cliente, image=image, background="white")
    image_label.image = image
    image_label.grid(row=0, column=2, rowspan=1, padx=10, pady=10)

    global nuevoCliente_entry, contactoCliente_entry

    # Espacios para ingresar las variables requeridas para reanudar calibración
    nuevoCliente_label = ttk.Label(ventana_cliente, text="Nombre del nuevo cliente:", background="white")
    nuevoCliente_label.grid(row=2, column=0, pady=5)
    nuevoCliente_entry = ttk.Entry(ventana_cliente, width=42)
    nuevoCliente_entry.grid(row=2, column=1, columnspan=2, pady=5, padx=(20, 5))

    contactoCliente_label = ttk.Label(ventana_cliente, text="Dirección del Cliente:",
                                      background="white")
    contactoCliente_label.grid(row=3, column=0, pady=5)
    contactoCliente_entry = ttk.Entry(ventana_cliente, width=42)
    contactoCliente_entry.grid(row=3, column=1, columnspan=2, pady=5, padx=(20, 5))

    #Botones
    continuar_button = ttk.Button(ventana_cliente, text="Ingresar cliente", command=ingresarCliente)
    continuar_button.grid(row=4, column=0, columnspan=1, pady=10, padx=10)
    
    clientes_actuales_button = ttk.Button(ventana_cliente, text="Ver clientes actuales", command=verClientes)
    clientes_actuales_button.grid(row=4, column=1, columnspan=1, pady=10, padx=10)

    regresar_button = ttk.Button(ventana_cliente, text="Regresar al menú de opciones",
                                 command=lambda: regresarVentanaPrincipal(root, ventana_cliente))
    regresar_button.grid(row=5, column=0, columnspan=1, pady=10)
    return


def ingresar_calibrando():
    # Ocultar la ventana del menú de opciones una vez que se selecciona una opción
    root.withdraw()

    # Crear una nueva ventana
    global ventana_calibrando
    ventana_calibrando = tk.Toplevel(root)
    ventana_calibrando.title("Ingresar Cliente")
    ventana_calibrando.configure(bg="white")
    ventana_calibrando.protocol("WM_DELETE_WINDOW", lambda: regresarVentanaPrincipal(root, ventana_calibrando)) #Cuando se cierre la ventana secundaria, vuelva al menú de opciones
    ventana_calibrando.iconphoto(False, winIcono)

    # Crear un nuevo layout para la ventana para ingresar un calibrando
    title_label = ttk.Label(ventana_calibrando, text="Comparador de bloques TESA", font=("Helvetica", 16, "bold"),
                            background="white")
    title_label.grid(row=0, column=0, columnspan=2, pady=20)

    subtitle_label = ttk.Label(ventana_calibrando, text="Ingresar calibrando", font=("Helvetica", 14),
                               background="white")
    subtitle_label.grid(row=1, column=0, columnspan=2, pady=10)

    image = Image.open("./assets/logoLCM.png")
    image = image.resize((int(image.width * 0.25), int(image.height * 0.25)))  # Ajustar el tamaño del logo
    image = ImageTk.PhotoImage(image)

    image_label = ttk.Label(ventana_calibrando, image=image, background="white")
    image_label.image = image
    image_label.grid(row=0, column=2, rowspan=1, padx=10, pady=10)

    global cliente_combobox, objeto_entry, marca_entry, numSerie_entry, materialCalibrando_combobox, cantidad_entry, identificacionInterna_entry
    global modelo_entry, grado_entry, unidad_combobox

    # Crear una lista con los nombres de los clientes ya registrados
    clientesRegistrados = []
    archivoClientes = openpyxl.load_workbook("Clientes/Clientes.xlsx")
    hojaClientes = archivoClientes.active

    numFila = 3  # Se empieza en la fila 3 porque antes están los encabezados
    for fila in hojaClientes.iter_rows(min_row=3,
                                       min_col=1,
                                       max_col=1):
        for celda in fila:
            if celda.value != None:
                clientesRegistrados.append(celda.value)
    archivoClientes.close()

    # Espacios para ingresar las variables requeridas para una nueva calibración
    cliente_label = ttk.Label(ventana_calibrando, text="Nombre del cliente:", anchor=tk.CENTER, background="white")
    cliente_label.grid(row=2, column=0, pady=5, sticky=tk.EW)
    cliente_combobox = ttk.Combobox(ventana_calibrando, values=clientesRegistrados, width=40 ,state= "readonly")
    cliente_combobox.grid(row=2, column=1, columnspan=2, pady=5, padx=(20, 5), sticky="ew")

    objeto_label = ttk.Label(ventana_calibrando, text="Objeto a calibrar:", background="white")
    objeto_label.grid(row=3, column=0, pady=5)
    objeto_entry = ttk.Entry(ventana_calibrando, width=42)
    objeto_entry.grid(row=3, column=1, columnspan=2, pady=5, padx=(20, 5))

    cantidad_label = ttk.Label(ventana_calibrando, text="Cantidad de bloques o instrumentos:", background="white")
    cantidad_label.grid(row=4, column=0, pady=5)
    cantidad_entry = ttk.Entry(ventana_calibrando, width=42)
    cantidad_entry.grid(row=4, column=1, columnspan=2, pady=5, padx=(20, 5))

    marca_label = ttk.Label(ventana_calibrando, text="Marca:", background="white")
    marca_label.grid(row=5, column=0, pady=5)
    marca_entry = ttk.Entry(ventana_calibrando, width=42)
    marca_entry.grid(row=5, column=1, columnspan=2, pady=5, padx=(20, 5))

    numSerie_label = ttk.Label(ventana_calibrando, text="Número de serie:", background="white")
    numSerie_label.grid(row=6, column=0, pady=5)
    numSerie_entry = ttk.Entry(ventana_calibrando, width=42)
    numSerie_entry.grid(row=6, column=1, columnspan=2, pady=5, padx=(20, 5))

    materialCalibrando_label = ttk.Label(ventana_calibrando, text="Material:", anchor=tk.CENTER, background="white")
    materialCalibrando_label.grid(row=7, column=0, pady=5, sticky=tk.EW)
    materialCalibrando_combobox = ttk.Combobox(ventana_calibrando, values=["Acero inoxidable", "Cerámica"], width=40 ,state= "readonly")
    materialCalibrando_combobox.grid(row=7, column=1, columnspan=2, pady=5, padx=(20, 5), sticky="ew")

    modelo_label = ttk.Label(ventana_calibrando, text="Modelo:", anchor=tk.CENTER, background="white")
    modelo_label.grid(row=8, column=0, pady=5, sticky=tk.EW)
    modelo_entry = ttk.Entry(ventana_calibrando, width=42)
    modelo_entry.grid(row=8, column=1, columnspan=2, pady=5, padx=(20, 5), sticky="ew")

    grado_label = ttk.Label(ventana_calibrando, text="Grado declarado: ", anchor=tk.CENTER, background="white")
    grado_label.grid(row=9, column=0, pady=5, sticky=tk.EW)
    grado_entry = ttk.Entry(ventana_calibrando, width=42)
    grado_entry.grid(row=9, column=1, columnspan=2, pady=5, padx=(20, 5), sticky="ew")

    identificacionInterna_label = ttk.Label(ventana_calibrando, text="Identificacion Interna: ", anchor=tk.CENTER, background="white")
    identificacionInterna_label.grid(row=10, column=0, pady=5, sticky=tk.EW)
    identificacionInterna_entry = ttk.Entry(ventana_calibrando, width=42)
    identificacionInterna_entry.grid(row=10, column=1, columnspan=2, pady=5, padx=(20, 5), sticky="ew")

    unidad_label = ttk.Label(ventana_calibrando, text="Unidad:", anchor=tk.CENTER, background="white")
    unidad_label.grid(row=11, column=0, pady=5, sticky=tk.EW)
    unidad_combobox = ttk.Combobox(ventana_calibrando, values=["mm", "pulg"], width=40 ,state= "readonly")
    unidad_combobox.grid(row=11, column=1, columnspan=2, pady=5, padx=(20, 5), sticky="ew")

    continuar_button = ttk.Button(ventana_calibrando, text="Ingresar calibrando", command=ingresarCalibrando)
    continuar_button.grid(row=12, column=1, columnspan=1, pady=10)

    regresar_button = ttk.Button(ventana_calibrando, text="Regresar al menú de opciones",
                                 command=lambda: regresarVentanaPrincipal(root, ventana_calibrando))
    regresar_button.grid(row=12, column=0, columnspan=1, pady=10)
    return


def ingresar_bloque_calibrando(ventana, cliente, numSerie, identificacionInterna, unidad):
    global ultimoBloqueDatos_label
    ventana.withdraw()
    
    top = tk.Toplevel()
    top.title("Información Bloque Individual")
    top.configure(bg="white")
    top.protocol("WM_DELETE_WINDOW", lambda: regresarVentanaPrincipal(root, top)) #Cuando se cierre la ventana secundaria, vuelva al menú de opciones
    top.iconphoto(False, winIcono)

    subtitle_label = ttk.Label(top, text="Información Bloques", font=("Helvetica", 14), background="white")
    subtitle_label.grid(row=0, column=0, columnspan=2, pady=10)


    cliente_label = ttk.Label(top, text="Nombre del cliente:", background="white")
    cliente_label.grid(row=4, column=0, pady=5)
    cliente_label = ttk.Label(top, text=cliente, background="white")
    cliente_label.grid(row=4, column=10, pady=5)

    idCalibrando_label = ttk.Label(top, text="ID Interno:", background="white")
    idCalibrando_label.grid(row=6, column=0, pady=5)
    idCalibrando_label = ttk.Label(top, text=identificacionInterna, background="white")
    idCalibrando_label.grid(row=6, column=10, pady=5)

    serieCalibrando_label = ttk.Label(top, text="Serie del calibrando:", background="white")
    serieCalibrando_label.grid(row=7, column=0, pady=5)
    serieCalibrando_label = ttk.Label(top, text=numSerie, background="white")
    serieCalibrando_label.grid(row=7, column=10, pady=5)

    subtituloEntrada_label = ttk.Label(top, text="Entrada de Bloque", font=("Helvetica", 12), background="white")
    subtituloEntrada_label.grid(row=9, column=0, columnspan=1, pady=10)

    longitudNominal_label = ttk.Label(top, text="Longitud Nominal", background="white")
    longitudNominal_label.grid(row=10, column=0, columnspan=1, pady=10)

    longitudNominal_entry = ttk.Entry(top, width=30)
    longitudNominal_entry.grid(row=10, column=10, columnspan=1, pady=10, padx=10)        
    
    idBloque_label = ttk.Label(top, text="ID del Bloque", background="white")
    idBloque_label.grid(row=20, column=0, columnspan=1, pady=10)

    idBloque_entry = ttk.Entry(top, width=30)
    idBloque_entry.grid(row=20, column=10, columnspan=1, pady=10, padx=10) 

    ultimoBloque_label = ttk.Label(top, text="Último Bloque Agregado:", background="white")
    ultimoBloque_label.grid(row=40, column=0, columnspan=1, pady=10) 
    ultimoBloqueDatos_label = ttk.Label(top, text="", background="white")
    ultimoBloqueDatos_label.grid(row=40, column=10, columnspan=1, pady=10)

    ingresarBloque_button = ttk.Button(top, text="Ingresar Bloque", command=lambda: ingresarBloque(top, cliente, numSerie, unidad, longitudNominal_entry, idBloque_entry, ultimoBloqueDatos_label))
    ingresarBloque_button.grid(row=30, column=10, columnspan=1, pady=10)

    finalizar_button = ttk.Button(top, text="Finalizar Adiciones", command=lambda: regresarVentanaPrincipal(root, top))
    finalizar_button.grid(row=50, column=10, columnspan=1, pady=10)

    return



def mover_motores():
    # Crear una nueva ventana
    ventana_moverMotores = tk.Toplevel(root)
    ventana_moverMotores.title("Posicionamiento de Motores")
    ventana_moverMotores.configure(bg="white")
    ventana_moverMotores.focus_set()

    main_label = ttk.Label(ventana_moverMotores,
                           text="Utilize las flechas del teclado para colocar los motores en la posición inicial.",
                           anchor=tk.CENTER, background="white")
    main_label.grid(row=0, column=0, pady=(30, 0), padx=30)
    flechas_label = ttk.Label(ventana_moverMotores, text="←↕→", anchor=tk.CENTER, background="white")
    flechas_label.grid(row=1, column=0, pady=(10, 10))
    exit_label = ttk.Label(ventana_moverMotores, text="Presione Enter ↲ para salir.", anchor=tk.CENTER,
                           background="white")
    exit_label.grid(row=2, column=0, pady=(0, 50), padx=30)

    print("Preparacion Pedal")
    ActivaPedal()
    GPIO.output(pin_enableCalibrationMotor, motorEnabledState)  # Encender motores

    def funcionMotores(event):
        print(event.keysym)
        print(type(event.keysym))
        moverManualInterfaz(event)

    def muere(event):
        print("Terminacion Pedal")
        ActivaPedal()
        GPIO.output(pin_enableCalibrationMotor, motorDisabledState)  # Apagar motores
        ventana_moverMotores.destroy()

    ventana_moverMotores.bind("<Up>", funcionMotores)
    ventana_moverMotores.bind("<Down>", funcionMotores)
    ventana_moverMotores.bind("<Left>", funcionMotores)
    ventana_moverMotores.bind("<Right>", funcionMotores)
    ventana_moverMotores.bind("<Return>", muere)
    ventana_moverMotores.bind("<q>", muere)

    return










########## Funciones de la Interfaz

def continuarNuevaCalibracion(ventana): # Función para continuar con el proceso de una calibración, dados todos los datos.
    cliente = cliente_combobox.get()
    certificado = certificado_entry.get()
    solicitud = solicitud_entry.get()
    idCalibrando = idCalibrando_entry.get()
    responsable = responsable_entry.get()
    revision = revision_entry.get()
    patron = patron_combobox.get()
    material = material_combobox.get()
    secuencia = secuencia_combobox.get()
    tInicial = tInicial_entry.get()
    tEstabilizacion = tEstabilizacion_entry.get()
    numReps = numReps_entry.get()

    #Verificar si existe idCalibrando y registrar unidades
    archivoCliente = BusquedaClientes(cliente)[2]
    workbookCliente = load_workbook(filename=archivoCliente)  #Apertura del archivo de excel del cliente

    #Revisar si ya existe algún calibrando registrado con el mismo número de serie
    existeCalibrando = False
    if idCalibrando in workbookCliente.sheetnames:
        existeCalibrando = True

    if not existeCalibrando:
        mostrarMensaje("El ID de calibrando no existe. \nPor favor registrar calibrando.")
        return
    
    hojaCalibrando = workbookCliente[idCalibrando]

    unidad = hojaCalibrando["Z1"].value


    # Verificar si existe archivo
    nombreArchivoCalibracion = "./Calibraciones en curso/" + str(certificado) + "_Info.xlsx"
    if os.path.exists(nombreArchivoCalibracion):
        mostrarMensaje("El archivo de calibración ya existe. \nPor favor REANUDAR la calibración.")
        return
    
    else:
        archivoCalibracion_datos, archivoCalibracion_info = CrearArchivoCalibracion(certificado)

        RellenarInfoCalibracion(archivoCalibracion_info, [cliente, certificado, solicitud, idCalibrando, responsable, revision, patron, material, secuencia, tInicial, tEstabilizacion, numReps, unidad])

    calibracion_abierta(ventana, archivoCalibracion_datos, cliente, certificado, solicitud, idCalibrando, responsable, revision, patron, material, secuencia, tInicial, tEstabilizacion, numReps, unidad)

    return 


def reanudarCalibracion(ventana):
    certificado = certificado_combobox.get()
    
    archivoCalibracion_datos, cliente, certificado, solicitud, idCalibrando, responsable, revision, patron, material, secuencia, tInicial, tEstabilizacion, numReps, unidad = obtenerInfoCalibracion(certificado)

    calibracion_abierta(ventana, archivoCalibracion_datos, cliente, certificado, solicitud, idCalibrando, responsable, revision, patron, material, secuencia, tInicial, tEstabilizacion, numReps, unidad)
    return


def calibrarBloque(archivoCalibracion_datos, secuencia, bloqueIdValor_combobox, seleccionarPlantilla_combobox, tInicial_entry, tEstabilizacion_entry, numReps_entry ):
    bloqueIdValor = bloqueIdValor_combobox.get()
    bloqueID, valorNominal, unidad = bloqueIdValor.split()
    plantilla = seleccionarPlantilla_combobox.get()
    tInicial = tInicial_entry.get()
    tEstabilizacion = tEstabilizacion_entry.get()
    numReps = numReps_entry.get()

    mensajeProceso = mostrarMensaje(f"Calibración del Bloque {bloqueID} en Proceso.\nPor favor espere a que finalice.")
    mensajeProceso.update()
    procesoCalibracion(RPi_url, archivoCalibracion_datos, secuencia, bloqueID, valorNominal, tInicial, tEstabilizacion, numReps, plantilla)
    mensajeProceso.destroy()
    mostrarMensaje(f"Calibración del Bloque {bloqueID} Finalizada.")
    return

def moverPlato():
    ventana = tk.Toplevel(root)
    ventana.title("Mover Plato")
    ventana.configure(bg="white")
    ventana.iconphoto(False, winIcono)

    title_label = ttk.Label(ventana, text="Seleccione la posición deseada", font=("Helvetica", 12, "bold"),
                            background="white")
    title_label.grid(row=0, column=0, columnspan=2, pady=10, padx=20)

    pos1 = ttk.Button(ventana, text="1", width=20,
                                 command=lambda: moverPlatoRemoto(RPi_url, "1"))
    pos1.grid(row=1, column=0, columnspan=1, pady=10)
    
    pos2 = ttk.Button(ventana, text="2", width=20,
                                 command=lambda: moverPlatoRemoto(RPi_url, "2"))
    pos2.grid(row=1, column=1, columnspan=1, pady=10)
    
    pos3 = ttk.Button(ventana, text="3", width=20,
                                 command=lambda: moverPlatoRemoto(RPi_url, "3"))
    pos3.grid(row=2, column=0, columnspan=1, pady=10)

    pos4 = ttk.Button(ventana, text="4", width=20,
                                 command=lambda: moverPlatoRemoto(RPi_url, "4"))
    pos4.grid(row=2, column=1, columnspan=1, pady=10)
    
    return



def finalizarCalibracion(root, ventana_CalibracionAbierta, archivoCalibracion_datos, cliente, certificado, secuencia, numReps):
    
    
    RellenarEncabezados(archivoCalibracion_datos, secuencia, numReps)

    # Combinar archivos
    archivoCalibracion_info=f"./Calibraciones en curso/{certificado}_Info.xlsx"

    workbookCombinado = load_workbook(filename=archivoCalibracion_datos)
    
    hojaDatosComb = workbookCombinado.active
    hojaDatosComb.title = "Datos"
    
    workbookInfo = load_workbook(filename=archivoCalibracion_info)
    hojaInfo = workbookInfo.active

    workbookCombinado.create_sheet("Información")
    hojaInfoComb = workbookCombinado["Información"]
    
    for row in hojaInfo.iter_rows(values_only=False):
        for cell in row:
            new_cell = hojaInfoComb.cell(row=cell.row, column=cell.column, value=cell.value)

            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.alignment = copy(cell.alignment)


    workbookCombinado.save(f"./Calibraciones Finalizadas/{certificado}-{cliente}.xlsx")
    workbookCombinado.close()
    workbookInfo.close()

    # Eliminar Archivos Restantes
    EliminarArchivo(archivoCalibracion_info)

    regresarVentanaPrincipal(root, ventana_CalibracionAbierta)
    return




    


def ingresarCliente():
    nuevoCliente = nuevoCliente_entry.get()
    contactoCliente = contactoCliente_entry.get()
    AgregarCliente(nuevoCliente, contactoCliente)
    return

def verClientes():
    # Crear una lista con los nombres de los clientes ya registrados
    clientesRegistrados = []
    archivoClientes = openpyxl.load_workbook("Clientes/Clientes.xlsx")
    hojaClientes = archivoClientes.active

    numFila = 3  # Se empieza en la fila 3 porque antes están los encabezados
    for fila in hojaClientes.iter_rows(min_row=3,
                                       min_col=1,
                                       max_col=1):
        for celda in fila:
            if celda.value != None:
                clientesRegistrados.append(celda.value)
    archivoClientes.close()

    textoClientes = "Clientes Registrados Actualmente:\n\n"

    for nombre in clientesRegistrados:
        textoClientes += "  · " + nombre + "\n"

    mostrarMensaje(textoClientes)
    return


def ingresarCalibrando():
    global isOK
    isOK = 1

    cliente = cliente_combobox.get()
    objeto = objeto_entry.get()
    cantidad = cantidad_entry.get()
    marca = marca_entry.get()
    numSerie = numSerie_entry.get()
    materialCalibrando = materialCalibrando_combobox.get()
    modelo = modelo_entry.get()
    grado = grado_entry.get()
    identificacionInterna = identificacionInterna_entry.get()
    unidad = unidad_combobox.get()
    IngresarCalibrando(cliente, objeto, cantidad, marca, numSerie, materialCalibrando, modelo, grado, identificacionInterna, unidad)
    
    if isOK:
        ingresar_bloque_calibrando(ventana_calibrando, cliente, numSerie, identificacionInterna, unidad)

    return



def regresarVentanaPrincipal(root, ventana):
    ventana.destroy()  # Destruir la ventana actual
    root.deiconify()  # Traer devuelta la ventana principal


# Status del RPi
def checkRPiStatus(url):
    try:
        url = url+"isUp"
        response = requests.get(url, timeout=5)
        data = response.json()
        status = data.get("status", "desconocido")
        status_label["text"] = f"Estado del Servidor: {status}"
        status_label["foreground"] = "green"
    except:
        status_label["text"] = "Estado del Servidor: error"
        status_label["foreground"] = "red"
    return

################## Ventana inicial ##################

root = tk.Tk()
#themed_style = ThemedStyle(root)
#themed_style.set_theme("adapta")  

global winIcono
winIcono = tk.PhotoImage(file = "./assets/logoLCM_r.png")
root.iconphoto(False, winIcono)


root.title("Comparador de bloques TESA")
root.configure(bg="white")

title_label = ttk.Label(root, text="Comparador de bloques TESA", font=("Helvetica", 16, "bold"), background="white")
title_label.grid(row=0, column=0, columnspan=2, pady=20, padx=20)

subtitle_label = ttk.Label(root, text="Menú de opciones", font=("Helvetica", 14), background="white")
subtitle_label.grid(row=1, column=0, columnspan=2, pady=10)

image = Image.open("./assets/logoLCM.png")
image = image.resize((int(image.width * 0.25), int(image.height * 0.25)))
image = ImageTk.PhotoImage(image)

image_label = ttk.Label(root, image=image, background="white")
image_label.image = image
image_label.grid(row=0, column=2, rowspan=1, padx=10, pady=10)

options = [
    ("Nueva calibración", nueva_calibracion),
    ("Reanudar calibración", reanudar_calibracion),
    ("Ingresar cliente", ingresar_cliente),
    ("Ingresar calibrando", ingresar_calibrando)
]

for i, (text, command) in enumerate(options):
    button = ttk.Button(root, text=text, command=command)
    button.grid(row=i + 2, column=0, columnspan=2, pady=5, padx=10, sticky="we")


status_label = ttk.Label(root, text="Estado del Servidor: desconocido", background="white", foreground="black")
status_label.grid(row=10, column=0, sticky=tk.W, pady=(10, 10), padx=(20,5))


check_button = ttk.Button(root, text="↻", command=lambda:checkRPiStatus(RPi_url))
check_button.grid(row=10, column=1, sticky=tk.W)


root.mainloop()